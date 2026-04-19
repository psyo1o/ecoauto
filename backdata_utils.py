# -*- coding: utf-8 -*-
"""
백데이터(수분량·THC) 추출 유틸
- eco_input.py에서 분리
- 수분 CSV, THC CSV/FID 저장 로직
- 클립보드 유틸(win32clipboard) 포함
"""
import os
import time
from datetime import datetime
import win32clipboard
from openpyxl import load_workbook
from excel_com_utils import get_excel_app
from data_utils import sample_to_datestr

# =====================================================================
# 상수
# =====================================================================
MOIST_ROOT = r"\\192.168.10.163\측정팀\2.성적서\0.수분량"
THC_ROOT   = r"\\192.168.10.163\측정팀\2.성적서\0.THC"

# Excel COM 상수
XL_UP     = -4162
XL_TOLEFT = -4159


# =====================================================================
# 클립보드 유틸
# =====================================================================
def _clipboard_clear():
    """클립보드 충돌 방지를 위한 재시도 로직 적용"""
    for _ in range(4):
        try:
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
            return
        except:
            try: win32clipboard.CloseClipboard()
            except: pass
            time.sleep(0.15)

def _clipboard_get_unicode_text():
    """클립보드 읽기 재시도 로직 적용"""
    for _ in range(4):
        try:
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                    return win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_TEXT):
                    b = win32clipboard.GetClipboardData(win32clipboard.CF_TEXT)
                    return b.decode("cp949", errors="ignore")
                return ""
            finally:
                win32clipboard.CloseClipboard()
        except:
            time.sleep(0.15)
    return ""

def _copy_range_and_get_text(excel_app, rng, wait_max=3.0, retries=4):
    """✅ Range.Copy() → 클립보드 텍스트를 폴링해서 가져옴(재시도/에러 무시 로직 대폭 강화)"""
    for attempt in range(retries):
        try: excel_app.CutCopyMode = False
        except: pass
        _clipboard_clear()

        # 복사 시도
        try:
            rng.Copy()
        except:
            time.sleep(0.3)
            try: rng.Copy()
            except: pass

        # 클립보드 폴링
        end = time.time() + float(wait_max)
        txt = ""
        while time.time() < end:
            time.sleep(0.1)
            txt = _clipboard_get_unicode_text()
            if txt and txt.strip():
                break

        # 성공했으면 복사모드 해제 후 리턴
        if txt and txt.strip():
            try: excel_app.CutCopyMode = False
            except: pass
            _clipboard_clear()
            return txt

        # 실패 시 약간 대기 후 다음 시도(attempt)
        time.sleep(0.5)

    try: excel_app.CutCopyMode = False
    except: pass
    _clipboard_clear()
    return ""


# =====================================================================
# 파일 저장 안전망
# =====================================================================
def _safe_write_file(path, text, encoding="utf-8-sig", newline="", retries=4):
    """네트워크 드라이브(NAS) 접근 지연 및 권한 오류 방지용 안전 쓰기"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    for attempt in range(retries):
        try:
            with open(path, "w", encoding=encoding, newline=newline) as f:
                f.write(text)
            return True
        except Exception as e:
            if attempt == retries - 1:
                raise RuntimeError(f"파일 저장 4회 시도 실패 ({path}): {e}")
            time.sleep(0.5)  # NAS 지연 대기 후 재시도
    return False

def _year_month_folder_from_sample(sample_no: str):
    """
    sample_no: AYYMMDDT-XX 에서
    YYYY, 'M월' 폴더명 반환
    """
    ds = sample_to_datestr(sample_no)
    if not ds:
        return None, None
    yyyy, mm, _ = ds.split("-")
    mm_int = int(mm)
    return yyyy, f"{mm_int}월"


# =====================================================================
# 수분 CSV(표시값 그대로)
# =====================================================================
def _export_moist_csv_from_open_ws(excel_app, ws, out_csv_path: str, max_rows=None):
    """
    ✅ '열린 ws(엑셀 COM)'에서 표시값을 복사해서 CSV로 저장
    - 날짜/시간/소수점 등 표시 서식 유지
    - 탭 → 콤마
    """
    os.makedirs(os.path.dirname(out_csv_path), exist_ok=True)

    # 마지막 행/열(값 기준, 빠름)
    last_row = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
    last_col = ws.Cells(1, ws.Columns.Count).End(XL_TOLEFT).Column

    if max_rows is not None:
        last_row = min(int(max_rows), int(last_row))

    if last_row < 1:
        last_row = 1
    if last_col < 1:
        last_col = 1

    rng = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
    txt = _copy_range_and_get_text(excel_app, rng, wait_max=3.0)

    if not txt or not txt.strip():
        raise RuntimeError("클립보드에서 수분 데이터를 복사해오지 못했습니다. (엑셀 응답 없음)")
        
    # 탭 → 콤마
    txt = txt.replace("\t", ",")
    # 줄바꿈 정리(윈도우 CRLF 유지)
    txt = txt.replace("\r\n", "\n").replace("\r", "\n")
    txt = "\r\n".join(txt.split("\n"))

    _safe_write_file(out_csv_path, txt, encoding="utf-8-sig", newline="")

    return out_csv_path


def export_csv_display_as_is(excel_path: str, sheet_name: str, out_csv_path: str, max_rows=None):
    """
    ✅ (외부에서도 쓸 수 있게 유지) 엑셀 파일 열어서 표시값 CSV 저장
    - 내부적으로는 Excel App 재사용
    """
    excel = get_excel_app()
    wb = None
    try:
        wb = excel.Workbooks.Open(excel_path, ReadOnly=True, UpdateLinks=0)
        ws = wb.Worksheets(sheet_name)
        return _export_moist_csv_from_open_ws(excel, ws, out_csv_path, max_rows=max_rows)
    finally:
        try:
            if wb is not None:
                try:
                    excel.CutCopyMode = False
                except:
                    pass
                wb.Close(False)
        except:
            pass


# =====================================================================
# THC PF → FID(150행 고정, 복사 기반)
# =====================================================================
def _export_pf_fid_from_open_ws(excel_app, ws, out_fid_path: str, fixed_rows=150):
    """
    ✅ 열린 ws(엑셀 COM)에서 150행 고정으로 복사한 텍스트를 .FID로 저장
    """
    os.makedirs(os.path.dirname(out_fid_path), exist_ok=True)

    last_col = ws.UsedRange.Columns.Count
    if not last_col or last_col < 1:
        last_col = 1

    rng = ws.Range(ws.Cells(1, 1), ws.Cells(int(fixed_rows), int(last_col)))
    txt = _copy_range_and_get_text(excel_app, rng, wait_max=3.0)

    if not txt or not txt.strip():
        raise RuntimeError("클립보드에서 THC(PF) 데이터를 복사해오지 못했습니다.")

    # 안전 저장 유틸 사용
    _safe_write_file(out_fid_path, txt, encoding="utf-8-sig", newline="")

    return out_fid_path


def export_fid_by_excel_copy(excel_path: str, sheet_name: str, out_fid_path: str):
    """
    ✅ (기존 시그니처 유지) 엑셀 열어서 PF 시트 복사→FID 저장
    - 내부적으로 Excel App 재사용
    """
    excel = get_excel_app()
    wb = None
    try:
        wb = excel.Workbooks.Open(excel_path, ReadOnly=True, UpdateLinks=0)
        ws = wb.Worksheets(sheet_name)
        return _export_pf_fid_from_open_ws(excel, ws, out_fid_path, fixed_rows=150)
    finally:
        try:
            if wb is not None:
                try:
                    excel.CutCopyMode = False
                except:
                    pass
                wb.Close(False)
        except:
            pass


# =====================================================================
# THC CSV(openpyxl)
# =====================================================================
def _safe_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v)


def _used_range_bounds(ws):
    max_row = 1
    max_col = 1
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                if cell.row > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column
    return max_row, max_col


def _export_ws_to_csv(ws, out_csv_path, encoding="utf-8-sig"):
    last_r, last_c = _used_range_bounds(ws)
    os.makedirs(os.path.dirname(out_csv_path), exist_ok=True)

    def q(s):
        s = _safe_str(s)
        if any(ch in s for ch in [",", '"', "\n", "\r"]):
            s = s.replace('"', '""')
            return f'"{s}"'
        return s

    lines = []
    for r in range(1, last_r + 1):
        row_vals = [q(ws.cell(r, c).value) for c in range(1, last_c + 1)]
        while row_vals and row_vals[-1] == "":
            row_vals.pop()
        lines.append(",".join(row_vals).rstrip())

    while lines and lines[-1].strip() == "":
        lines.pop()

    text_data = "\n".join(lines)
    _safe_write_file(out_csv_path, text_data, encoding=encoding, newline="\n")


# =====================================================================
# export_backdata_moist_thc (최적화 적용)
# =====================================================================
def export_backdata_moist_thc(excel_path: str, sample_no: str):
    """
    ✅ 규칙 그대로 + 최적화:
    - openpyxl로 조건 판단
    - 수분/THC PF(FID) 둘 중 하나라도 필요하면
      Excel 파일을 "한 번만 Open → 필요한 시트들 처리 → Close"
    - Excel 앱은 전역 1개 재사용(프로그램 끝에서 close_excel_app()로 닫기)
    """
    # ✅ 비산먼지 스킵
    if ("비산먼지" in os.path.basename(excel_path)) or ("비산" in os.path.basename(excel_path)):
        print("▶ 비산먼지 성적서 → 백데이터 스킵")
        return

    yyyy, mm_folder = _year_month_folder_from_sample(sample_no)
    if not yyyy:
        print(f"⚠ 백데이터: 시료번호 날짜 파싱 실패 → {sample_no}")
        return
    print(f"▶백데이터: 시료번호 {sample_no} 진행중")
    # ---- 1) openpyxl로 조건 판단(빠름) ----
    wb = load_workbook(excel_path, data_only=True)
    
    # 수분 조건
    need_moist = False
    try:
        ws_in = wb["입력"] if "입력" in wb.sheetnames else None
        if ws_in is not None:
            b18 = "" if ws_in["B18"].value is None else str(ws_in["B18"].value).strip()
            need_moist = (b18 == "사용") and ("수분량자동측정" in wb.sheetnames)
    except:
        need_moist = False

    # THC 조건
    ws_an = None
    if "입력(분석값)" in wb.sheetnames:
        ws_an = wb["입력(분석값)"]
    else:
        for n in wb.sheetnames:
            if n.replace(" ", "") == "입력(분석값)".replace(" ", ""):
                ws_an = wb[n]
                break

    need_thc = False
    is_fid_mode = False
    if ws_an is not None:
        a64 = "" if ws_an["A64"].value is None else str(ws_an["A64"].value).strip()
        if a64 != "":
            need_thc = True
            c65 = ws_an["C65"].value
            try:
                is_fid_mode = (int(float(c65)) == 1)
            except:
                is_fid_mode = (str(c65).strip() == "1")

    # ---- 2) 엑셀 COM은 "필요할 때만" 파일 1회 Open/Close ----
    excel = None
    wb_xl = None
    try:
        if need_moist or (need_thc and not is_fid_mode):
            excel = get_excel_app()
            wb_xl = excel.Workbooks.Open(excel_path, ReadOnly=True, UpdateLinks=0)

        # -----------------------
        # (A) 수분 CSV (표시값 그대로)
        # -----------------------
        try:
            if need_moist:
                out_dir = os.path.join(MOIST_ROOT, yyyy, mm_folder)
                out_csv = os.path.join(out_dir, f"{sample_no}.csv")

                ws_m_xl = wb_xl.Worksheets("수분량자동측정")
                _export_moist_csv_from_open_ws(excel, ws_m_xl, out_csv, max_rows=6)
                print(f"✅ 수분 CSV 저장: {out_csv}")
            else:
                if ws_in is None:
                    print("⚠ 수분: '입력' 시트 없음 → 스킵")
                else:
                    b18 = "" if ws_in["B18"].value is None else str(ws_in["B18"].value).strip()
                    if b18 != "사용":
                        print("▶ 수분: 입력!B18 != '사용' → 스킵")
                    elif "수분량자동측정" not in wb.sheetnames:
                        print("⚠ 수분: '수분량자동측정' 시트 없음 → 스킵")
        except Exception as e:
            print(f"❌ 수분 백데이터 실패: {e}")

        # -----------------------
        # (B) THC
        # -----------------------
        try:
            if not need_thc:
                if ws_an is None:
                    print("⚠ THC: '입력(분석값)' 시트 없음 → 스킵")
                else:
                    print("▶ THC: 입력(분석값)!A64 빈칸 → 스킵")
                return

            out_dir = os.path.join(THC_ROOT, yyyy, mm_folder)
            os.makedirs(out_dir, exist_ok=True)

            if is_fid_mode:
                # C65==1 → THC 측정값(FID) → CSV (openpyxl 그대로)
                sheet_name = "THC 측정값(FID)"
                if sheet_name not in wb.sheetnames:
                    cand = [n for n in wb.sheetnames if n.replace(" ", "") == sheet_name.replace(" ", "")]
                    if cand:
                        sheet_name = cand[0]
                    else:
                        print(f"⚠ THC: 시트 없음({sheet_name}) → 스킵")
                        return

                ws_t = wb[sheet_name]
                out_csv = os.path.join(out_dir, f"{sample_no}.csv")
                _export_ws_to_csv(ws_t, out_csv)
                print(f"✅ THC(FID모드) CSV 저장: {out_csv}")

            else:
                # C65!=1 → THC 측정값(PF) → .FID (Excel 복사 기반, 150행 고정)
                sheet_name = "THC 측정값(PF)"
                try:
                    ws_pf_xl = wb_xl.Worksheets(sheet_name)
                except:
                    ws_pf_xl = None
                    for sh in wb_xl.Worksheets:
                        if str(sh.Name).replace(" ", "") == sheet_name.replace(" ", ""):
                            ws_pf_xl = sh
                            break
                    if ws_pf_xl is None:
                        print(f"⚠ THC: 시트 없음({sheet_name}) → 스킵")
                        return

                out_fid = os.path.join(out_dir, f"{sample_no}.FID")
                _export_pf_fid_from_open_ws(excel, ws_pf_xl, out_fid, fixed_rows=150)
                print(f"✅ THC(PF모드) FID 저장: {out_fid}")

        except Exception as e:
            print(f"❌ THC 백데이터 실패: {e}")

    finally:
        try:
            if wb_xl is not None:
                try:
                    excel.CutCopyMode = False
                except:
                    pass
                wb_xl.Close(False)
        except:
            pass
