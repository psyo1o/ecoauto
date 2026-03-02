import os
import re
import io
import csv
import datetime as dt
import tkinter as tk
from tkinter import filedialog, messagebox

# -----------------------------
# UI 로그(콘솔 대신 창 아래에 표시)
# -----------------------------
import sys
import threading
import queue
import traceback
from tkinter.scrolledtext import ScrolledText

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# ★추가(샘플 형태 비교용)
import hashlib

#===================경고 제거========================
# ✅ openpyxl 조건부서식 경고 숨김 (오류 아님, 콘솔 정리용)
import warnings
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Conditional Formatting extension is not supported and will be removed"
)

# ============================================================
# 공통 유틸 모듈 (모듈화)
# ============================================================
from data_utils import (
    parse_sn_date, parse_sn_team, extract_sn_text,
    excel_value_to_time, parse_time_range as _parse_time_range_data
)
from log_utils import log_error


# -----------------------------
# 환경 설정
# -----------------------------
BASE_ROOT = r"\\192.168.10.163\측정팀\2.성적서"
REPORT_DIRS = ["0 0.입력중", "0 1.완료", "0 2.검토중", "0 3.검토완료", "0 4.출력완료&에코랩입력중", "0 5.최종완료"]
MOISTURE_DIR = "0.수분량"
THC_DIR = "0.THC"
OUTPUT_ROOT = r"\\192.168.10.163\측정팀\10.검토\2.발송대장 검토"

# ★ 샘플 파일(폴더 고정, 파일명만 맞춰두면 됨)
#   - 수분 CSV 샘플: 0.수분량\수분량샘플.csv
#   - THC CSV 샘플 : 0.THC\FID샘플.csv
#   - THC FID 샘플 : 0.THC\PF샘플.FID
MOISTURE_SAMPLE = os.path.join(BASE_ROOT, MOISTURE_DIR, "수분량샘플.csv")
THC_CSV_SAMPLE  = os.path.join(BASE_ROOT, THC_DIR, "FID샘플.csv")
THC_FID_SAMPLE  = os.path.join(BASE_ROOT, THC_DIR, "PF샘플.FID")

# ★ 샘플 signature 캐시(속도)
_TEMPLATE_SIG = {"moisture_csv": None, "thc_csv": None, "thc_fid": None}


# -----------------------------
# 유틸 함수들
# -----------------------------
def ensure_output_dir():
    os.makedirs(OUTPUT_ROOT, exist_ok=True)


def _cell_has_equipment(v) -> bool:
    """
    장비표 셀값이 '있음'인지 판정.
    - 빈값/없음/미사용/0/X 등은 없음으로 취급
    """
    if v is None:
        return False
    s = str(v).strip()
    if s == "":
        return False
    low = s.lower()
    if low in ("0", "x", "no", "n", "없음", "미사용", "미 보유", "미보유"):
        return False
    return True


def has_team_equipment(sh_in, team: int, keywords: list[str]):
    """
    [입력] 시트 장비표(AH9:AQ13)를 기준으로 팀(team)의 장비 보유 여부 확인.

    - 팀행: 9~13행 (team=1 -> 9행, team=5 -> 13행)
    - 장비 헤더(장비 종류)는 보통 AH8:AQ8 이지만, 파일마다 다를 수 있어
      AH7:AQ9 범위에서 '키워드가 포함된 헤더'를 우선 탐색한다.

    반환: (True/False, detail_msg)
      - 열(장비명) 자체를 못 찾으면 False, "장비항목없음"
      - 팀행 범위 밖이면 False, "팀행없음"
      - 값이 비어있거나 '없음/미사용'이면 False, "장비없음"
      - 예외 발생/범위 접근 실패면 False, "장비표없음"
    """
    def _norm(s: str) -> str:
        # 공백/괄호 등을 제거해서 헤더 표기 흔들림에 강하게
        s = (s or "").strip()
        s = re.sub(r"\s+", "", s)          # 모든 공백 제거
        s = s.replace("(", "").replace(")", "")
        return s

    try:
        # 1) 헤더(장비 종류) 열 찾기: AH7:AQ9 중에서 keywords가 포함된 셀
        target_col = None
        for hdr_row in (7, 8, 9):
            hdr_cells = list(sh_in[f"AH{hdr_row}":f"AQ{hdr_row}"][0])
            for c in hdr_cells:
                name = _norm(str(c.value or ""))
                if not name:
                    continue
                for kw in keywords:
                    if not kw:
                        continue
                    if _norm(kw) in name:
                        target_col = c.column
                        break
                if target_col is not None:
                    break
            if target_col is not None:
                break

        if target_col is None:
            return False, "장비항목없음"

        # 2) 팀행 결정 (AH9:AQ13)
        team_row = 8 + int(team)  # team=1 -> 9
        if team_row < 9 or team_row > 13:
            return False, "팀행없음"

        v = sh_in.cell(row=team_row, column=target_col).value
        if _cell_has_equipment(v):
            return True, ""
        return False, "장비없음"
    except Exception:
        return False, "장비표없음"


# parse_time_range: data_utils에서 import (_parse_time_range_data)
def parse_time_range(text: str):
    """receipt.py 내부 호환성 래퍼"""
    return _parse_time_range_data(text)


def time_in_range(inner_start: dt.time, inner_end: dt.time,
                  outer_start: dt.time, outer_end: dt.time):
    if None in (inner_start, inner_end, outer_start, outer_end):
        return False
    is_min = inner_start.hour * 60 + inner_start.minute
    ie_min = inner_end.hour * 60 + inner_end.minute
    os_min = outer_start.hour * 60 + outer_start.minute
    oe_min = outer_end.hour * 60 + outer_end.minute
    return is_min >= os_min and ie_min <= oe_min


def find_latest_report_file(sn: str):
    """
    시료번호(sn)과 정확히 일치하면서, sn 바로 뒤에 숫자가 붙지 않은 파일만 찾는다.
    A2512053-01 뒤에 숫자오면(A2512053-011) 제외.
    """
    pattern = re.compile(rf"^{re.escape(sn)}(?!\d)", re.IGNORECASE)
    candidates = []

    for d in REPORT_DIRS:
        folder = os.path.join(BASE_ROOT, d)
        if not os.path.isdir(folder):
            continue

        for name in os.listdir(folder):
            low = name.lower()
            if not low.endswith((".xlsx", ".xlsm")):
                continue

            stem, _ext = os.path.splitext(name)
            if not pattern.match(stem):
                continue

            full = os.path.join(folder, name)
            try:
                mtime = os.path.getmtime(full)
            except OSError:
                continue

            candidates.append((mtime, full))

    if not candidates:
        return None

    candidates.sort(reverse=True, key=lambda x: x[0])
    return candidates[0][1]


def moisture_file_path(sn: str):
    d = parse_sn_date(sn)
    if d is None:
        return None
    folder = os.path.join(
        BASE_ROOT,
        MOISTURE_DIR,
        f"{d.year}",
        f"{d.month}월"
    )
    return os.path.join(folder, f"{sn}.csv")


def find_thc_file(sn: str, prefer_ext: str | None = None):
    d = parse_sn_date(sn)
    if d is None:
        return None, "파일없음"

    folder = os.path.join(
        BASE_ROOT,
        THC_DIR,
        f"{d.year}",
        f"{d.month}월"
    )
    if not os.path.isdir(folder):
        return None, "파일없음"

    files = []
    for name in os.listdir(folder):
        lower = name.lower()
        if not lower.startswith(sn.lower()):
            continue
        if not (lower.endswith(".fid") or lower.endswith(".csv")):
            continue
        full = os.path.join(folder, name)
        try:
            mtime = os.path.getmtime(full)
        except OSError:
            continue
        files.append((mtime, full))

    if not files:
        return None, "파일없음"

    # 반대 확장자 존재 여부 체크
    has_fid = any(f[1].lower().endswith(".fid") for f in files)
    has_csv = any(f[1].lower().endswith(".csv") for f in files)

    if prefer_ext == "fid":
        if has_csv:
            return None, "확장자오류"
        fid_files = [f for f in files if f[1].lower().endswith(".fid")]
        if not fid_files:
            return None, "파일없음"
        fid_files.sort(reverse=True, key=lambda x: x[0])
        return fid_files[0][1], "OK"

    elif prefer_ext == "csv":
        if has_fid:
            return None, "확장자오류"
        csv_files = [f for f in files if f[1].lower().endswith(".csv")]
        if not csv_files:
            return None, "파일없음"
        csv_files.sort(reverse=True, key=lambda x: x[0])
        return csv_files[0][1], "OK"

    files.sort(reverse=True, key=lambda x: x[0])
    return files[0][1], "OK"


def get_file_times(path: str):
    try:
        cts = os.path.getctime(path)
        mts = os.path.getmtime(path)
        return dt.datetime.fromtimestamp(cts), dt.datetime.fromtimestamp(mts)
    except OSError:
        return None, None


def format_time_or_blank(t: dt.time | None):
    return "" if t is None else t.strftime("%H:%M")


def format_dt_or_blank(x: dt.datetime | None):
    return "" if x is None else x.strftime("%Y-%m-%d %H:%M")


# ✅ 시료번호 추출/비교 유틸
def report_sn_from_sheet_b1(sh_in) -> str:
    """성적서 [입력] 시트 B1에서 시료번호 추출"""
    return extract_sn_text(sh_in["B1"].value)


def build_report_status(base_status: str, sn_dj: str, rep_sn_b1: str) -> str:
    """
    성적서상태 단일 컬럼에서
    - base_status (예: OK, 성적서시간없음, 입력시트없음 등)
    - SN 비교 결과 (OK, SN불일치, SN없음)
    를 합쳐 표기.
    """
    sn_dj_u = (sn_dj or "").strip().upper()
    rep_u = (rep_sn_b1 or "").strip().upper()

    if rep_u == "":
        sn_flag = "SN없음"
    elif rep_u == sn_dj_u:
        sn_flag = "OK"
    else:
        sn_flag = "SN불일치"

    if base_status == "OK":
        return sn_flag  # OK 여부는 SN 비교로 결정
    else:
        if sn_flag == "OK":
            return base_status
        return f"{base_status}|{sn_flag}"


# =========================================================
# ★추가: 샘플 기반 "형태" 비교(숫자만 바뀌는 파일 대응)
# =========================================================
def _read_text_any_encoding(path: str) -> str:
    data = open(path, "rb").read()
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-16", "latin1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")

def _normalize_text_for_signature(text: str, kind: str) -> str:
    s = text.replace("\r\n", "\n").replace("\r", "\n")

    # CSV 계열은 구조(헤더/컬럼수/데이터 컬럼수 일관성)만 남김
    if kind in ("moisture_csv", "thc_csv"):
        # 구분자 흔들림 대비(탭/세미콜론)
        s2 = s.replace("\t", ",")
        if s2.count(";") > s2.count(","):
            s2 = s2.replace(";", ",")

        rows = []
        for row in csv.reader(io.StringIO(s2)):
            row = [c.strip() for c in row]
            # 끝 빈컬럼 제거(,, 제거)
            while row and row[-1] == "":
                row.pop()
            # 완전 빈 줄 제거(,,,, 같은 줄)
            if not row or all(c == "" for c in row):
                continue
            rows.append(row)

        if not rows:
            return "csv|empty"

        header = rows[0]
        col_n = len(header)

        # 데이터 줄들의 컬럼수 패턴만 기록(행 개수는 무시)
        data_col_counts = set()
        for r in rows[1:]:
            data_col_counts.add(len(r))

        header_join = ",".join(header).strip()

        # 헤더에 시료번호/숫자 등이 섞여도 구조로 취급되게 토큰화
        header_join = re.sub(r"[A-Za-z]\d{7}-\d{2,3}", "<SN>", header_join)
        header_join = re.sub(r"(?<![A-Za-z])[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", "<NUM>", header_join)

        counts = ",".join(str(x) for x in sorted(data_col_counts)) if data_col_counts else "none"
        return f"csv|cols={col_n}|header={header_join.lower()}|data_cols={counts}"

    # FID는 기존처럼 “형태+토큰화”로 비교(원하면 이것도 구조형으로 바꿀 수 있음)
    s = re.sub(r"[A-Za-z]\d{7}-\d{2,3}", "<SN>", s)
    s = re.sub(r"\b\d{4}[/-]\d{2}[/-]\d{2}\b", "<DATE>", s)
    s = re.sub(r"\b\d{2}[/-]\d{2}[/-]\d{2,4}\b", "<DATE>", s)
    s = re.sub(r"\b\d{1,2}:\d{2}:\d{2}\b", "<TIME>", s)
    s = re.sub(r"\b\d{1,2}:\d{2}\b", "<TIME>", s)
    s = re.sub(r"(?<![A-Za-z])[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", "<NUM>", s)
    s = s.replace("\t", " ")
    # FID 단위/깨진문자 정규화 (추가)
    s = s.replace("캜", "c").replace("�C", "c")
    s = re.sub(r"[ ]{2,}", " ", s)

    lines = []
    for line in s.split("\n"):
        line = line.strip()
        if not line:
            continue
        lines.append(line)

    return "\n".join(lines).lower()



def _sig(norm_text: str) -> str:
    return hashlib.sha1(norm_text.encode("utf-8")).hexdigest()


def _get_template_signature(kind: str) -> tuple[str | None, str]:
    """
    kind:
      - moisture_csv : MOISTURE_SAMPLE
      - thc_csv      : THC_CSV_SAMPLE
      - thc_fid      : THC_FID_SAMPLE
    """
    if _TEMPLATE_SIG.get(kind):
        return _TEMPLATE_SIG[kind], "OK"

    if kind == "moisture_csv":
        sample_path = MOISTURE_SAMPLE
    elif kind == "thc_csv":
        sample_path = THC_CSV_SAMPLE
    elif kind == "thc_fid":
        sample_path = THC_FID_SAMPLE
    else:
        return None, "kind오류"

    if not os.path.exists(sample_path):
        return None, f"샘플없음:{sample_path}"

    try:
        raw = _read_text_any_encoding(sample_path)
        norm = _normalize_text_for_signature(raw, kind)
        sigv = _sig(norm)
        _TEMPLATE_SIG[kind] = sigv
        return sigv, "OK"
    except Exception:
        return None, f"샘플읽기오류:{sample_path}"


def validate_file_content_by_sample(path: str, kind: str) -> tuple[bool, str]:
    """
    샘플 기반 형태 검증
    - 샘플 없으면 NG
    - 형태 다르면 NG
    """
    tmpl_sig, msg = _get_template_signature(kind)
    if tmpl_sig is None:
        return False, f"샘플없음({msg})"

    try:
        raw = _read_text_any_encoding(path)
        norm = _normalize_text_for_signature(raw, kind)
        sigv = _sig(norm)
        if sigv == tmpl_sig:
            return True, ""
        return False, "내용형태오류"
    except Exception:
        return False, "내용읽기오류"


def _save_unique(report_wb: Workbook, out_path: str) -> str:
    """같은 파일명이 있으면 _2, _3 붙여서 저장"""
    if not os.path.exists(out_path):
        report_wb.save(out_path)
        return out_path
    base, ext = os.path.splitext(out_path)
    k = 2
    while True:
        p = f"{base}_{k}{ext}"
        if not os.path.exists(p):
            report_wb.save(p)
            return p
        k += 1


# -----------------------------
# 핵심 검사 로직
# -----------------------------
def process_daejang(daejang_path: str,
                    start_day: int | None,
                    end_day: int | None) -> list[str]:

    wb_dj = load_workbook(daejang_path, data_only=True)

    # 헤더(원래 그대로)
    headers = [
        "대장시트", "시료번호", "엑셀시료번호",
        "대장시작", "대장끝",
        "성적서시작", "성적서끝",
        "수분시작", "수분끝",
        "THC시작", "THC끝",
        "성적서상태", "수분상태", "THC상태",
        "수분_생성", "수분_수정",
        "THC_생성", "THC_수정",
        "THC_B66"
    ]

    ensure_output_dir()
    out_paths: list[str] = []

    for sheet_name in wb_dj.sheetnames:
        if not sheet_name.endswith("일"):
            continue
        day_str = sheet_name[:-1]
        if not day_str.isdigit():
            continue
        day_int = int(day_str)

        if start_day is not None and day_int < start_day:
            continue
        if end_day is not None and day_int > end_day:
            continue

        # ✅ 하루마다 워크북 새로 만들어서 저장 (로직 자체는 동일, 결과만 분리)
        report_wb = Workbook()
        report_ws = report_wb.active
        report_ws.title = "Report"
        report_ws.append(headers)

        day_sn_dates: list[dt.date] = []

        ws = wb_dj[sheet_name]
        max_row = ws.max_row

        for row in range(2, max_row + 1):
            sn = str(ws.cell(row=row, column=1).value or "").strip()
            time_str = str(ws.cell(row=row, column=2).value or "").strip()

            if not sn or not time_str:
                continue

            sn_dj = sn.strip().upper()

            sn_date = parse_sn_date(sn)
            if sn_date is not None:
                day_sn_dates.append(sn_date)

            dj_start_t, dj_end_t = parse_time_range(time_str)
            if dj_start_t is None or dj_end_t is None:
                report_ws.append([
                    sheet_name, sn_dj, "",
                    "", "", "", "", "", "", "", "",
                    "대장시간오류", "", "",
                    "", "", "", "", ""
                ])
                continue

            rep_path = find_latest_report_file(sn)
            if rep_path is None:
                report_ws.append([
                    sheet_name, sn_dj, "",
                    format_time_or_blank(dj_start_t),
                    format_time_or_blank(dj_end_t),
                    "", "", "", "", "", "",
                    "성적서없음", "", "",
                    "", "", "", "", ""
                ])
                continue

            print(">> 검사 중 파일:", rep_path)

            try:
                wb_rep = load_workbook(rep_path, data_only=True)
            except Exception:
                report_ws.append([
                    sheet_name, sn_dj, "",
                    format_time_or_blank(dj_start_t),
                    format_time_or_blank(dj_end_t),
                    "", "", "", "", "", "",
                    "성적서열기오류", "", "",
                    "", "", "", "", ""
                ])
                continue

            try:
                sh_in = wb_rep["입력"]
            except KeyError:
                report_ws.append([
                    sheet_name, sn_dj, "",
                    format_time_or_blank(dj_start_t),
                    format_time_or_blank(dj_end_t),
                    "", "", "", "", "", "",
                    "입력시트없음", "", "",
                    "", "", "", "", ""
                ])
                wb_rep.close()
                continue

            rep_sn_b1 = report_sn_from_sheet_b1(sh_in)

            rep_start_t = excel_value_to_time(sh_in["E5"].value)
            rep_end_t = excel_value_to_time(sh_in["F5"].value)
            if rep_start_t is None or rep_end_t is None:
                report_ws.append([
                    sheet_name, sn_dj, rep_sn_b1,
                    format_time_or_blank(dj_start_t),
                    format_time_or_blank(dj_end_t),
                    "", "", "", "", "", "",
                    build_report_status("성적서시간없음", sn_dj, rep_sn_b1), "", "",
                    "", "", "", "", ""
                ])
                wb_rep.close()
                continue

            base_date = sn_date if sn_date is not None else dt.date(2000, 1, 1)

            team_no = parse_sn_team(sn_dj)

            # 장비표 기반 체크(입력 시트 AH9:AQ13)
            # - 수분량 사용이면: 수분량자동측정기 필요
            # - THC 사용이면: 대기배출가스(THC)측정기 필요
            #   (팀 번호를 못 뽑으면 장비 체크는 스킵)


            # -----------------------------
            # 수분량 검사 (원래 로직 유지 + OK일 때만 샘플 비교 추가)
            # -----------------------------
            moisture_status = ""
            m_start_t = None
            m_end_t = None
            m_ct = None
            m_mt = None

            # ★추가: 수분량 파일 존재 여부(미사용/장비없음 상태에서도 파일 있으면 오류 처리)
            m_path_any = moisture_file_path(sn)
            moisture_file_exists = bool(m_path_any and os.path.exists(m_path_any))
            if moisture_file_exists:
                m_ct, m_mt = get_file_times(m_path_any)

            use_flag = str(sh_in["B18"].value or "").strip()
            if use_flag != "사용":
                moisture_status = "미사용"
                if moisture_file_exists:
                    moisture_status = "미사용(파일있음)오류"
            else:
                # ★추가: 장비 체크 (수분량 사용이면 수분량자동측정기 필수)
                if team_no is not None:
                    ok_eq, eq_msg = has_team_equipment(sh_in, team_no, ["수분량자동측정기"])
                    if not ok_eq:
                        moisture_status = "장비항목없음" if eq_msg == "장비항목없음" else "장비없음"
                        if moisture_file_exists:
                            moisture_status += "(파일있음)"

                if moisture_status == "":
                    m_start_t = excel_value_to_time(sh_in["B23"].value)
                    if m_start_t is None:
                        moisture_status = "시간없음"
                    else:
                        m_end_t = (dt.datetime.combine(base_date, m_start_t) +
                                   dt.timedelta(minutes=5)).time()

                        if not time_in_range(m_start_t, m_end_t, rep_start_t, rep_end_t):
                            moisture_status = "시간오류"
                        else:
                            m_path = m_path_any
                            if not m_path or not os.path.exists(m_path):
                                moisture_status = "파일없음"
                            else:
                                m_ct, m_mt = get_file_times(m_path)
                                if m_ct is None or m_mt is None:
                                    moisture_status = "파일시간오류"
                                else:
                                    m_end_dt = dt.datetime.combine(base_date, m_end_t)
                                    if m_ct >= m_end_dt and m_mt >= m_end_dt:
                                        moisture_status = "OK"
                                        # ★추가: 확장자/파일시간 OK인 경우에만 샘플 형태 비교
                                        ok, reason = validate_file_content_by_sample(m_path, "moisture_csv")
                                        if not ok:
                                            moisture_status = reason
                                    else:
                                        moisture_status = "생성/수정 시간오류"
            # THC 검사 (원래 로직 유지 + OK일 때만 샘플 비교 추가)
            # -----------------------------
            thc_status = ""
            thc_start_t = None
            thc_end_t = None
            thc_ct = None
            thc_mt = None
            thc_b66 = ""

            try:
                sh_thc = wb_rep["입력(분석값)"]
            except KeyError:
                thc_status = "시트없음"
            else:
                a64_hidden = bool(getattr(sh_thc.row_dimensions.get(64), "hidden", False)) or \
                             bool(getattr(sh_thc.column_dimensions.get("A"), "hidden", False))
                a64_raw = str(sh_thc["A64"].value or "").strip()

                # ★추가: THC 파일 존재 여부(미사용/장비없음 상태에서도 파일 있으면 오류 처리)
                t_any_path, t_any_status = find_thc_file(sn, prefer_ext=None)
                thc_file_exists = bool(t_any_path and t_any_status == "OK" and os.path.exists(t_any_path))
                if thc_file_exists:
                    thc_ct, thc_mt = get_file_times(t_any_path)

                if a64_hidden or a64_raw == "":
                    thc_status = "미사용"
                    if thc_file_exists:
                        thc_status = "미사용(파일있음)오류"
                else:
                    # ★추가: 장비 체크 (THC 사용이면 대기배출가스(THC)측정기 필수)
                    if team_no is not None:
                        ok_eq, eq_msg = has_team_equipment(sh_in, team_no, ["총탄화수소", "THC"])
                        if not ok_eq:
                            thc_status = "장비없음" if eq_msg != "장비항목없음" else "장비항목없음"
                            if thc_file_exists:
                                thc_status += "(파일있음)"

                    if thc_status == "":
                        c65_num = 1 if str(sh_thc["C65"].value).strip() == "1" else 0

                        if c65_num == 0:
                            thc_b66 = str(sh_thc["B66"].value or "").strip()

                        val_i64 = sh_thc["I64"].value
                        val_k64 = sh_thc["K64"].value

                        if isinstance(val_i64, str):
                            val_i64 = val_i64.strip()
                        if isinstance(val_k64, str):
                            val_k64 = val_k64.strip()

                        thc_start_t = excel_value_to_time(val_i64)
                        raw_k64 = excel_value_to_time(val_k64)

                        if thc_start_t is None or raw_k64 is None:
                            thc_status = "시간없음"
                        else:
                            if c65_num == 1:
                                thc_end_t = raw_k64
                            else:
                                thc_end_dt = dt.datetime.combine(base_date, raw_k64) + dt.timedelta(minutes=3)
                                thc_end_t = thc_end_dt.time()

                            if not time_in_range(thc_start_t, thc_end_t, rep_start_t, rep_end_t):
                                thc_status = "시간오류"
                            else:
                                prefer_ext = "fid" if c65_num == 0 else "csv"
                                t_path, ext_status = find_thc_file(sn, prefer_ext=prefer_ext)

                                if ext_status == "확장자오류":
                                    thc_status = "확장자오류"
                                elif ext_status == "파일없음":
                                    thc_status = "파일없음"
                                else:
                                    thc_ct, thc_mt = get_file_times(t_path)
                                    if thc_ct is None or thc_mt is None:
                                        thc_status = "파일시간오류"
                                    else:
                                        thc_end_dt2 = dt.datetime.combine(base_date, thc_end_t)
                                        if thc_ct >= thc_end_dt2 and thc_mt >= thc_end_dt2:
                                            thc_status = "OK"
                                            # ★추가: 확장자/파일시간 OK인 경우에만 샘플 형태 비교
                                            kind = "thc_fid" if t_path.lower().endswith(".fid") else "thc_csv"
                                            ok, reason = validate_file_content_by_sample(t_path, kind)
                                            if not ok:
                                                thc_status = reason
                                        else:
                                            thc_status = "생성/수정 시간오류"

            wb_rep.close()

            report_status = build_report_status("OK", sn_dj, rep_sn_b1)

            report_ws.append([
                sheet_name,
                sn_dj,
                rep_sn_b1,
                format_time_or_blank(dj_start_t),
                format_time_or_blank(dj_end_t),
                format_time_or_blank(rep_start_t),
                format_time_or_blank(rep_end_t),
                format_time_or_blank(m_start_t),
                format_time_or_blank(m_end_t),
                format_time_or_blank(thc_start_t),
                format_time_or_blank(thc_end_t),
                report_status,
                moisture_status,
                thc_status,
                format_dt_or_blank(m_ct),
                format_dt_or_blank(m_mt),
                format_dt_or_blank(thc_ct),
                format_dt_or_blank(thc_mt),
                thc_b66
            ])

        # 색상 처리(원래 그대로)
        red_fill = PatternFill(start_color="FF9696", end_color="FF9696", fill_type="solid")
        blue_fill = PatternFill(start_color="C0E6FF", end_color="C0E6FF", fill_type="solid")
        gray_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        for r in range(2, report_ws.max_row + 1):
            for c in range(12, 15):
                cell = report_ws.cell(row=r, column=c)
                val = str(cell.value or "").lower()

                if ("미사용" == val):
                    cell.fill = gray_fill
                elif ("ok" in val):
                    cell.fill = blue_fill
                elif ("오류" in val) or ("없음" in val) or ("불일치" in val) or ("sn없음" in val) or ("시간" in val and "ok" not in val):
                    cell.fill = red_fill

        # ✅ 파일명: 하루마다 YYMMDD 발송대장 체크.xlsx
        if day_sn_dates:
            d = min(day_sn_dates)
            fname = f"{str(d.year)[2:]}{d.month:02d}{d.day:02d} 발송대장 체크.xlsx"
        else:
            # 날짜 못 뽑으면 시트명 기반
            fname = f"{sheet_name} 발송대장 체크.xlsx"

        out_path = os.path.join(OUTPUT_ROOT, fname)
        saved = _save_unique(report_wb, out_path)
        out_paths.append(saved)

    return out_paths


# -----------------------------
# 콘솔(stdout/stderr) 출력 → UI 로그창으로 전달
# -----------------------------
class QueueWriter:
    """print() 출력(stdout/stderr)을 queue로 모으는 writer"""

    def __init__(self, q: queue.Queue):
        self.q = q

    def write(self, msg: str):
        if msg:
            self.q.put(msg)

    def flush(self):
        pass


# -----------------------------
# Tkinter UI
# -----------------------------
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.attributes('-topmost', True)
        root.title("발송대장/성적서/수분THC백데이터 검토")

        self.daejang_path = tk.StringVar()
        self.start_day = tk.StringVar()
        self.end_day = tk.StringVar()
        self.status = tk.StringVar(value="대장 파일을 선택하세요.")

        frm_file = tk.Frame(root)
        frm_file.pack(padx=10, pady=5, fill="x")

        tk.Label(frm_file, text="대장 파일:").pack(side="left")
        tk.Entry(frm_file, textvariable=self.daejang_path, width=60).pack(side="left", padx=5)
        tk.Button(frm_file, text="찾기", command=self.browse_file).pack(side="left")

        frm_day = tk.Frame(root)
        frm_day.pack(padx=10, pady=5, fill="x")

        tk.Label(frm_day, text="시작일(숫자):").pack(side="left")
        tk.Entry(frm_day, textvariable=self.start_day, width=5).pack(side="left", padx=5)
        tk.Label(frm_day, text="종료일(숫자):").pack(side="left")
        tk.Entry(frm_day, textvariable=self.end_day, width=5).pack(side="left", padx=5)
        tk.Label(frm_day, text="(비워두면 전체 일자)").pack(side="left")

        frm_ctrl = tk.Frame(root)
        frm_ctrl.pack(padx=10, pady=10, fill="x")

        tk.Button(frm_ctrl, text="검사 시작", command=self.run_check).pack(side="left")
        tk.Label(frm_ctrl, textvariable=self.status, fg="blue").pack(side="left", padx=10)

        # -----------------------------
        # 로그창(콘솔 대신 UI에 출력)
        # -----------------------------
        self.log_q = queue.Queue()

        frm_log = tk.Frame(root)
        frm_log.pack(padx=10, pady=(0, 10), fill="both", expand=True)

        tk.Label(frm_log, text="로그:").pack(anchor="w")

        self.log_text = ScrolledText(frm_log, height=12)
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")

        # print()/에러 출력이 UI 로그로 가도록 리다이렉트
        sys.stdout = QueueWriter(self.log_q)
        sys.stderr = QueueWriter(self.log_q)

        # queue에 쌓인 로그를 주기적으로 UI에 뿌려줌
        self._pump_log()

    def _pump_log(self):
        """queue에 쌓인 로그를 Text 위젯에 뿌리는 루프(메인스레드)"""
        try:
            while True:
                msg = self.log_q.get_nowait()
                self.log_text.configure(state="normal")
                self.log_text.insert("end", msg)
                self.log_text.see("end")
                self.log_text.configure(state="disabled")
        except queue.Empty:
            pass

        self.root.after(50, self._pump_log)

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="시료접수발송대장(대기) 파일 선택",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xls")]
        )
        if path:
            self.daejang_path.set(path)
            self.status.set("검사 준비 완료. '검사 시작' 버튼을 눌러주세요.")

    def run_check(self):
        path = self.daejang_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("오류", "대장 파일을 먼저 선택하세요.")
            return

        def parse_int(s):
            s = s.strip()
            if not s:
                return None
            try:
                v = int(s)
                if 1 <= v <= 31:
                    return v
            except ValueError:
                return None
            return None

        s_day = parse_int(self.start_day.get())
        e_day = parse_int(self.end_day.get())

        if s_day is not None and e_day is not None and s_day > e_day:
            messagebox.showerror("오류", "시작일이 종료일보다 클 수 없습니다.")
            return

        self.status.set("검사 중... 잠시만 기다려주세요.")
        self.root.update_idletasks()

        # UI 멈춤 방지: 검사 로직은 백그라운드 스레드에서 실행
        def worker():
            try:
                out_paths = process_daejang(path, s_day, e_day)
                self.root.after(0, lambda: self._on_done(out_paths))
            except Exception as e:
                tb = traceback.format_exc()
                self.root.after(0, lambda: self._on_error(e, tb))

        threading.Thread(target=worker, daemon=True).start()

    def _on_error(self, e: Exception, tb: str):
        self.status.set("오류 발생")
        # 상세 traceback은 로그창에 남기고, 팝업은 기존처럼 간단히 표시
        print(tb)
        messagebox.showerror("오류", f"검사 중 오류 발생:\n{e}")

    def _on_done(self, out_paths):
        if not out_paths:
            self.status.set("결과 없음")
            messagebox.showinfo("완료", "해당 일자에 검사할 데이터가 없습니다.")
            return

        self.status.set("검사 완료")
        messagebox.showinfo("완료", "검사 완료!\n결과 파일:\n" + "\n".join(out_paths))


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log_error("receipt.main", e)
        raise