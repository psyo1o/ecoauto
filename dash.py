# dash_v4.py
# 발송대장(Report) + 측정인 검토(eco_check 결과, 요약 시트)들을 합쳐 대시보드 생성
# - eco_check 팀별 결과 파일(여러 개) 선택 가능
# - 장비/인력/차량 "같은 날짜에 여러 팀이 사용" 중복 검사 추가(시간 겹침은 보지 않음)
#
# 출력 시트(기존 그대로): 대시보드 / 상세 / 원인집계

import re
import os
from datetime import datetime, date, time
from typing import List, Tuple, Dict, Any, Optional, Set

import pandas as pd
import openpyxl
import threading
import queue
import tkinter as tk
from tkinter import ttk

from tkinter import Tk
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from tkinter.messagebox import showerror, showinfo

from data_utils import parse_sn as _parse_sn_raw, parse_time as _parse_time_raw
from gui_common import LogPanel, set_window_topmost
from log_utils import log_error


SN_RE = re.compile(r'^A\d{7}-\d{2}$')  # 예: A2512063-01  (YYMMDD + 팀1자리)


# 기본 저장 폴더(요청: 네트워크 경로로 자동 저장)
DEFAULT_OUTPUT_DIR = r"\\192.168.10.163\측정팀\10.검토\5.종합 검토"
DEFAULT_SEND_REVIEW_DIR = r"\\192.168.10.163\측정팀\10.검토\2.발송대장 검토"
DEFAULT_MEASUREMENT_REVIEW_DIR = r"\\192.168.10.163\측정팀\10.검토\3.측정인 검토"

def _extract_yymmdd_from_sample(sn: Any) -> Optional[str]:
    """시료번호/문자열에서 YYMMDD(6자리) 추출. 예: A2601053-01 -> 260105"""
    if sn is None:
        return None
    s = str(sn).strip()
    m = re.match(r"^[A-Za-z](\d{6})\d?-\d{2}$", s)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{6})", s)
    return m2.group(1) if m2 else None

def _best_date_tag(send_df: pd.DataFrame, review_paths: List[str]) -> str:
    """저장 파일명에 넣을 날짜 태그(YYMMDD). 우선순위: 발송대장 시료번호 -> 검토파일명 -> 오늘"""
    for col in ["시료번호", "시료 번호", "Sample", "샘플", "SN"]:
        if col in send_df.columns:
            vals = send_df[col].dropna().astype(str).tolist()
            for v in vals[:80]:
                tag = _extract_yymmdd_from_sample(v)
                if tag:
                    return tag
    for p in review_paths:
        tag = _extract_yymmdd_from_sample(os.path.basename(p))
        if tag:
            return tag
    return datetime.now().strftime("%y%m%d")

def _unique_path(path: str) -> str:
    """같은 이름이 있으면 _01, _02...를 붙여서 충돌 방지"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    for i in range(1, 100):
        cand = f"{base}_{i:02d}{ext}"
        if not os.path.exists(cand):
            return cand
    return f"{base}_{int(datetime.now().timestamp())}{ext}"



def _sheet_to_df(wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"'{sheet_name}' 시트를 찾지 못했습니다.")
    ws = wb[sheet_name]
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    return df


def _parse_sn(sn: str) -> Tuple[Optional[str], Optional[str]]:
    """시료번호에서 날짜(YYYY-MM-DD), 팀 문자열을 추출. data_utils.parse_sn 래퍼."""
    d_obj, team, _ = _parse_sn_raw(str(sn).strip())
    if d_obj is None:
        return None, None
    return d_obj.isoformat(), str(team)


def _parse_time(v: Any) -> Optional[str]:
    """HH:MM 문자열로 정규화. data_utils.parse_time 래퍼."""
    t = _parse_time_raw(v)
    if t is None:
        return None
    return f"{t.hour:02d}:{t.minute:02d}"


def _parse_list(v: Any) -> List[str]:
    """'a, b, c' -> ['a','b','c']"""
    if v is None:
        return []
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return []
    # eco_check는 보통 ", "로 join함
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p]


def read_send_report(send_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(send_path, data_only=True)
    df = _sheet_to_df(wb, "Report")

    if "시료번호" not in df.columns:
        raise ValueError("발송대장 'Report' 시트에 '시료번호' 컬럼이 없습니다.")

    df["시료번호"] = df["시료번호"].astype(str)
    df = df[df["시료번호"].str.match(SN_RE)].copy()

    def send_overall(row):
        def ok(val, okset):
            if val is None:
                return False
            v = str(val).strip()
            # 값 표준화: '사용안함/미실시/해당없음' 등을 '미사용'으로 통일
            if v in {'사용안함', '미실시', '해당없음', '없음'}:
                v = '미사용'
            return v in okset

        if not ok(row.get("성적서상태"), {"OK"}):
            return "NG"
        if not ok(row.get("수분상태"), {"OK", "미사용"}):
            return "NG"
        if not ok(row.get("THC상태"), {"OK", "미사용"}):
            return "NG"
        return "OK"

    df["발송_종합"] = df.apply(send_overall, axis=1)
    return df


def read_review_summary_one(review_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(review_path, data_only=True)
    df = _sheet_to_df(wb, "요약")

    for col in ["시료번호", "항목", "비교"]:
        if col not in df.columns:
            raise ValueError(f"검토파일 '요약' 시트에 '{col}' 컬럼이 없습니다. ({os.path.basename(review_path)})")

    # 사이트값/엑셀값 컬럼은 있을 수도/없을 수도 있음
    if "사이트값" not in df.columns:
        df["사이트값"] = None
    if "엑셀값" not in df.columns:
        df["엑셀값"] = None

    df["비교"] = df["비교"].astype(str).str.replace("확인 불가", "확인불가")
    df["시료번호"] = df["시료번호"].astype(str)
    df["항목"] = df["항목"].astype(str)
    return df


def _agg_compare(series: pd.Series) -> str:
    s = series.astype(str).tolist()
    if any(x == "NG" for x in s):
        return "NG"
    if any(x == "확인불가" for x in s):
        return "확인불가"
    return "OK"


def read_review_summary_multi(review_paths: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    return:
      df_norm: (시료번호, 항목, 비교, 사이트값, 엑셀값) - (시료번호,항목) 단위로 정규화/병합
      df_sum : (시료번호별 요약)
      cause_review_ng : (검토 NG 항목 집계)
    """
    dfs = []
    for p in review_paths:
        d = read_review_summary_one(p).copy()
        d["소스파일"] = os.path.basename(p)
        dfs.append(d)
    df_all = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    if df_all.empty:
        raise ValueError("검토 파일(요약)에서 읽어온 데이터가 없습니다.")

    # (시료번호, 항목) 단위로 통합(팀별 파일이 여러개여도 1줄로)
    def first_nonnull(x: pd.Series):
        for v in x.tolist():
            if v is None:
                continue
            s = str(v).strip()
            if s and s.lower() != "nan":
                return v
        return None

    df_norm = (
        df_all
        .groupby(["시료번호", "항목"], as_index=False)
        .agg({
            "비교": _agg_compare,
            "사이트값": first_nonnull,
            "엑셀값": first_nonnull,
        })
    )

    # 시료번호별 요약
    summary_rows = []
    for sn, g in df_norm.groupby("시료번호"):
        vc = g["비교"].value_counts().to_dict()
        ok_cnt = int(vc.get("OK", 0))
        ng_cnt = int(vc.get("NG", 0))
        unk_cnt = int(vc.get("확인불가", 0))
        overall = "NG" if ng_cnt > 0 else ("WARN" if unk_cnt > 0 else "OK")

        ng_items = g.loc[g["비교"] == "NG", "항목"].dropna().astype(str).tolist()
        unk_items = g.loc[g["비교"] == "확인불가", "항목"].dropna().astype(str).tolist()

        summary_rows.append({
            "시료번호": sn,
            "검토_OK": ok_cnt,
            "검토_NG": ng_cnt,
            "검토_확인불가": unk_cnt,
            "검토_종합": overall,
            "NG_항목(상위10)": ", ".join(ng_items[:10]),
            "확인불가_항목(상위10)": ", ".join(unk_items[:10]),
        })

    df_sum = pd.DataFrame(summary_rows)

    # 원인 집계(검토 NG 항목)
    cause_review_ng = df_norm[df_norm["비교"] == "NG"]["항목"].astype(str).value_counts().reset_index()
    cause_review_ng.columns = ["검토_NG항목", "건수"]

    return df_norm, df_sum, cause_review_ng


def build_dup_tables(df_norm: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame,
                                                   Set[str], Set[str], Set[str],
                                                   Set[str], Set[str], Set[str]]:
    """
    같은 날짜에 여러 팀이 같은 장비/인력/차량을 사용했는지 검사.
    - 장비/인력/차량이 같은 날 겹치면 무조건 표에 기록하되, 시간이 안 겹치면 'OK(정상교대)'로 명시
    """
    needed = {"시료번호", "항목"}
    if not needed.issubset(df_norm.columns):
        raise ValueError("검토 요약(df_norm) 필수 컬럼이 없습니다.")

    samples = sorted(df_norm["시료번호"].dropna().astype(str).unique().tolist())

    # 시료별 자원/시간 추출
    by_sn: Dict[str, Dict[str, Any]] = {}
    for sn in samples:
        d, team = _parse_sn(sn)
        if not d or not team:
            continue

        def pick(item: str) -> Any:
            sub = df_norm[(df_norm["시료번호"] == sn) & (df_norm["항목"] == item)]
            if sub.empty:
                return None
            v = sub["엑셀값"].iloc[0] if "엑셀값" in sub.columns else None
            if v is None or str(v).strip() == "" or str(v).lower() == "nan":
                v = sub["사이트값"].iloc[0] if "사이트값" in sub.columns else None
            return v

        # 가스, THC, 비산먼지 개별 시간 추출
        gas_start_cands = [
            _parse_time(pick("[RealGrid] 황산화물 / 시작시간")),
            _parse_time(pick("[RealGrid] 질소산화물 / 시작시간")),
            _parse_time(pick("[RealGrid] 일산화탄소 / 시작시간"))
        ]
        gas_end_cands = [
            _parse_time(pick("[RealGrid] 황산화물 / 종료시간")),
            _parse_time(pick("[RealGrid] 질소산화물 / 종료시간")),
            _parse_time(pick("[RealGrid] 일산화탄소 / 종료시간"))
        ]
        
        valid_gas_starts = [s for s in gas_start_cands if s]
        valid_gas_ends = [e for e in gas_end_cands if e]
        
        by_sn[sn] = {
            "날짜": d,
            "팀": team,
            "채취시작": _parse_time(pick("채취시작")),
            "채취끝": _parse_time(pick("채취끝")),
            
            "가스시작": min(valid_gas_starts) if valid_gas_starts else None,
            "가스끝": max(valid_gas_ends) if valid_gas_ends else None,
            "THC시작": _parse_time(pick("[RealGrid] 총탄화수소 / 시작시간")),
            "THC끝": _parse_time(pick("[RealGrid] 총탄화수소 / 종료시간")),
            "비산먼지시작": _parse_time(pick("[RealGrid] 비산먼지 / 시작시간")),
            "비산먼지끝": _parse_time(pick("[RealGrid] 비산먼지 / 종료시간")),
            
            "장비": _parse_list(pick("장비")),
            "인력": _parse_list(pick("인력")),
            "차량": _parse_list(pick("차량")),
        }

    def to_min(hhmm: Optional[str]) -> Optional[int]:
        if not hhmm:
            return None
        m = re.match(r"^(\d{2}):(\d{2})$", str(hhmm))
        if not m:
            return None
        return int(m.group(1)) * 60 + int(m.group(2))

    def norm_end(s_min: Optional[int], e_min: Optional[int]) -> Optional[int]:
        if s_min is None or e_min is None:
            return None
        if e_min < s_min:
            return e_min + 24 * 60
        return e_min

    def overlap_minutes(a_s: int, a_e: int, b_s: int, b_e: int) -> int:
        s = max(a_s, b_s)
        e = min(a_e, b_e)
        return max(0, e - s)

    def make_dups(key: str) -> Tuple[pd.DataFrame, Set[str], Set[str]]:
        groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        for sn, info in by_sn.items():
            d = info.get("날짜")
            t = info.get("팀")
            if not d or not t:
                continue

            for r in info.get(key, []):
                if not r:
                    continue
                
                r_str = str(r).replace(" ", "").upper()
                s_str = info.get("채취시작")
                e_str = info.get("채취끝")
                
                if key == "장비":
                    if "대기배출가스" in r_str and "THC" not in r_str:
                        if not info.get("가스시작") and not info.get("가스끝"):
                            continue
                        s_str = info.get("가스시작")
                        e_str = info.get("가스끝")
                    elif "THC" in r_str:
                        if not info.get("THC시작") and not info.get("THC끝"):
                            continue
                        s_str = info.get("THC시작")
                        e_str = info.get("THC끝")
                    elif "비산먼지" in r_str:
                        if not info.get("비산먼지시작") and not info.get("비산먼지끝"):
                            continue
                        s_str = info.get("비산먼지시작")
                        e_str = info.get("비산먼지끝")

                s = to_min(s_str)
                e = norm_end(s, to_min(e_str))

                k = (d, str(r))
                groups.setdefault(k, []).append({
                    "시료번호": sn,
                    "팀": t,
                    "시작": s_str,
                    "끝": e_str,
                    "s_min": s,
                    "e_min": e,
                })

        rows: List[Dict[str, Any]] = []
        dup_samples: Set[str] = set()
        unk_samples: Set[str] = set()

        for (d, r), items in groups.items():
            teams = sorted({it["팀"] for it in items if it.get("팀")})
            if len(teams) < 2:
                continue

            items_sorted = sorted(items, key=lambda x: (x["s_min"] is None, x["s_min"] if x["s_min"] is not None else 10**9))
            n = len(items_sorted)
            for i in range(n):
                a = items_sorted[i]
                for j in range(i + 1, n):
                    b = items_sorted[j]
                    if a.get("팀") == b.get("팀"):
                        continue

                    a_s, a_e = a.get("s_min"), a.get("e_min")
                    b_s, b_e = b.get("s_min"), b.get("e_min")

                    # 시간 미기재
                    if a_s is None or a_e is None or b_s is None or b_e is None:
                        rows.append({
                            "날짜": d,
                            "이름": r,
                            "시료A": a.get("시료번호"),
                            "팀A": a.get("팀"),
                            "시작A": a.get("시작"),
                            "끝A": a.get("끝"),
                            "시료B": b.get("시료번호"),
                            "팀B": b.get("팀"),
                            "시작B": b.get("시작"),
                            "끝B": b.get("끝"),
                            "겹침분": None,
                            "판정": "확인필요(시간미기재)",
                        })
                        unk_samples.add(str(a.get("시료번호")))
                        unk_samples.add(str(b.get("시료번호")))
                        continue

                    # 시간 겹침 (NG)
                    if max(a_s, b_s) < min(a_e, b_e):
                        ov = overlap_minutes(a_s, a_e, b_s, b_e)
                        rows.append({
                            "날짜": d,
                            "이름": r,
                            "시료A": a.get("시료번호"),
                            "팀A": a.get("팀"),
                            "시작A": a.get("시작"),
                            "끝A": a.get("끝"),
                            "시료B": b.get("시료번호"),
                            "팀B": b.get("팀"),
                            "시작B": b.get("시작"),
                            "끝B": b.get("끝"),
                            "겹침분": ov,
                            "판정": "NG(시간겹침)",
                        })
                        dup_samples.add(str(a.get("시료번호")))
                        dup_samples.add(str(b.get("시료번호")))
                    
                    # [핵심 추가] 같은 날 썼지만 시간은 안 겹침 (OK)
                    else:
                        rows.append({
                            "날짜": d,
                            "이름": r,
                            "시료A": a.get("시료번호"),
                            "팀A": a.get("팀"),
                            "시작A": a.get("시작"),
                            "끝A": a.get("끝"),
                            "시료B": b.get("시료번호"),
                            "팀B": b.get("팀"),
                            "시작B": b.get("시작"),
                            "끝B": b.get("끝"),
                            "겹침분": 0,
                            "판정": "OK(정상교대)",
                        })
                        # 주의: OK 상황이므로 dup_samples(NG목록)에는 추가하지 않음!

        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values(by=["날짜", "이름", "판정"], ascending=[True, True, True]).reset_index(drop=True)
        return df, dup_samples, unk_samples

    eq_cases, eq_dup, eq_unk = make_dups("장비")
    person_cases, person_dup, person_unk = make_dups("인력")
    vehicle_cases, vehicle_dup, vehicle_unk = make_dups("차량")

    if not eq_cases.empty:
        eq_cases = eq_cases.rename(columns={"이름": "장비명"})
    if not person_cases.empty:
        person_cases = person_cases.rename(columns={"이름": "인력명"})
    if not vehicle_cases.empty:
        vehicle_cases = vehicle_cases.rename(columns={"이름": "차량"})

    return (eq_cases, person_cases, vehicle_cases,
            eq_dup, person_dup, vehicle_dup,
            eq_unk, person_unk, vehicle_unk)

def build_dashboard(send_df: pd.DataFrame,
                    review_sum: pd.DataFrame,
                    review_ng_causes: pd.DataFrame,
                    eq_cases: pd.DataFrame,
                    person_cases: pd.DataFrame,
                    vehicle_cases: pd.DataFrame,
                    eq_samples: Set[str],
                    person_samples: Set[str],
                    vehicle_samples: Set[str],
                    eq_unknown: Set[str],
                    person_unknown: Set[str],
                    vehicle_unknown: Set[str]) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame]]:
    samples1 = set(send_df["시료번호"].astype(str))
    samples2 = set(review_sum["시료번호"].astype(str))
    union = sorted(samples1 | samples2)

    detail = pd.DataFrame({"시료번호": union})

    keep_cols = [c for c in ["시료번호", "대장시트", "SN날짜(YYMMDD)", "대장시작", "대장끝",
                             "성적서상태", "수분상태", "THC상태", "발송_종합"] if c in send_df.columns]
    detail = detail.merge(send_df[keep_cols], on="시료번호", how="left")
    detail = detail.merge(review_sum, on="시료번호", how="left")

    # 중복 결과(검토파일이 있는 시료만 판단 가능)
    def _dup_flag(sn: str, dup_set: Set[str], unk_set: Set[str]) -> str:
        sn = str(sn)
        if sn in dup_set:
            return "NG"
        if sn in unk_set:
            return "확인필요"
        if sn in samples2:  # 검토파일에 존재하지만 중복 아님
            return "OK"
        return ""  # 검토파일이 없으면 판단 불가

    detail["장비중복"] = detail["시료번호"].apply(lambda x: _dup_flag(x, eq_samples, eq_unknown))
    detail["인력중복"] = detail["시료번호"].apply(lambda x: _dup_flag(x, person_samples, person_unknown))
    detail["차량중복"] = detail["시료번호"].apply(lambda x: _dup_flag(x, vehicle_samples, vehicle_unknown))
    detail["중복_종합"] = detail.apply(
        lambda r: "NG" if ("NG" in [r.get("장비중복"), r.get("인력중복"), r.get("차량중복")])
        else ("OK" if (r.get("장비중복") == "OK" and r.get("인력중복") == "OK" and r.get("차량중복") == "OK") else ""),
        axis=1
    )

    def final_overall(row):
        s1 = "WARN" if pd.isna(row.get("발송_종합")) else str(row.get("발송_종합"))
        s2 = "WARN" if pd.isna(row.get("검토_종합")) else str(row.get("검토_종합"))
        s3 = str(row.get("중복_종합")) if row.get("중복_종합") else "OK"  # 중복 판단 불가면 OK로 취급(최종은 따로 확인)
        if "NG" in (s1, s2, s3):
            return "NG"
        if "WARN" in (s1, s2):
            return "WARN"
        return "OK"

    detail["최종_종합"] = detail.apply(final_overall, axis=1)

    summary = pd.DataFrame([
        ["총 시료(Union)", len(union)],
        ["둘 다 존재(매칭)", len(samples1 & samples2)],
        ["발송대장만", len(samples1 - samples2)],
        ["검토요약만", len(samples2 - samples1)],
        ["최종 OK", int((detail["최종_종합"] == "OK").sum())],
        ["최종 WARN", int((detail["최종_종합"] == "WARN").sum())],
        ["최종 NG", int((detail["최종_종합"] == "NG").sum())],
        ["장비중복 NG", int((detail["장비중복"] == "NG").sum())],
        ["인력중복 NG", int((detail["인력중복"] == "NG").sum())],
        ["차량중복 NG", int((detail["차량중복"] == "NG").sum())],
    ], columns=["항목", "값"])

    
    # 원인집계 테이블들
    def _safe_series(df: pd.DataFrame, col: str) -> pd.Series:
        if col in df.columns:
            return df[col]
        return pd.Series([], dtype=str)

    def _vc_table(series: pd.Series, key_name: str) -> pd.DataFrame:
        # pandas 버전에 상관없이 [key_name, 건수] 컬럼으로 안정적으로 생성
        s = series.fillna("미기재").astype(str)
        return s.value_counts().rename_axis(key_name).reset_index(name="건수")

    causes: Dict[str, pd.DataFrame] = {
        "검토_NG항목": review_ng_causes,
        "THC상태": _vc_table(_safe_series(send_df, "THC상태"), "THC상태"),
        "수분상태": _vc_table(_safe_series(send_df, "수분상태"), "수분상태"),
        "성적서상태": _vc_table(_safe_series(send_df, "성적서상태"), "성적서상태"),
    }

    # 중복 원인(자원명별 건수 + 사례)
    if not eq_cases.empty:
        causes["장비중복(장비명별)"] = _vc_table(eq_cases["장비명"], "장비명")
        causes["장비중복(사례)"] = eq_cases
    else:
        causes["장비중복(장비명별)"] = pd.DataFrame(columns=["장비명", "건수"])
        causes["장비중복(사례)"] = pd.DataFrame(columns=["날짜", "장비명", "시료A", "팀A", "시작A", "끝A",
                                                       "시료B", "팀B", "시작B", "끝B", "겹침분", "판정"])

    if not person_cases.empty:
        causes["인력중복(인력명별)"] = _vc_table(person_cases["인력명"], "인력명")
        causes["인력중복(사례)"] = person_cases
    else:
        causes["인력중복(인력명별)"] = pd.DataFrame(columns=["인력명", "건수"])
        causes["인력중복(사례)"] = pd.DataFrame(columns=["날짜", "인력명", "시료A", "팀A", "시작A", "끝A",
                                                       "시료B", "팀B", "시작B", "끝B", "겹침분", "판정"])

    if not vehicle_cases.empty:
        causes["차량중복(차량별)"] = _vc_table(vehicle_cases["차량"], "차량")
        causes["차량중복(사례)"] = vehicle_cases
    else:
        causes["차량중복(차량별)"] = pd.DataFrame(columns=["차량", "건수"])
        causes["차량중복(사례)"] = pd.DataFrame(columns=["날짜", "차량", "시료A", "팀A", "시작A", "끝A",
                                                       "시료B", "팀B", "시작B", "끝B", "겹침분", "판정"])

    return summary, detail, causes



def write_dashboard_excel(out_path: str, summary_df: pd.DataFrame, detail_df: pd.DataFrame, causes: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name="대시보드", index=False)
        detail_df.to_excel(w, sheet_name="상세", index=False)
        # -------------------------------
        # 원인집계 시트(겹치지 않게 배치)
        #   - 상단: 기존 4블록(검토/THC/수분/성적서) 가로 배치
        #   - 하단: 장비/인력/차량 중복은 세로로 섹션 분리(사례표가 가로로 길어서 겹침 방지)
        # -------------------------------
        top_blocks = [
            ("검토_NG항목", causes["검토_NG항목"], 0),
            ("THC상태", causes["THC상태"], 4),
            ("수분상태", causes["수분상태"], 7),
            ("성적서상태", causes["성적서상태"], 10),
        ]
        for _name, _df, _sc in top_blocks:
            _df.to_excel(w, sheet_name="원인집계", index=False, startcol=_sc, startrow=0)

        ws = w.sheets["원인집계"]
        top_h = max((len(_df) + 1) for _, _df, _ in top_blocks)  # header 포함
        cur_row = top_h + 2  # (0-index) 아래로 2줄 띄우기

        def _write_section(title: str, df_count: pd.DataFrame, df_cases: pd.DataFrame, startrow: int) -> int:
            # 제목(1-index 셀)
            ws.cell(row=startrow + 1, column=1).value = f"[{title}]"

            # 표는 제목 아래로
            df_count.to_excel(w, sheet_name="원인집계", index=False, startcol=0, startrow=startrow + 1)
            startcol_cases = int(df_count.shape[1]) + 2
            df_cases.to_excel(w, sheet_name="원인집계", index=False, startcol=startcol_cases, startrow=startrow + 1)

            h1 = len(df_count) + 1
            h2 = len(df_cases) + 1
            return startrow + max(h1, h2) + 3  # 다음 섹션까지 2줄 띄움

        cur_row = _write_section("장비중복", causes["장비중복(장비명별)"], causes["장비중복(사례)"], cur_row)
        cur_row = _write_section("인력중복", causes["인력중복(인력명별)"], causes["인력중복(사례)"], cur_row)
        cur_row = _write_section("차량중복", causes["차량중복(차량별)"], causes["차량중복(사례)"], cur_row)

#--------------GUI-------------
class DashGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("발송대장/측정인 종합 검토")
        self.root.geometry("860x620")
        self.root.minsize(820, 560)
        set_window_topmost(self.root)

        self.review_paths: List[str] = []
        self.result_queue: "queue.Queue[Tuple[Any, ...]]" = queue.Queue()
        self._is_running = False

        self.var_send = tk.StringVar()
        self.var_review = tk.StringVar(value="0개 선택됨")
        self.status_var = tk.StringVar(value="대기 중")

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(150, self._poll_results)

    def _build_ui(self):
        outer = ttk.Frame(self.root, padding=15)
        outer.grid(row=0, column=0, sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(3, weight=1)

        file_frame = ttk.LabelFrame(outer, text="입력 파일", padding=10)
        file_frame.grid(row=0, column=0, sticky="ew")
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="대장검토 파일").grid(row=0, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(file_frame, textvariable=self.var_send, width=70).grid(row=0, column=1, sticky="ew", padx=5, pady=6)
        self.btn_pick_send = ttk.Button(file_frame, text="대장검토 파일 찾기", command=self._choose_send)
        self.btn_pick_send.grid(row=0, column=2, padx=5, pady=6)

        ttk.Label(file_frame, text="측정인검토 파일(여러개)").grid(row=1, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(file_frame, textvariable=self.var_review, width=70, state="readonly").grid(row=1, column=1, sticky="ew", padx=5, pady=6)
        self.btn_pick_review = ttk.Button(file_frame, text="측정인검토 파일 선택", command=self._choose_review)
        self.btn_pick_review.grid(row=1, column=2, padx=5, pady=6)

        action_frame = ttk.Frame(outer)
        action_frame.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        action_frame.columnconfigure(1, weight=1)

        self.btn_start = ttk.Button(action_frame, text="검사 시작", command=self._on_start)
        self.btn_start.grid(row=0, column=0, padx=(0, 8))

        ttk.Label(action_frame, text="상태").grid(row=0, column=1, sticky="e", padx=(12, 4))
        ttk.Label(action_frame, textvariable=self.status_var).grid(row=0, column=2, sticky="w")

        self.progress_bar = ttk.Progressbar(outer, mode="determinate", maximum=100)
        self.progress_bar.grid(row=2, column=0, sticky="ew", pady=(10, 0))

        log_frame = ttk.LabelFrame(outer, text="로그", padding=10)
        log_frame.grid(row=3, column=0, sticky="nsew", pady=(12, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_panel = LogPanel(log_frame, height=20)
        self.log_panel.grid(row=0, column=0, sticky="nsew")
        self.log_panel.start_pumping()

    def _on_close(self):
        try:
            self.log_panel.restore_stdio()
        finally:
            self.root.destroy()

    def _choose_send(self):
        path = askopenfilename(
            title="발송대장 검토 파일 찾기",
            initialdir=DEFAULT_SEND_REVIEW_DIR,
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self.var_send.set(path)
        print(f"[선택] 대장검토 파일: {os.path.basename(path)}")

    def _choose_review(self):
        paths = askopenfilenames(
            title="측정인 검토 파일 선택(팀별 여러 개 선택 가능)",
            initialdir=DEFAULT_MEASUREMENT_REVIEW_DIR,
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls"), ("All files", "*.*")]
        )
        if not paths:
            return
        self.review_paths = list(paths)
        self.var_review.set(f"{len(self.review_paths)}개 선택됨")
        print(f"[선택] 측정인 검토 파일 {len(self.review_paths)}개")
        for path in self.review_paths:
            print(f"  - {os.path.basename(path)}")

    def _set_running(self, running: bool):
        self._is_running = running
        state = "disabled" if running else "normal"
        self.btn_pick_send.config(state=state)
        self.btn_pick_review.config(state=state)
        self.btn_start.config(state=state, text="실행 중..." if running else "검사 시작")

    def _on_start(self):
        send_path = self.var_send.get().strip()
        review_paths = list(self.review_paths)

        if not send_path:
            showerror("오류", "발송대장 검토 파일을 선택하세요.")
            return
        if not review_paths:
            showerror("오류", "측정인 검토 파일을 1개 이상 선택하세요.")
            return
        if self._is_running:
            return

        self._set_running(True)
        self.status_var.set("시작 준비 중...")
        self.progress_bar["value"] = 0

        print("\n" + "=" * 70)
        print("[시작] 종합 검토 생성")
        print(f"[입력] 대장검토 파일: {send_path}")
        print(f"[입력] 측정인 검토 파일 수: {len(review_paths)}")

        thread = threading.Thread(target=self._worker, args=(send_path, review_paths), daemon=True)
        thread.start()

    def _post_progress(self, message: str, percent: int):
        self.result_queue.put(("PROG", message, percent))

    def _worker(self, send_path: str, review_paths: List[str]):
        try:
            self._post_progress("1/7 발송대장 읽는 중...", 10)
            print("[1/7] 발송대장 읽는 중...")
            send_df = read_send_report(send_path)

            self._post_progress("2/7 검토 요약 병합 중...", 25)
            print("[2/7] 검토 요약 병합 중...")
            df_norm, review_sum, review_ng_causes = read_review_summary_multi(list(review_paths))

            self._post_progress("3/7 중복 테이블 생성 중...", 45)
            print("[3/7] 중복(장비/인력/차량) 테이블 생성 중...")
            (
                eq_cases,
                person_cases,
                vehicle_cases,
                eq_samples,
                person_samples,
                vehicle_samples,
                eq_unknown,
                person_unknown,
                vehicle_unknown,
            ) = build_dup_tables(df_norm)

            self._post_progress("4/7 대시보드 데이터 생성 중...", 65)
            print("[4/7] 대시보드 데이터 생성 중...")
            summary_df, detail_df, causes = build_dashboard(
                send_df, review_sum, review_ng_causes,
                eq_cases, person_cases, vehicle_cases,
                eq_samples, person_samples, vehicle_samples,
                eq_unknown, person_unknown, vehicle_unknown,
            )

            self._post_progress("5/7 저장 파일명 결정 중...", 78)
            print("[5/7] 저장 경로/파일명 결정 중...")
            date_tag = _best_date_tag(send_df, list(review_paths))
            try:
                os.makedirs(DEFAULT_OUTPUT_DIR, exist_ok=True)
            except Exception:
                pass

            auto_path = os.path.join(DEFAULT_OUTPUT_DIR, f"{date_tag} 종합 검토.xlsx")
            out_path = _unique_path(auto_path)

            self._post_progress("6/7 엑셀 파일 저장 중...", 92)
            print(f"[6/7] 엑셀 파일 저장 중... -> {out_path}")
            try:
                write_dashboard_excel(out_path, summary_df, detail_df, causes)
            except Exception as save_err:
                print(f"[경고] 자동 저장 실패: {save_err}")
                self.result_queue.put(("SAVE_FALLBACK", str(save_err), out_path, summary_df, detail_df, causes))
                return

            self.result_queue.put(("OK", out_path))
        except Exception as exc:
            import traceback

            traceback.print_exc()
            self.result_queue.put(("ERR", str(exc)))

    def _handle_save_fallback(self, save_err: str, default_path: str,
                              summary_df: pd.DataFrame, detail_df: pd.DataFrame,
                              causes: Dict[str, pd.DataFrame]):
        self.status_var.set("자동 저장 실패 - 저장 위치 선택 필요")
        self.progress_bar["value"] = 92

        fallback = asksaveasfilename(
            title="종합 검토 저장 위치/파일명 선택 (자동 저장 실패)",
            initialdir=DEFAULT_OUTPUT_DIR,
            initialfile=os.path.basename(default_path),
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not fallback:
            print("[중단] 저장 위치 선택이 취소되어 작업을 마쳤습니다.")
            self.status_var.set("저장 취소")
            self._set_running(False)
            return

        try:
            print(f"[재시도] 수동 저장: {fallback}")
            write_dashboard_excel(fallback, summary_df, detail_df, causes)
        except Exception as exc:
            import traceback

            traceback.print_exc()
            showerror("오류", str(exc))
            self.status_var.set("오류")
            self._set_running(False)
            return

        print(f"[완료] 종합 검토 저장 완료: {fallback}")
        print("[안내] 창은 닫히지 않습니다. 파일만 다시 선택해서 이어서 실행하면 됩니다.")
        self.status_var.set("완료")
        self.progress_bar["value"] = 100
        self._set_running(False)

    def _poll_results(self):
        try:
            while True:
                item = self.result_queue.get_nowait()
                kind = item[0]

                if kind == "PROG":
                    _, message, percent = item
                    self.status_var.set(message)
                    self.progress_bar["value"] = int(percent)
                    continue

                if kind == "SAVE_FALLBACK":
                    _, save_err, default_path, summary_df, detail_df, causes = item
                    self._handle_save_fallback(save_err, default_path, summary_df, detail_df, causes)
                    continue

                if kind == "OK":
                    _, out_path = item
                    print(f"[완료] 종합 검토 저장 완료: {out_path}")
                    print("[안내] 창은 닫히지 않습니다. 파일만 다시 선택해서 이어서 실행하면 됩니다.")
                    self.status_var.set("완료")
                    self.progress_bar["value"] = 100
                    self._set_running(False)
                    continue

                if kind == "ERR":
                    _, message = item
                    print(f"[오류] {message}")
                    self.status_var.set("오류")
                    self._set_running(False)
                    showerror("오류", message)
        except queue.Empty:
            pass

        if self.root.winfo_exists():
            self.root.after(150, self._poll_results)

    def run(self):
        self.root.mainloop()



def main():
    app = DashGUI()
    app.run()



if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log_error("dash.main", e)
        raise