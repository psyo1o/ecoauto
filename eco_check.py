# -*- coding: utf-8 -*-
"""
측정인.kr 자동 비교 시스템 - FINAL + 파일명 자동 생성 + PDF 다운로드/하이퍼링크 + NG 빨간색 표시
"""
import warnings
warnings.filterwarnings("ignore")  # 무조건 모든 경고 차단 (조건 없음)
warnings.showwarning = lambda *args, **kwargs: None
import os
import re
import time
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# ============================================================
# 공통 유틸 모듈 (모듈화)
# ============================================================
from selenium_utils import safe_click, wait_el as wait_until_exists, close_popup, set_date_js, wait
from format_utils import format_time as trim_time_to_hm, to_float1, to_float2
from data_utils import norm_ymd, sample_to_datestr, clean_leading_mark
from excel_utils import find_sheet_by_candidates as _find_sheet_by_candidates_openpyxl
from file_utils import find_excel_for_sample as _find_excel_util
from measin_utils import (
    login, search_date, wait_grid_loaded, get_samples_current_page,
    open_sample_detail, go_back_to_list, collect_samples_from_files,
    LOGIN_URL, FIELD_URL, NAS_BASE, NAS_DIRS
)
from excel_utils import find_sheet_by_candidates, parse_measuring_record
from realgrid_utils import rg_api_read_data
from log_utils import log_error
from measin_constants import (
    SKIP_VOL_AND_SPEED, SKIP_SPEED_ONLY, DUST_SKIP_FIELDS,
    SM3_ITEMS as sm3_items,
    SEL_DATE, SEL_START_TIME, SEL_END_TIME,
    SEL_O2_STD, SEL_O2_MEAS, SEL_GAS_VOL_PRE, SEL_GAS_VOL_POST,
    SEL_MOISTURE, SEL_GAS_TEMP, SEL_GAS_SPEED
)

# ------------------------------------------------------------
# 설정
# ------------------------------------------------------------


PDF_DIR = r"\\192.168.10.163\측정팀\10.검토\3.측정인 검토\PDF"
if not os.path.isdir(PDF_DIR):
    os.makedirs(PDF_DIR)

PDF_MAP = {}
COMPANY_MAP = {}   # ★ 추가: 시료번호 -> 업소명(표시용)


# ============================================================
# RealGrid 비교 예외 → measin_constants.py 에서 import 완료
# ============================================================

# ------------------------------------------------------------
# 공통 유틸 (wait → selenium_utils, clean_leading_mark → data_utils)
# ------------------------------------------------------------


def init_driver():
    """eco_check 전용 드라이버 초기화 (PDF 다운로드 경로 설정 포함)"""
    from selenium_utils import init_driver as _base_init
    d = _base_init()
    # PDF 다운로드 경로를 CDP로 설정
    try:
        d.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": PDF_DIR}
        )
    except:
        pass
    return d


def _wait_new_pdf(download_dir, before_set, timeout=60):
    """
    클릭 직전(before_set) 대비 새로 생긴 PDF를 찾아서
    .crdownload가 사라지고 완성된 파일만 반환
    """
    t0 = time.time()

    while time.time() - t0 < timeout:
        now = set(os.listdir(download_dir))

        # 새로 생긴 파일 후보
        added = list(now - before_set)

        # 크롬 다운로드 진행중이면 .crdownload 존재
        crs = [f for f in added if f.lower().endswith(".crdownload")]
        pdfs = [f for f in added if f.lower().endswith(".pdf")]

        if crs:
            time.sleep(0.3)
            continue

        if pdfs:
            paths = [os.path.join(download_dir, f) for f in pdfs]
            latest = max(paths, key=os.path.getmtime)
            # 파일 쓰기 마무리 안정화
            time.sleep(0.5)
            return latest

        time.sleep(0.3)

    return ""


# ------------------------------------------------------------
# PDF 다운로드
# ------------------------------------------------------------
def download_pdf(driver, sample_no):
    """
    PDF 다운로드 버튼 클릭 후
    '이번 클릭으로 새로 생성된 PDF'만 잡아서 sample_no.pdf로 저장
    """
    print(f"   [PDF] 다운로드 시도: {sample_no}")

    target_path = os.path.join(PDF_DIR, f"{sample_no}.pdf")

    # 기존 타겟 파일 있으면 삭제(있어도 상관 없지만 깔끔하게)
    if os.path.isfile(target_path):
        try:
            os.remove(target_path)
        except:
            pass

    # ✅ 1) 클릭 직전 폴더 상태 스냅샷
    try:
        before = set(os.listdir(PDF_DIR))
    except Exception as e:
        print(f"   ❌ PDF_DIR 접근 실패: {e}")
        return ""

    # ✅ 2) PDF 버튼 클릭
    if not safe_click(driver, "#fileArea > section > div > div.row.fr > input:nth-child(3)"):
        print("   ❌ PDF 버튼 클릭 실패")
        return ""

    # ✅ 3) 새로 생긴 PDF만 대기해서 잡기
    new_pdf = _wait_new_pdf(PDF_DIR, before, timeout=60)
    if not new_pdf or not os.path.isfile(new_pdf):
        print("   ❌ PDF 다운로드 완료/파일 탐지 실패")
        return ""

    # ✅ 4) 최종 파일명으로 rename (동일 디스크면 replace가 제일 안전)
    try:
        if os.path.abspath(new_pdf) != os.path.abspath(target_path):
            os.replace(new_pdf, target_path)
        print(f"   ✔ PDF 저장: {target_path}")
        return target_path
    except Exception as e:
        print(f"   ❌ PDF 이름 변경 실패: {e}")
        # 이름 변경 실패해도 실제 받은 파일 경로라도 리턴
        return new_pdf


# ------------------------------------------------------------
# 사이트 데이터 수집
# ------------------------------------------------------------
def gv(driver, selector):
    try:
        el = driver.find_element(By.CSS_SELECTOR, selector)
        v = el.get_attribute("value")
        if not v:
            v = el.text
        return (v or "").strip()
    except:
        return ""


def click_tab(driver, tab_id):
    try:
        el = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, f"a#{tab_id}"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", el)
        time.sleep(0.2)
        driver.execute_script("arguments[0].click();", el)

        if tab_id == "ui-id-1":
            wait_until_exists(driver, "#machineDiv", timeout=10)
        elif tab_id == "ui-id-2":
            wait_until_exists(driver, "#meas_start_time", timeout=10)
        elif tab_id == "ui-id-3":
            wait_until_exists(driver, "td#officer_dt", timeout=10)
        time.sleep(0.3)
    except:
        print(" ❌ 탭 전환 실패:", tab_id)


def get_wind_direction_text(driver):
    try:
        s = driver.find_element(
            By.CSS_SELECTOR,
            "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 "
            "> table > tbody > tr > td:nth-child(5) > select"
        )
        v = s.get_attribute("value")
        op = s.find_element(By.CSS_SELECTOR, f"option[value='{v}']")
        return op.text.strip()
    except:
        return ""


_DT_RE = re.compile(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}")

def _extract_time(txt: str) -> str:
    """'촬영일시: 2026-01-09 08:35' 같은 텍스트에서 '2026-01-09 08:35'만 뽑기"""
    if not txt:
        return ""
    m = _DT_RE.search(txt)
    if m:
        return m.group(0)
    # fallback
    if ":" in txt:
        return txt.split(":", 1)[1].strip()
    return txt.strip()

def get_mobile_times(driver):
    result = {
        "환경기술인입력일시": "",
        "GPS위치확인일시": "",
        "촬영일시목록": []
    }

    # ------------------------------------------------------------
    # 1) 환경기술인 입력일시: common-work-info 영역의 "입력일시"만
    # ------------------------------------------------------------
    try:
        env_els = driver.find_elements(
            By.XPATH,
            "//div[contains(@class,'common-work-info')]"
            "//td[@id='officer_dt' and contains(normalize-space(.),'입력일시')]"
        )
        if env_els:
            result["환경기술인입력일시"] = _extract_time(env_els[0].text.strip())
    except:
        pass

    # ------------------------------------------------------------
    # 2) 촬영일시 목록: pic_area 영역의 "촬영일시"만 전부 수집
    #    (pic_area 안에 입력일시가 있어도 무시됨)
    # ------------------------------------------------------------
    try:
        photo_els = driver.find_elements(
            By.XPATH,
            "//*[@id='pic_area']"
            "//td[@id='officer_dt' and contains(normalize-space(.),'촬영일시')]"
        )
        for el in photo_els:
            t = _extract_time(el.text.strip())
            if t:
                result["촬영일시목록"].append(t)
    except:
        pass

    # ------------------------------------------------------------
    # 3) GPS 위치확인일시: 기존 방식 유지
    # ------------------------------------------------------------
    try:
        gps = driver.find_element(By.CSS_SELECTOR, "td#gps_dt")
        tx = gps.text.strip()
        result["GPS위치확인일시"] = _extract_time(tx)
    except:
        pass

    return result

def _norm_company_key(s: str) -> str:
    """업소명 매칭용 간단 정규화(공백/법인표기 제거)."""
    if s is None:
        return ""
    t = str(s).strip()
    t = t.replace("(주)", "").replace("㈜", "").replace("주식회사", "")
    t = re.sub(r"\s+", "", t)
    return t

def relax_env_input_time_by_company(sample_rows_map: dict, excel_meta_map: dict):
    """
    동일 날짜 + 동일 업소 방문 케이스에서
    환경기술인 입력일시가 '해당 시료' 채취시간을 벗어나도
    같은 업소의 다른 시료 채취시간 범위 안이면 OK로 완화한다.
    """
    from collections import defaultdict

    windows = defaultdict(list)  # (date, compkey) -> [(start_dt, end_dt), ...]

    for sn, meta in (excel_meta_map or {}).items():
        if not isinstance(meta, dict):
            continue
        date = (meta.get("날짜") or "").strip()
        comp = _norm_company_key(meta.get("업소명", ""))
        st = _pd(meta.get("측정시작DT", ""))
        ed = _pd(meta.get("측정종료DT", ""))
        if date and comp and st and ed:
            windows[(date, comp)].append((st, ed))

    if not windows:
        return

    for sn, rows in (sample_rows_map or {}).items():
        meta = (excel_meta_map or {}).get(sn, {})
        if not isinstance(meta, dict):
            continue

        date = (meta.get("날짜") or "").strip()
        comp = _norm_company_key(meta.get("업소명", ""))
        if not date or not comp:
            continue

        key = (date, comp)
        if key not in windows:
            continue

        for r in rows:
            if not isinstance(r, dict):
                continue
            if r.get("항목") != "환경기술인입력일시":
                continue
            if r.get("비교") != "NG":
                continue

            dt = _pd(r.get("사이트값", ""))
            if not dt:
                continue

            if any(st <= dt <= ed for st, ed in windows[key]):
                r["비교"] = "OK"
                r["사이트만존재"] = ""


def read_site_data(driver, sample_no):
    data = {}

    # ------------------------------------------------------------
    # ① 탭1 데이터 수집 (장비 / 차량 / 인력 / 측정항목)
    # ------------------------------------------------------------
    click_tab(driver, "ui-id-1")
    time.sleep(1.5)
    try:
        els = driver.find_elements(By.CSS_SELECTOR,
                                   "#machineDiv > div > span > span.selection > span > ul > li")
        data["장비"] = [clean_leading_mark(e.text) for e in els if e.text.strip()]
    except:
        data["장비"] = []

    try:
        els = driver.find_elements(By.CSS_SELECTOR,
                                   "#carSection > div > span > span.selection > span > ul > li")
        data["차량"] = [clean_leading_mark(e.text) for e in els if e.text.strip()]
    except:
        data["차량"] = []

    try:
        els = driver.find_elements(
            By.CSS_SELECTOR,
            "#wid-id-4 > div > div.widget-body.no-padding > div > fieldset "
            "> div.row.input-full > section:nth-child(2) "
            "> span > span.selection > span > ul > li"
        )
        data["인력"] = [clean_leading_mark(e.text) for e in els if e.text.strip()]
    except:
        data["인력"] = []

    try:
        els = driver.find_elements(By.CSS_SELECTOR,
                                   "#inairTargetItem > div:nth-child(2) > div > span > span.selection > span > ul > li")
        arr = []
        for e in els:
            t = clean_leading_mark(e.text.strip())
            if t:
                arr.append(t)
        data["측정항목"] = arr
    except:
        data["측정항목"] = []

    # ------------------------------------------------------------
    # ② 탭2 데이터 수집 (채취시간/기상/산소농도 등)
    # ------------------------------------------------------------
    click_tab(driver, "ui-id-2")
    time.sleep(1)
    data["날짜"] = norm_ymd(gv(driver, SEL_DATE))
    data["기상"] = gv(driver,
        "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 > table > tbody > tr > td:nth-child(1) > input")
    data["기온"] = gv(driver,
        "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 > table > tbody > tr > td:nth-child(2) > input")
    data["습도"] = gv(driver,
        "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 > table > tbody > tr > td:nth-child(3) > input")
    data["기압"] = gv(driver,
        "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 > table > tbody > tr > td:nth-child(4) > input")

    data["풍향"] = get_wind_direction_text(driver)
    data["풍속"] = to_float1(gv(driver,
        "#idWHArea > div > div:nth-child(2) > fieldset > label.col.col-12 > table > tbody > tr > td:nth-child(6) > input"))

    data["채취시작"] = trim_time_to_hm(gv(driver, SEL_START_TIME))
    data["채취끝"] = trim_time_to_hm(gv(driver, SEL_END_TIME))

    data["표준산소농도"] = to_float1(gv(driver, SEL_O2_STD))
    data["실측산소농도"] = to_float1(gv(driver, SEL_O2_MEAS))
    data["배출가스유량전"] = to_float1(gv(driver, SEL_GAS_VOL_PRE))
    data["배출가스유량후"] = to_float1(gv(driver, SEL_GAS_VOL_POST))
    data["수분량"] = gv(driver, SEL_MOISTURE)
    data["배출가스온도"] = gv(driver, SEL_GAS_TEMP)
    data["배출가스유속"] = to_float2(gv(driver, SEL_GAS_SPEED))

    # ------------------------------------------------------------
    # ③ 탭2에서 PDF 다운로드 실행
    # ------------------------------------------------------------

    try:
        data["PDF경로"] = download_pdf(driver, sample_no)
    except Exception as e:
        print("⚠ PDF 다운로드 실패:", e)
        data["PDF경로"] = ""

    # ------------------------------------------------------------
    # ④ 탭3 데이터 수집 (모바일 입력/GPS/촬영일시)
    # ------------------------------------------------------------
    click_tab(driver, "ui-id-3")
    time.sleep(2)
    mob = get_mobile_times(driver)
    data["환경기술인입력일시"] = mob["환경기술인입력일시"]
    data["GPS위치확인일시"] = mob["GPS위치확인일시"]
    data["촬영일시목록"] = mob["촬영일시목록"]

    return data


# ------------------------------------------------------------
# NAS 검색 / 엑셀 읽기 / 비교 / 저장 (원본 그대로 유지)
# ------------------------------------------------------------
def find_excel_for_sample(sample_no):
    """file_utils.find_excel_for_sample 래퍼 (eco_check 호환)"""
    result = _find_excel_util(sample_no, nas_base=NAS_BASE, nas_dirs=NAS_DIRS, strict=True)
    if not result:
        print(" ❌ 엑셀 없음:", sample_no)
    return result


def get_team_no_from_sample(sample_no):
    try:
        return sample_no[7]
    except:
        return ""




def parse_team_input(s: str):
    """팀 입력 문자열을 팀번호 리스트로 파싱.
    허용 예)
      - "" / 공백 : 전체
      - "3"
      - "1,3,5" / "1 3 5"
      - "1-3" (범위)
    반환: ["1","3","5"] 처럼 문자열 리스트(중복 제거, 정렬)
    """
    if s is None:
        return []
    s = str(s).strip()
    if not s:
        return []
    s = s.replace(" ", ",")
    parts = [p.strip() for p in s.split(",") if p.strip()]
    teams = set()
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            if a.strip().isdigit() and b.strip().isdigit():
                a_i, b_i = int(a), int(b)
                lo, hi = (a_i, b_i) if a_i <= b_i else (b_i, a_i)
                for t in range(lo, hi + 1):
                    if 1 <= t <= 5:
                        teams.add(str(t))
            continue
        if p.isdigit():
            t = int(p)
            if 1 <= t <= 5:
                teams.add(str(t))
    return sorted(teams, key=lambda x: int(x))



# ------------------------------------------------------------
# 비교 관련 로직 (원본 유지)
# ------------------------------------------------------------
SIMPLE_FIELDS = [
    "날짜",
    "기상", "기온", "습도", "기압", "풍향", "풍속",
    "표준산소농도", "실측산소농도",
    "배출가스유량전", "배출가스유량후",
    "수분량", "배출가스온도", "배출가스유속",
    "채취시작", "채취끝",
]




# =====================================================================
# 탭2 RealGrid(측정항목별 테이블) 읽기 + 성적서(엑셀) 기대값 생성 + 비교
#  - eco_check: "수기 입력 오타" 잡는 용도
# =====================================================================

REALGRID_ROOT_CSS = "#measGridAnalySampAnzeDataAirItemList1"

def _rg_norm_date(v):
    if v is None:
        return ""
    s = str(v).strip()
    # '2026-01-02' 형태로 들어오면 그대로, '2026.01.02' 등은 치환
    s = s.replace(".", "-").replace("/", "-")
    # 'YYYY-MM-DD HH:MM' 같이 오면 날짜만
    if " " in s:
        s = s.split(" ")[0]
    return s

def _rg_norm_time(v):
    if v is None:
        return ""
    s = str(v).strip()
    # 'HH:MM:SS' -> 'HH:MM'
    if ":" in s:
        parts = s.split(":")
        if len(parts) >= 2:
            return f"{parts[0]}:{parts[1]}"
    return s

def _rg_norm_num(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    s = s.replace(",", "")
    try:
        # 1.0 / 1.00 같은 표현을 통일
        f = float(s)
        # 정수면 정수 문자열로
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
        # 소수는 불필요한 0 제거
        out = f"{f:.10f}".rstrip("0").rstrip(".")
        return out
    except:
        return s

def _rg_norm_vol_text(v):
    """시료채취량: 소수점 4자리로 반올림 통일 (사이트·엑셀 비교용)"""
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    s = s.replace(",", "")
    try:
        return f"{float(s):.4f}"
    except ValueError:
        return s



def _spd_unit_equiv_for_compare(sv: str, ev: str) -> bool:
    """흡인속도 단위 비교 예외: L-MIN 과 L/min 은 동일로 취급(표시는 그대로)."""
    s = "" if sv is None else str(sv).strip()
    e = "" if ev is None else str(ev).strip()
    if not s or not e:
        return False
    # 'L-MIN'만 예외 허용 (대소문자 무시). 다른 변형(L/MIN 등)은 건드리지 않음.
    if s.upper() == "L-MIN" and e.lower() == "l/min":
        return True
    if e.upper() == "L-MIN" and s.lower() == "l/min":
        return True
    return False


def build_excel_realgird_expected(excel_path: str, sample_no: str, is_dust: bool) -> dict:
    """
    성적서 엑셀에서 탭2 RealGrid에 들어가야 하는 값(기대값) 생성.
    - 일반: 입력(분석값) 시트의 헤더(측정시작/측정 종료/시료흡인속도/시료채취량) 기반
    """
    wb = load_workbook(excel_path, data_only=True)

    ws = _find_sheet_by_candidates_openpyxl(wb, ["입력(분석값)", "입력"])
    if ws is None:
        raise RuntimeError("엑셀에서 '입력(분석값)' 시트를 찾지 못함")

    # 1행 헤더 → 열번호 매핑
    header_map = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if not v:
            continue
        t = str(v).strip()
        header_map[t] = col

    def col_of(*names):
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_start = col_of("측정시작", "측정 시작")
    c_end   = col_of("측정 종료", "측정종료")
    c_spd   = col_of("시료흡인속도", "시료 흡인속도")
    c_vol   = col_of("시료채취량", "시료 채취량")

    if not (c_start and c_end and c_spd and c_vol):
        raise RuntimeError(f"입력(분석값) 1행 헤더 매칭 실패: start={c_start}, end={c_end}, spd={c_spd}, vol={c_vol} / 헤더={list(header_map.keys())}")

    date_str = sample_to_datestr(sample_no) or ""

    out = {}
    for r in range(2, 65):

        # ✅ A열이 빈칸이면 "측정 안함" → excel_rg에서 제외
        a_flag = ws.cell(row=r, column=1).value  # A열
        if a_flag is None or str(a_flag).strip() == "":
            continue

        item = ws.cell(row=r, column=2).value  # B열(측정항목)
        if not item:
            continue
        item = str(item).strip()
        if not item:
            continue

        st = _rg_norm_time(ws.cell(row=r, column=c_start).value)
        et = _rg_norm_time(ws.cell(row=r, column=c_end).value)
        spd = _rg_norm_num(ws.cell(row=r, column=c_spd).value)
        vol = _rg_norm_vol_text(ws.cell(row=r, column=c_vol).value)

        vol_unit = "Sm³" if item in sm3_items else "L"

        out[item] = {
            "sd": date_str, "st": st,
            "ed": date_str, "et": et,
            "vol": vol, "vol_u": vol_unit,
            "spd": spd, "spd_u": "L/min",
        }

    return out

def build_realgird_compare_rows(sample_no: str, site_rg: dict, excel_rg: dict) -> list:
    """
    RealGrid 비교 결과를 eco_check의 결과 row 포맷으로 반환.
    SKIP_SPEED_ONLY: 흡인속도/단위 비교 PASS
    SKIP_VOL_AND_SPEED: 시료채취량/단위 + 흡인속도/단위 비교 PASS
    """
    rows = []

    site_items = set(site_rg.keys()) if isinstance(site_rg, dict) else set()
    excel_items = set(excel_rg.keys()) if isinstance(excel_rg, dict) else set()
    all_items = sorted(site_items | excel_items)

    fields = [
        ("sd", "측정일(시작)", _rg_norm_date),
        ("st", "시작시간", _rg_norm_time),
        ("ed", "측정일(종료)", _rg_norm_date),
        ("et", "종료시간", _rg_norm_time),
        ("vol", "시료채취량", _rg_norm_vol_text),
        ("vol_u", "채취량단위", lambda x: "" if x is None else str(x).strip()),
        ("spd", "흡인속도", _rg_norm_num),
        ("spd_u", "흡인속도단위", lambda x: "" if x is None else str(x).strip()),
    ]

    for item in all_items:
        # ✅ 항상 먼저 정의 (NameError 방지)
        item_name = ("" if item is None else str(item)).strip()

        s = site_rg.get(item) if item in site_items else None
        e = excel_rg.get(item) if item in excel_items else None

        if s is None:
            rows.append({
                "sample": sample_no,
                "항목": f"[RealGrid] {item_name} (사이트에 없음)",
                "사이트값": "",
                "엑셀값": "존재",
                "비교": "NG",
                "사이트만존재": "",
                "엑셀만존재": "O"
            })
            continue

        if e is None:
            rows.append({
                "sample": sample_no,
                "항목": f"[RealGrid] {item_name} (엑셀에 없음)",
                "사이트값": "존재",
                "엑셀값": "",
                "비교": "NG",
                "사이트만존재": "O",
                "엑셀만존재": ""
            })
            continue

        def is_blank(v):
            return v is None or str(v).strip() == ""

        # ✅ 스킵 규칙
        skip_speed = (item_name in SKIP_SPEED_ONLY)
        skip_vol_and_speed = (item_name in SKIP_VOL_AND_SPEED)

        for key, label, norm_fn in fields:

            # SKIP_SPEED_ONLY → 흡인속도/단위 PASS (단, 사이트값 있으면 NG)
            if skip_speed and key in ("spd", "spd_u"):
                raw = s.get(key) if isinstance(s, dict) else ""
                try:
                    sv = norm_fn(raw)
                except Exception:
                    sv = raw

                if not is_blank(sv):
                    rows.append({
                        "sample": sample_no,
                        "항목": f"[RealGrid] {item_name} / {label} (예외항목: 사이트는 빈칸이어야 함)",
                        "사이트값": sv,
                        "엑셀값": "",
                        "비교": "NG",
                        "사이트만존재": sv,
                        "엑셀만존재": ""
                    })
                continue

            # SKIP_VOL_AND_SPEED → 채취량/단위 + 흡인속도/단위 PASS (단, 사이트값 있으면 NG)
            if skip_vol_and_speed and key in ("vol", "vol_u", "spd", "spd_u"):
                raw = s.get(key) if isinstance(s, dict) else ""
                try:
                    sv = norm_fn(raw)
                except Exception:
                    sv = raw

                if not is_blank(sv):
                    rows.append({
                        "sample": sample_no,
                        "항목": f"[RealGrid] {item_name} / {label} (예외항목: 사이트는 빈칸이어야 함)",
                        "사이트값": sv,
                        "엑셀값": "",
                        "비교": "NG",
                        "사이트만존재": sv,
                        "엑셀만존재": ""
                    })
                continue


            sv = norm_fn(s.get(key)) if isinstance(s, dict) else ""
            ev = norm_fn(e.get(key)) if isinstance(e, dict) else ""
            ok = (sv == ev) or (key == "spd_u" and _spd_unit_equiv_for_compare(sv, ev))

            rows.append({
                "sample": sample_no,
                "항목": f"[RealGrid] {item_name} / {label}",
                "사이트값": sv,
                "엑셀값": ev,
                "비교": "OK" if ok else "NG",
                "사이트만존재": "",
                "엑셀만존재": ""
            })

    return rows

def compare_scalar(sample, field, site_val, excel_val):
    return {
        "sample": sample,
        "항목": field,
        "사이트값": site_val,
        "엑셀값": excel_val,
        "비교": "OK" if str(site_val) == str(excel_val) else "NG",
        "사이트만존재": "",
        "엑셀만존재": "",
    }


def compare_list(sample, field, site_list, excel_list):
    s = set([x.strip() for x in site_list if x.strip()])
    e = set([x.strip() for x in excel_list if x.strip()])

    only_s = sorted(list(s - e))
    only_e = sorted(list(e - s))

    return {
        "sample": sample,
        "항목": field,
        "사이트값": ", ".join(sorted(s)),
        "엑셀값": ", ".join(sorted(e)),
        "비교": "OK" if not only_s and not only_e else "NG",
        "사이트만존재": ", ".join(only_s),
        "엑셀만존재": ", ".join(only_e),
    }


def _pd(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None


def compare_mobile_single(sample, label, t, es, ee):
    if not t:
        return {
            "sample": sample, "항목": label,
            "사이트값": "", "엑셀값": f"{es} ~ {ee}",
            "비교": "NG", "사이트만존재": "", "엑셀만존재": ""
        }

    dt = _pd(t)
    st = _pd(es)
    ed = _pd(ee)
    if not dt or not st or not ed:
        ok = "확인불가"
    else:
        ok = "OK" if st <= dt <= ed else "NG"

    return {
        "sample": sample,
        "항목": label,
        "사이트값": t,
        "엑셀값": f"{es} ~ {ee}",
        "비교": ok,
        "사이트만존재": "" if ok == "OK" else t,
        "엑셀만존재": "",
    }


def compare_mobile_photos(sample, arr, es, ee):
    if not arr:
        return {
            "sample": sample, "항목": "사진촬영일시(전체)",
            "사이트값": "", "엑셀값": f"{es} ~ {ee}",
            "비교": "NG", "사이트만존재": "", "엑셀만존재": ""
        }

    st = _pd(es)
    ed = _pd(ee)
    if not st or not ed:
        return {
            "sample": sample,
            "항목": "사진촬영일시(전체)",
            "사이트값": ", ".join(arr),
            "엑셀값": f"{es} ~ {ee}",
            "비교": "확인불가",
            "사이트만존재": "",
            "엑셀만존재": "",
        }

    out_range = []
    for t in arr:
        dt = _pd(t)
        if not dt or dt < st or dt > ed:
            out_range.append(t)

    return {
        "sample": sample,
        "항목": "사진촬영일시(전체)",
        "사이트값": ", ".join(arr),
        "엑셀값": f"{es} ~ {ee}",
        "비교": "OK" if not out_range else "NG",
        "사이트만존재": ", ".join(out_range),
        "엑셀만존재": "",
    }


def build_comparison_rows(sample_no, site, excel):
    rows = []

    rows.append({
        "sample": sample_no,
        "항목": "엑셀 시료번호 일치 여부",
        "사이트값": sample_no,
        "엑셀값": excel.get("엑셀시료번호", ""),
        "비교": "OK" if sample_no == excel.get("엑셀시료번호", "") else "NG",
        "사이트만존재": "",
        "엑셀만존재": "",
    })

    # ★ 비산먼지면 특정 필드는 비교 PASS
    is_dust = bool(excel.get("is_dust") or site.get("is_dust"))

    for f in SIMPLE_FIELDS:
        if is_dust and f in DUST_SKIP_FIELDS:
            continue
        rows.append(compare_scalar(sample_no, f, site.get(f, ""), excel.get(f, "")))

    rows.append(compare_list(sample_no, "측정항목", site.get("측정항목", []), excel.get("측정항목", [])))
    rows.append(compare_list(sample_no, "장비", site.get("장비", []), excel.get("장비", [])))
    rows.append(compare_list(sample_no, "차량", site.get("차량", []), excel.get("차량", [])))
    rows.append(compare_list(sample_no, "인력", site.get("인력", []), excel.get("인력", [])))

    es = excel.get("측정시작DT", "")
    ee = excel.get("측정종료DT", "")

    rows.append(compare_mobile_single(sample_no, "환경기술인입력일시",
                                      site.get("환경기술인입력일시", ""), es, ee))
    rows.append(compare_mobile_single(sample_no, "GPS위치확인일시",
                                      site.get("GPS위치확인일시", ""), es, ee))
    rows.append(compare_mobile_photos(sample_no, site.get("촬영일시목록", []), es, ee))

    # RealGrid 비교는 그대로 (아래 2)에서 로직 추가)
    try:
        site_rg = site.get("realgrid", {}) if isinstance(site, dict) else {}
        excel_rg = excel.get("realgrid", {}) if isinstance(excel, dict) else {}
        if site_rg or excel_rg:
            rows.extend(build_realgird_compare_rows(sample_no, site_rg, excel_rg))
    except Exception as e:
        rows.append({
            "시료번호": sample_no,
            "항목": "[RealGrid] 비교 중 예외",
            "사이트값": str(e),
            "엑셀값": "",
            "비교": "NG",
            "사이트만존재": "",
            "엑셀만존재": ""
        })

    return rows
          

# ------------------------------------------------------------
# 결과 저장
# ------------------------------------------------------------
def _next_available_path(path: str) -> str:
    base, ext = os.path.splitext(path)
    for i in range(1, 1000):
        cand = f"{base}_{i}{ext}"
        if not os.path.exists(cand):
            return cand
    return f"{base}_{int(time.time())}{ext}"

def save_results(sample_rows_map, out_path):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "요약"
    ws_sum.append(["시료번호", "업소명", "항목", "비교", "사이트값", "엑셀값", "사이트만존재", "엑셀만존재"])


    for sample_no, rows in sample_rows_map.items():
        ws = wb.create_sheet(sample_no)
        ws.append(["항목", "사이트값", "엑셀값", "비교", "사이트만존재", "엑셀만존재"])
        # ✅ 개별 시료 시트 1행 1열 틀고정 (B2 셀 기준)
        ws.freeze_panes = "B2"
        for r in rows:
            ws.append([r["항목"], r["사이트값"], r["엑셀값"], r["비교"],
                       r["사이트만존재"], r["엑셀만존재"]])

            company = COMPANY_MAP.get(sample_no, "")
            ws_sum.append([
                sample_no, company,
                r["항목"], r["비교"],
                r["사이트값"], r["엑셀값"],
                r["사이트만존재"], r["엑셀만존재"]
            ])

        # PDF 하이퍼링크 행 추가
        pdf_path = PDF_MAP.get(sample_no, "")
        if pdf_path:
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value="PDF 열기")
            link_cell = ws.cell(row=row_idx, column=2, value=pdf_path)
            link_cell.hyperlink = pdf_path
            link_cell.style = "Hyperlink"

            sum_row = ws_sum.max_row + 1
            company = COMPANY_MAP.get(sample_no, "")
            ws_sum.cell(row=sum_row, column=1, value=sample_no)
            ws_sum.cell(row=sum_row, column=2, value=company)
            ws_sum.cell(row=sum_row, column=3, value="PDF 열기")
            ws_sum.cell(row=sum_row, column=4, value="OK")
            sum_link_cell = ws_sum.cell(row=sum_row, column=5, value=pdf_path)
            sum_link_cell.hyperlink = pdf_path
            sum_link_cell.style = "Hyperlink"

    # NG 빨간색 조건부서식
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 요약 시트(C열)
    if ws_sum.max_row > 1:
        rule_sum = FormulaRule(formula=['$D2="NG"'], fill=red_fill)
        ws_sum.conditional_formatting.add(f"D2:D{ws_sum.max_row}", rule_sum)
# 각 시료 시트(D열) - 요약/중복검사 시트들은 제외
    for name in wb.sheetnames:
        if name == "요약":
            continue

        ws_sample = wb[name]
        if ws_sample.max_row > 1:
            rule_sample = FormulaRule(formula=['$D2="NG"'], fill=red_fill)
            ws_sample.conditional_formatting.add(f"D2:D{ws_sample.max_row}", rule_sample)

    try:
        wb.save(out_path)
        print(f"\n[완료] 결과 저장 → {out_path}")
    except PermissionError as e:
        alt = _next_available_path(out_path)
        wb.save(alt)
        print(f"\n⚠️ 저장 실패(잠금/권한): {out_path}")
        print(f"[대체 저장] → {alt}")



# ------------------------------------------------------------
# 메인 실행
# ------------------------------------------------------------
def main():
    print("=== 측정인.kr 자동 비교 시스템 시작 ===")

    login_id = input("측정인아이디: ").strip()
    login_pw = input("측정인비밀번호: ").strip()
    start_date = input("시작일 (YYYY-MM-DD): ").strip()
    end_date = input("종료일 (YYYY-MM-DD): ").strip()
    team_input = input("팀번호(1-5, 예: 3 / 1,3,5 / 1-3, 엔터=전체): ").strip()

    yyyymmdd = start_date.replace("-", "")

    # ------------------------------------------------------------
    # 0) 파일 기반 시료번호 목록 먼저 생성 (팀 필터 포함)
    #    - 선택한 팀이 파일 목록에 없으면: 사이트 로그인/브라우저 실행 없이 종료
    # ------------------------------------------------------------
    samples = collect_samples_from_files(start_date, nas_base=NAS_BASE, nas_dirs=NAS_DIRS)
    print(f"[파일 기반 시료번호] {len(samples)}개 → {samples}")

    teams = parse_team_input(team_input)

    if teams:
        team_set = set(teams)
        samples = [sn for sn in samples if get_team_no_from_sample(sn) in team_set]

        # 필터 후 실제 남은 팀만 추출
        run_teams = sorted(
            {get_team_no_from_sample(sn) for sn in samples if str(get_team_no_from_sample(sn)).isdigit()},
            key=lambda x: int(x)
        )

        print(f"[팀 선택] 실행 팀 → {','.join(run_teams) if run_teams else '-'}팀  |  {len(samples)}개")

    else:
        teams = sorted(
            {get_team_no_from_sample(sn) for sn in samples if str(get_team_no_from_sample(sn)).isdigit()},
            key=lambda x: int(x)
        )
        print(f"[전체 선택] 발견된 팀 자동 분리 대상 → {','.join(teams) if teams else '-'}팀  |  {len(samples)}개")

    if not samples:
        if team_input.strip():
            print(f"❌ 선택한 팀({team_input})에 해당하는 시료번호가 파일에서 발견되지 않음 → 사이트 로그인 없이 종료")
        else:
            print("❌ 해당 날짜 파일에서 시료번호 없음 → 사이트 로그인 없이 종료")
        return

    # 단일 팀이면 기존과 동일하게 team_tag 사용(파일명/로그용)
    team_tag = teams[0] if len(teams) == 1 else ""

    out_dir = r"\\192.168.10.163\측정팀\10.검토\3.측정인 검토"
    out_name = f"{yyyymmdd} 팀{team_tag} 검토 파일.xlsx" if team_tag else f"{yyyymmdd} 검토 파일.xlsx"
    RESULT_XLSX = os.path.join(out_dir, out_name)

    driver = init_driver()
    sample_rows = {}
    excel_meta_map = {}  # 시료별 (날짜/업소/채취시간) 메타

    try:
        login(driver, login_id, login_pw)
        search_date(driver, start_date, end_date)

        for sample_no in samples:
            print("\n-------------------------------------------")
            print(" 시료:", sample_no)
            print("-------------------------------------------")

            # 상세페이지 진입
            if not open_sample_detail(driver, sample_no):
                continue
            time.sleep(1)
            site = read_site_data(driver, sample_no)
            PDF_MAP[sample_no] = site.get("PDF경로", "")
            xlsx = find_excel_for_sample(sample_no)
            if not xlsx:
                go_back_to_list(driver)
                continue
            is_dust = ("비산" in os.path.basename(str(xlsx)))
            excel = parse_measuring_record(str(xlsx), sample_no)
            COMPANY_MAP[sample_no] = excel.get("업소명", "")
            # --------------------------------------------------
            # 시료별 메타 저장(날짜/업소/채취시간) - 환경기술인 입력일시 완화용
            # --------------------------------------------------
            excel_meta_map[sample_no] = {
                "날짜": excel.get("날짜", ""),
                "업소명": excel.get("업소명", ""),
                "측정시작DT": excel.get("측정시작DT", ""),
                "측정종료DT": excel.get("측정종료DT", ""),
            }

            excel["is_dust"] = is_dust
            site["is_dust"] = is_dust
            # --------------------------------------------------
            # 탭2 RealGrid(항목별 표)도 같이 비교: 수기 입력 오타 탐지용
            # --------------------------------------------------
            time.sleep(1)

            try:
                site_rg = rg_api_read_data(driver, REALGRID_ROOT_CSS)
            except Exception as e:
                site_rg = {}
                # RealGrid를 못 읽어도 전체 비교는 진행
                print("⚠ 탭2 RealGrid 읽기 실패:", e)

            try:
                excel_rg = build_excel_realgird_expected(str(xlsx), sample_no, is_dust=is_dust)
            except Exception as e:
                excel_rg = {}
                print("⚠ 엑셀 RealGrid 기대값 생성 실패:", e)

            site["realgrid"] = site_rg
            excel["realgrid"] = excel_rg

            rows = build_comparison_rows(sample_no, site, excel)
            sample_rows[sample_no] = rows
            relax_env_input_time_by_company(sample_rows, excel_meta_map)
            print(" 완료 : ", sample_no)
            go_back_to_list(driver)

        if sample_rows:
            # teams가 여러 개인 경우 → 팀별로 파일을 따로 저장 (dash에서 여러 파일 선택해서 종합검토)
            if teams and len(teams) > 1:
                base_dir = os.path.dirname(RESULT_XLSX)
                for t in teams:
                    team_map = {sn: rows for sn, rows in sample_rows.items() if get_team_no_from_sample(sn) == t}
                    if not team_map:
                        continue
                    out_name = f"{yyyymmdd} 팀{t} 검토 파일.xlsx"
                    out_path = os.path.join(base_dir, out_name)
                    save_results(team_map, out_path)
                    print(f"✅ 저장 완료: {out_path}")
            else:
                save_results(sample_rows, RESULT_XLSX)
                print(f"✅ 저장 완료: {RESULT_XLSX}")
        else:
            print("⚠ 저장할 결과 없음")

    finally:
        print("작업 종료. 브라우저는 직접 닫아도 됨.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log_error("eco_check.main", e)
        raise
