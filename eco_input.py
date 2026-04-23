# -*- coding: utf-8 -*-
"""
측정인.kr 자동 입력 – 팀별 시료번호 자동 처리 / 비산먼지 전용 로직 포함(최종)
요청된 부분만 수정하고 기존 안정동작 부분은 전혀 손대지 않음.
"""
import os
import time
import re
import traceback
import pyperclip
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pythoncom
import win32com.client as win32
import win32gui, win32con
from pywinauto.application import Application
from pywinauto import Desktop
from pywinauto.keyboard import send_keys
from selenium.common.exceptions import (
    UnexpectedAlertPresentException,
    NoAlertPresentException,
    TimeoutException,
)
import datetime as dt
from excel_utils import find_sheet_by_candidates, parse_measuring_record

#===================경고 제거========================
# ✅ openpyxl 조건부서식 경고 숨김 (오류 아님, 콘솔 정리용)
import warnings
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Conditional Formatting extension is not supported and will be removed"
)

# ============================================================
# 공통 유틸 모듈 (모듈화)
# ============================================================
from selenium_utils import (
    init_driver, safe_click, wait_el, set_date_js,
    accept_all_alerts as _accept_all_alerts,
    accept_all_alerts as _accept_alerts_base,
    close_popup, wait
)
from measin_utils import (
    login, search_date, wait_grid_loaded,
    open_sample_detail as _open_sample_detail,
    go_back_to_list as _go_back_to_list,
    collect_samples_from_files,
    LOGIN_URL, FIELD_URL, NAS_BASE, NAS_DIRS
)
from format_utils import format_time as trim_hm, to_f1, to_f2
from data_utils import parse_ymd_date, sample_to_datestr, clean_leading_mark as clean
from realgrid_utils import (
    rg_dump_headers, rg_find_col, rg_find_cols, rg_get_body, rg_api_write_data, 
    rg_scroll_top, rg_find_tr_by_item, rg_paste_to_tr, rg_paste_to_tr_tab4, rg_set_cell_by_keys
)
from select2_utils import Select2Handler as _Select2Handler
from pdf_utils import merge_pdfs as _merge_pdfs, PDFExporter as _PDFExporter
from excel_utils import find_sheet_by_candidates as _find_sheet_by_candidates
from file_utils import collect_samples_from_nas as _collect_samples_from_nas_util, find_best_matching_file as _find_best_file_util, is_fugitive_dust_file
from excel_com_utils import get_excel_app, kill_excel_app
from log_utils import log_error
from measin_constants import (
    SKIP_VOL_AND_SPEED, SKIP_SPEED_ONLY, HEAVY_METALS,
    SM3_ITEMS as sm3_items,
    SEL_DATE, SEL_START_TIME, SEL_END_TIME,
    SEL_O2_STD, SEL_O2_MEAS, SEL_GAS_VOL_PRE, SEL_GAS_VOL_POST,
    SEL_MOISTURE, SEL_GAS_TEMP, SEL_GAS_SPEED
)
from backdata_utils import (
    export_backdata_moist_thc, export_csv_display_as_is, export_fid_by_excel_copy
)
from tab4_utils import (
    read_tab4_from_macro_xlsm, fill_tab4_dates, fill_tab4_grid_only,
    tab4_find_tr_by_item, tab4_get_api_items, tab4_paste_row_using_tab2,
    tab4_temp_save, tab4_comp_save,
    TAB4_SELECTOR, TAB4_GRID_ROOT
)

# =====================================================================
# 설정
# =====================================================================
#===============================팝업닫기========================

POPUP_CLOSE_SEL = "body > div > div > div.modal-body > div.modal-footer.row > form > input"

# PDF 임시 생성 폴더(로컬 권장: 업로드 창에서 경로 인식/권한 문제 줄어듦)
PDF_TMP_DIR = r"C:\measin_upload_tmp"

# 백데이터/탭4 상수 → backdata_utils.py, tab4_utils.py 에서 import 완료

TAB4_FILE_BTN1 = "#newFile1"   # 시험 분석일지(PDF) 업로드 버튼(열기창)
TAB4_FILE_BTN2 = "#newFile2"   # 측정기록부(PDF) 업로드 버튼(열기창)

FINAL_DONE_DIR = r"\\192.168.10.163\측정팀\2.성적서\0 5.최종완료"
PDF_BASE_DIR   = r"\\192.168.10.163\측정팀\2.성적서\0.PDF\2.대기pdf"

# ── 수질 ──────────────────────────────────────────────────
FIELD_URL_WATER    = "https://측정인.kr/ms/field_outwater.do"
# 수질 성적서: NAS_BASE_WATER\YYYY년\업체명\파일.xlsm
NAS_BASE_WATER     = r"\\192.168.10.163\측정팀\2.성적서\14.수질성적서"
# 수질 PDF:    PDF_BASE_DIR_WATER\YYYY년\업체명\파일.pdf
PDF_BASE_DIR_WATER = r"\\192.168.10.163\측정팀\2.성적서\0.PDF\1.수질pdf"
# 수질 탭4 PDF 업로드 셀렉터
WATER_FILE_BTN1    = "#anzeFile1"   # 분석일지
WATER_FILE_BTN2    = "#anzeFile2"   # 대행기록부
# 수질 엑셀 시트명
WATER_SHEET_ANALY  = "분석일지"
WATER_SHEET_RECORD = "대행기록부"

# ======================== 공통 import ========================
import atexit
import win32clipboard  # 기존 로직 유지용

# ======================== Excel COM 재사용(전역) ========================
_EXCEL_APP = None

# XL_UP, XL_TOLEFT → backdata_utils.py 로 이동
# wait_until_sheet_updates, _cell_text, read_tab4_from_macro_xlsm → tab4_utils.py 로 이동
# 클립보드 유틸(_clipboard_clear 등) → backdata_utils.py 로 이동


# =====================================================================
# RealGrid 예외 규칙 → measin_constants.py 에서 import 완료
# =====================================================================

def rg2_fill_measure_grid_api(driver, sample_no: str, per_item: dict):
    """RealGrid API로 직접 입력 (js 모듈화 완료)"""
    date_str = sample_to_datestr(sample_no)
    if not date_str:
        raise RuntimeError("시료번호에서 날짜 파싱 실패")

    payload = []
    for item_name, v in per_item.items():
        st = v.get("시작시간", "")
        et = v.get("종료시간", "")
        vol = v.get("시료채취량", "")
        spd = v.get("흡인속도", "")
        vol_u = v.get("채취량단위", "L")
        spd_u = v.get("흡인속도단위", "L/min")

        if item_name in SKIP_VOL_AND_SPEED:
            vol, vol_u, spd, spd_u = "", "", "", ""
        elif item_name in SKIP_SPEED_ONLY:
            spd, spd_u = "", ""

        payload.append({
            "item": item_name, "sd": date_str, "st": st, "ed": date_str, "et": et,
            "vol": vol, "vol_u": vol_u, "spd": spd, "spd_u": spd_u,
        })

    # 길었던 JS 블록 삭제! 대신 유틸리티 함수 호출
    result = rg_api_write_data(driver, payload)
    
    print("▶ RealGrid API 입력 결과:", result)
    missing = result.get("missing")
    if missing:
        print("⚠ RealGrid에 없어서 미적용된 항목들:")
        for x in missing:
            print("   -", x)
    return result


# _norm_rg, tab4_find_tr_by_item, tab4_paste_row_using_tab2,
# tab4_get_api_items, fill_tab4_grid_only → tab4_utils.py 로 이동


# =====================================================================
# 공통 유틸
# =====================================================================

def parse_sample_input(s: str):
    """
    "A2512313-01, A2512313-02" / 공백 / 줄바꿈 섞여도 여러개 파싱
    """
    if not s:
        return []
    parts = re.split(r"[,\s]+", s.strip())
    out, seen = [], set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def find_sample_file_in_nas(sample_no: str):
    """대기 NAS에서 sample_no 접두 일치 엑셀 파일 검색 (loose)"""
    return _find_best_file_util(
        sample_no,
        nas_base=NAS_BASE,
        nas_dirs=NAS_DIRS,
        extensions=(".xlsm", ".xlsx", ".xls"),
        strict=False,
    )


def find_sample_file_in_water_nas(sample_no: str):
    """수질 NAS에서 엑셀 파일 검색.
    경로 구조: NAS_BASE_WATER\YYYY년\업체명\파일.xlsm  (하위 전체 재귀)
    """
    return _find_best_file_util(
        sample_no,
        nas_base=NAS_BASE_WATER,
        nas_dirs=[""],
        extensions=(".xlsm", ".xlsx", ".xls"),
        strict=False,
    )


def ask_yesno(msg, default_yes=True):
    """
    예/아니오 입력 받기.
    - yes: 예, y, yes, ㅇ, 1
    - no : 아니오, n, no, ㄴ, 0
    엔터만 치면 default_yes 적용
    """
    while True:
        s = input(msg).strip().lower()
        if s == "":
            return default_yes
        if s in ("예", "y", "yes", "ㅇ", "1"):
            return True
        if s in ("아니오", "n", "no", "ㄴ", "0"):
            return False
        print("⚠ 입력은 예/아니오(또는 y/n)로 해줘.")


# wait → selenium_utils.wait() 로 이동
# clean → data_utils.clean_leading_mark() 로 이동 (import 시 clean 으로 alias)

# accept_any_alert: eco_input 전용 (단일 알림 수락 + 텍스트 반환)
def accept_any_alert(driver, timeout=2):
    end = time.time() + timeout
    while time.time() < end:
        try:
            a = driver.switch_to.alert
            txt = a.text
            a.accept()
            time.sleep(0.2)
            return txt
        except:
            time.sleep(0.1)
    return None

# =====================================================================
# 시료번호 목록 자동 생성
# =====================================================================

def extract_samples_from_nas(team_nos, date_str): # 파라미터 이름 변경
    """
    파일명에서 시작부분 AYYMMDDT-XX 추출.
    team_nos(리스트 또는 문자열 튜플), 날짜(YYMMDD) 일치하는 파일만 목록에 추가.
    파일명에 정확히 '비산먼지' 포함 시 dust=True
    """
    return _collect_samples_from_nas_util(
        nas_base=NAS_BASE,
        nas_dirs=NAS_DIRS,
        date_str=date_str,
        team_nos=team_nos,
        sample_pattern=r"^(A\d{6}\d-\d{2})",
        dust_keywords=("비산먼지",),
    )

# =====================================================================
# 사이트 상세 진입 – 시료번호 검색 방식
# =====================================================================

def open_detail_by_sample(d, sample_no):
    """measin_utils.open_sample_detail 래퍼 (eco_input 호환성 유지)"""
    return _open_sample_detail(d, sample_no)

def back_to_list(d, btn_selector="#btnMsFieldDocCancel"):
    """measin_utils.go_back_to_list 래퍼 (eco_input 호환성 유지)"""
    selectors = [btn_selector]
    if btn_selector != "#btnGoList":
        selectors.append("#btnGoList")
    # accept_any_alert 먼저 처리
    accept_any_alert(d, timeout=2)
    result = _go_back_to_list(d, btn_selectors=selectors)
    accept_any_alert(d, timeout=2)
    return result

# =====================================================================
# 탭1 입력
# =====================================================================

def set_dust_radio(d, is_dust):
    try:
        if is_dust:
            sel = "#rowAirRpt > section:nth-child(1) > div > label:nth-child(1)"
        else:
            sel = "#rowAirRpt > section:nth-child(1) > div > label.radio.ml-1 > i"
        safe_click(d, sel)
        print("  → 비산먼지 설정 완료:", is_dust)
    except:
        print("⚠ 비산먼지 설정 실패")

def clear_multi(d, ul_sel):
    _Select2Handler(d).clear(ul_sel)

def add_multi(d, search_sel, val):
    _Select2Handler(d).add(search_sel, val)

def fill_multi(d, ul_sel, search_sel, values):
    _Select2Handler(d).fill(ul_sel, search_sel, values)

def fill_tab1(d, data, is_dust):
    print("▶ 탭1 입력 시작")
    set_dust_radio(d, is_dust)

    # --------------------------------------
    # ① 비산먼지인 경우
    # --------------------------------------
    if is_dust:
        print("▶ 비산먼지 → 기존 선택 삭제 + 전용 셀렉터로 입력")


        # 측정항목
        fill_multi(
            d,
            "#inairTargetItem > div:nth-child(2) > div > span > span.selection > span > ul",
            "#inairTargetItem > div:nth-child(2) > div > span > span.selection > span > ul > li.select2-search.select2-search--inline > input",
            data["측정항목"]
        )

        # 인력
        fill_multi(
            d,
            "#wid-id-4 > div > div.widget-body.no-padding > div > fieldset > div.row.input-full > section:nth-child(2) > span > span.selection > span > ul",
            "#wid-id-4 > div > div.widget-body.no-padding > div > fieldset > div.row.input-full > section:nth-child(2) > span > span.selection > span > ul > li.select2-search.select2-search--inline > input",
            data["인력"]
        )

        # 차량
        fill_multi(
            d,
            "#carSection > div > span > span.selection > span > ul",
            "#carSection > div > span > span.selection > span > ul > li.select2-search.select2-search--inline > input",
            data["차량"]
        )

        # 장비
        fill_multi(
            d,
            "#machineDiv > div > span > span.selection > span > ul",
            "#machineDiv > div > span > span.selection > span > ul > li.select2-search.select2-search--inline > input",
            data["장비"]
        )


        print("▶ 비산먼지 탭1 입력 완료")

    # --------------------------------------
    # ② 비산먼지 아닌 경우 (원래 잘 되던 로직 그대로)
    # --------------------------------------
    else:
        # 측정항목
        fill_multi(
            d,
            "#inairTargetItem > div:nth-child(2) > div > span > span.selection > span > ul",
            "#inairTargetItem input[type='search']",
            data["측정항목"]
        )

        # 인력
        fill_multi(
            d,
            "#edit_emp_id + span ul.select2-selection__rendered",
            "#edit_emp_id + span input.select2-search__field",
            data["인력"]
        )

        # 차량
        fill_multi(
            d,
            "#carSection > div > span > span.selection > span > ul",
            "#carSection input[type='search']",
            data["차량"]
        )

        # 장비
        fill_multi(
            d,
            "#machineDiv > div > span > span.selection > span > ul",
            "#machineDiv input[type='search']",
            data["장비"]
        )

    print("▶ 탭1 입력 완료")

    # 탭1 저장 버튼 (비산먼지/일반 공통)
    try:
        safe_click(d, "#updateFieldPlanBtn")
        try:
            d.switch_to.alert.accept(); wait(0.5)
        except:
            pass
        try:
            d.switch_to.alert.accept(); wait(0.5)
        except:
            pass
        print("✅ 탭1 저장 완료")
    except:
        print("⚠ 탭1 저장 실패")


# =====================================================================
# 탭2 입력
# =====================================================================

def set_wind(d, txt):
    if not txt: return
    mp={"북":0,"북-북동":1,"북동":2,"동-북동":3,"동":4,"동-남동":5,"남동":6,
        "남-남동":7,"남":8,"남-남서":9,"남서":10,"서-남서":11,"서":12,
        "서-북서":13,"북서":14,"북-북서":15,"정온":99}
    v=mp.get(txt)
    if v is None: return
    try:
        el=WebDriverWait(d,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"select.meas_wdir")))
        Select(el).select_by_value(str(v))
        d.execute_script(
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el)
    except: pass

def fill_tab2(d, data, is_dust):
    print("▶ 탭2 입력 시작")

    def sv(sel,val):
        try:
            e=d.find_element(By.CSS_SELECTOR,sel)
            e.clear(); e.send_keys(str(val)); wait(0.1)
        except: pass

    # 공통
    sv("input[name='meas_wthr']", data["기상"])
    sv("input[name='meas_temper']", data["기온"])
    sv("input.meas_humd", data["습도"])
    sv("input.meas_atoms", data["기압"])
    set_wind(d, data["풍향"])
    sv("input.meas_wspd", data["풍속"])

    sv(SEL_START_TIME, data["채취시작"])
    sv(SEL_END_TIME, data["채취끝"])

    if not is_dust:
        sv(SEL_O2_STD, data["표준산소농도"])
        sv(SEL_O2_MEAS, data["실측산소농도"])
        sv(SEL_GAS_VOL_PRE, data["배출가스유량전"])
        sv(SEL_GAS_VOL_POST, data["배출가스유량후"])
        sv(SEL_MOISTURE, data["수분량"])
        sv(SEL_GAS_TEMP, data["배출가스온도"])
        sv(SEL_GAS_SPEED, data["배출가스유속"])

    print("✅ 탭2 입력 완료")


def ensure_gas_flow_checkbox_checked(driver, selector="#meas_gas_fvol_yn"):
    try:
        checkbox = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
        )
        is_checked = driver.execute_script("return !!arguments[0].checked;", checkbox)
        if not is_checked:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", checkbox)
            driver.execute_script("arguments[0].click();", checkbox)
            time.sleep(0.2)
            print("✅ 배출가스 유량 체크 완료")
        else:
            print("▶ 배출가스 유량 이미 체크됨")
    except Exception as e:
        print("⚠ 배출가스 유량 체크 실패:", e)



# ------------------------------------------------------------
# 🔽🔽 여기부터 새 기능 추가 (기존 코드 절대 수정 없음, 아래만 수정)
# ------------------------------------------------------------
def parse_facility_from_excel(path):
    """
    엑셀의 '대기측정기록부'에서 26~29행 D,F,H,J,L,N 가져오기
    (시설/가동상황 표)
    """
    wb = load_workbook(path, data_only=True)

    if "대기측정기록부" in wb.sheetnames:
        ws = wb["대기측정기록부"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(26, 30):  # 26,27,28,29
        row = {
            "fuel": ws[f"D{r}"].value,        # 연료사용량
            "prod": ws[f"F{r}"].value,        # 제품생산량
            "inc": ws[f"H{r}"].value,         # 소각량
            "raw": ws[f"J{r}"].value,         # 원료투입량
            "kind": ws[f"L{r}"].value,        # 종류
            "unit": ws[f"N{r}"].value,        # 단위
        }
        rows.append(row)
    return rows


def select_first_prev_facility(driver, row_idx):
    """
    방지시설 Select2:
    첫 번째 옵션 클릭 후 ENTER로 선택 확정
    """
    try:
        search_input = f"#prev_fac_no_by_emis_fac_{row_idx} + span input.select2-search__field"

        # 검색창 클릭(열기)
        inp = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, search_input))
        )
        inp.click()
        time.sleep(0.2)

        # 옵션 목록 가져오기
        options = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "ul.select2-results__options li")
            )
        )

        if not options:
            print(f"⚠ row{row_idx}: 방지시설 옵션 없음")
            return

        first = options[0]

        # 1차 클릭(선택 후보)
        driver.execute_script("arguments[0].click();", first)
        time.sleep(0.2)

        # ENTER로 선택 확정
        inp.send_keys(Keys.ENTER)
        time.sleep(0.2)

        print(f" - row{row_idx}: 방지시설 첫 항목 ENTER로 최종 선택 완료")

    except Exception as e:
        print(f"⚠ row{row_idx}: 방지시설 선택 실패 → {e}")

def fill_facility_rows(driver, fac_rows):
    """
    rowfac0 ~ rowfac3 까지 4행 자동 입력
    fac_rows: parse_facility_from_excel 결과 리스트
    """

    print("▶ 배출시설 가동상황 자동 입력")

    for idx, row in enumerate(fac_rows):
        try:
            fuel = "" if row["fuel"] is None else str(row["fuel"]).strip()
            prod = "" if row["prod"] is None else str(row["prod"]).strip()
            inc  = "" if row["inc"]  is None else str(row["inc"]).strip()
            raw  = "" if row["raw"]  is None else str(row["raw"]).strip()
            kind = "" if row["kind"] is None else str(row["kind"]).strip()
            unit = "" if row["unit"] is None else str(row["unit"]).strip()

            # 각 칸 입력 (기존 기능 유지)
            driver.find_element(By.CSS_SELECTOR, f"#fuel_used_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#fuel_used_{idx}").send_keys(fuel)

            driver.find_element(By.CSS_SELECTOR, f"#production_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#production_{idx}").send_keys(prod)

            driver.find_element(By.CSS_SELECTOR, f"#incinerate_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#incinerate_{idx}").send_keys(inc)

            driver.find_element(By.CSS_SELECTOR, f"#raw_material_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#raw_material_{idx}").send_keys(raw)

            driver.find_element(By.CSS_SELECTOR, f"#kind_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#kind_{idx}").send_keys(kind)

            driver.find_element(By.CSS_SELECTOR, f"#unit_{idx}").clear()
            driver.find_element(By.CSS_SELECTOR, f"#unit_{idx}").send_keys(unit)

            # ✅ 연료사용량이 있으면 → 방지시설 첫 번째 항목 선택
            if fuel != "":
                select_first_prev_facility(driver, idx)
            else:
                print(f" - row{idx}: 연료사용량 없음 → 방지시설 선택 안 함")

        except Exception as e:
            print(f"⚠ 배출시설 row{idx} 입력 실패:", e)

    print("✅ 배출시설 가동상황 입력 완료")


def write_sampler_comment(driver, text="특이사항 없음"):
    """
    채취자 의견
    """
    try:
        opin = driver.find_element(By.CSS_SELECTOR, "#smpl_ctor_opin")
        opin.clear()
        opin.send_keys(text)
        time.sleep(0.3)
        print("✅ 채취자 의견 입력 완료")

    except Exception as e:
        print("⚠ 채취자 의견 입력 실패:", e)


# fill_tab4_dates → tab4_utils.py 로 이동


# ------------------------------------------------------------
# 🔽 탭2 저장 (입력완료→확인2번)
# ------------------------------------------------------------

def _should_draft_by_sampling_end(date_str: str, end_hm: str) -> bool:
    """
    현재시간이 '시료채취(해당 날짜의) 끝 시간'보다 이르면 True(임시저장)
    - date_str: 'YYYY-MM-DD'
    - end_hm: 'HH:MM' 또는 'HH:MM:SS'
    """

    if not date_str or not end_hm:
        return False  # 값 없으면 최종저장 쪽(원하면 True로 바꿔도 됨)

    date_str = str(date_str).strip()
    end_hm = str(end_hm).strip()

    # 날짜 파싱
    meas_date = parse_ymd_date(date_str)
    if meas_date is None:
        return False

    # 시간 파싱
    try:
        parts = end_hm.split(":")
        hh = int(parts[0]); mm = int(parts[1]); ss = int(parts[2]) if len(parts) >= 3 else 0
        end_t = dt.time(hh, mm, ss)
    except Exception:
        return False

    now = datetime.now()

    # ✅ 날짜가 과거면 이미 끝난걸로 보고 "최종저장" 가능 (임시저장 X)
    if meas_date < now.date():
        return False

    # ✅ 날짜가 미래면 아직 측정일도 안됨 → 임시저장 강제
    if meas_date > now.date():
        return True

    # ✅ 날짜가 오늘이면 시간 비교
    end_dt = datetime.combine(meas_date, end_t)
    return now < end_dt



def save_tab2(driver, use_final_save=False):
    """
    use_final_save=False  → 임시저장(#btnDraftMsFieldDoc)
    use_final_save=True   → 입력완료(#btnSaveMsFieldDoc)  (PDF 업로드 동반 시)
    """
    btn_sel = "#btnSaveMsFieldDoc" if use_final_save else "#btnDraftMsFieldDoc"
    print(f"▶ 탭2 저장 시작 ({'입력완료' if use_final_save else '임시저장'})")

    try:
        btn = driver.find_element(By.CSS_SELECTOR, btn_sel)
        driver.execute_script("arguments[0].click();", btn)

        # ✅ 클릭 직후 바로 뜨는 알럿들 처리
        _accept_all_alerts(driver, total_wait=2.5, poll=0.15, label="1차")

        # ✅ 업로드/서버 처리 대기(입력완료면 더 길게)
        #    이 구간에도 알럿이 늦게 뜰 수 있으니 '기다리기+알럿처리'로 통합
        wait_time = 1.5 if not use_final_save else 7.0
        _accept_all_alerts(driver, total_wait=wait_time, poll=0.2, label="대기중")

        # ✅ 마무리로 한 번 더(잔여 알럿 방지)
        _accept_all_alerts(driver, total_wait=2.0, poll=0.2, label="마무리")

        print("▶ 탭2 저장 완료")
        return True

    except Exception as e:
        print("⚠ 탭2 저장 실패:", e)
        return False



        
        
# ============================================================
# PDF 생성(엑셀 COM) + 사이트 업로드(파일창 자동) + 업로드 후 PDF 삭제
# ============================================================

def _find_sheet_name(wb, prefer_names):
    names = [sh.Name for sh in wb.Worksheets]
    for pn in prefer_names:
        if pn in names:
            return pn
    for pn in prefer_names:
        p = pn.replace(" ", "")
        for n in names:
            if p in n.replace(" ", ""):
                return n
    return None


def export_pdf_from_excel(excel_path: str, sample_no: str, is_dust: bool) -> str:
    out_dir = PDF_TMP_DIR
    os.makedirs(out_dir, exist_ok=True)
    out_pdf = os.path.join(out_dir, f"{sample_no}.pdf")

    if is_dust:
        sheet_a_candidates = ["대기시료채취 및 분석일지", "대기시료채취및분석일지"]
        sheet_b_candidates = ["교정"]
    else:
        sheet_a_candidates = ["대기시료채취 및 분석일지", "대기시료채취및분석일지"]
        sheet_b_candidates = ["먼지시료채취기록지"]

    excel = None
    wb = None
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(excel_path, ReadOnly=True)

        sh_a = _find_sheet_name(wb, sheet_a_candidates)
        sh_b = _find_sheet_name(wb, sheet_b_candidates)

        if not sh_a:
            raise RuntimeError(f"필수 시트 없음: {sheet_a_candidates}")
        if not sh_b:
            raise RuntimeError(f"필수 시트 없음: {sheet_b_candidates}")

        # ✅ VARIANT 없이 "시트 2개" 멀티선택 (안정적)
        wb.Worksheets(sh_a).Select()          # 첫 시트 선택
        wb.Worksheets(sh_b).Select(False)     # Replace=False → 선택에 추가

        # 0 = xlTypePDF
        wb.ActiveSheet.ExportAsFixedFormat(0, out_pdf)


        return out_pdf

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        try:
            pythoncom.CoUninitialize()
        except:
            pass


def _win32_is_file_open_dialog(hwnd) -> bool:
    """#32770 공용 대화상자 중 '열기/Open' 버튼 + Edit(파일명 입력칸) 있는지로 판별"""
    try:
        if not win32gui.IsWindowVisible(hwnd):
            return False
        if win32gui.GetClassName(hwnd) != "#32770":
            return False

        found_open_btn = False
        found_edit = False

        def enum_child(ch, _):
            nonlocal found_open_btn, found_edit
            cls = win32gui.GetClassName(ch)
            txt = win32gui.GetWindowText(ch)

            if cls == "Button" and (("열기" in txt) or ("Open" in txt)):
                found_open_btn = True

            # 파일이름 입력칸은 Edit가 직접 있거나 ComboBox 안에 숨어있음
            if cls == "Edit":
                found_edit = True

        win32gui.EnumChildWindows(hwnd, enum_child, None)

        # Edit가 바로 없으면 ComboBox/ComboBoxEx32 안의 Edit도 탐색
        if not found_edit:
            def enum_child2(ch, _):
                nonlocal found_edit
                cls = win32gui.GetClassName(ch)
                if cls in ("ComboBoxEx32", "ComboBox"):
                    def enum_grand(gch, __):
                        nonlocal found_edit
                        if win32gui.GetClassName(gch) == "Edit":
                            found_edit = True
                    win32gui.EnumChildWindows(ch, enum_grand, None)
            win32gui.EnumChildWindows(hwnd, enum_child2, None)

        return found_open_btn and found_edit
    except:
        return False


def _find_open_dialog(timeout=30):
    """
    ✅ 제목이 비어 있어도 찾도록 개선:
    1) Win32 EnumWindows로 #32770 중 '열기/Open 버튼 + 파일명 입력칸' 가진 창을 찾음
    2) 실패 시 UIA로 '열기/Open 버튼' 가진 창을 찾음(제목 무시)
    """
    end = time.time() + timeout

    while time.time() < end:
        # 1) Win32: 최우선(가장 빠르고 확실)
        try:
            found = []
            def enum_top(hwnd, _):
                if _win32_is_file_open_dialog(hwnd):
                    found.append(hwnd)
            win32gui.EnumWindows(enum_top, None)

            if found:
                # 가장 최근에 뜬 애를 잡기(보통 마지막이 최신)
                hwnd = found[-1]
                # pywinauto win32 wrapper 반환
                return Desktop(backend="win32").window(handle=hwnd)
        except:
            pass

        # 2) UIA: 제목이 없어도 '열기/Open' 버튼이 있는 창 찾기
        try:
            for w in Desktop(backend="uia").windows():
                try:
                    # 버튼 이름으로 찾기(열기/Open)
                    btns = w.descendants(control_type="Button")
                    if any(("열기" in b.window_text()) or ("Open" in b.window_text()) for b in btns):
                        return w
                except:
                    continue
        except:
            pass

        time.sleep(0.1)

    raise RuntimeError("열기(Open) 창을 찾지 못함(timeout). (제목이 비거나 다른 구조일 가능성)")



def pick_file_in_open_dialog(file_path: str, timeout=30):
    """
    ✅ 가장 안정적인 입력: Alt+N → Ctrl+V → Enter
    (UIA Edit set_text는 PC마다 안 먹는 경우가 많아서 키보드 입력으로 고정)
    """
    dlg = _find_open_dialog(timeout=timeout)

    # 전면/포커스 강제
    try:
        hwnd = dlg.handle
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.SetForegroundWindow(hwnd)
    except:
        pass

    try:
        dlg.set_focus()
    except:
        pass

    time.sleep(0.2)

    # 파일 이름 칸 포커스(Alt+N) → 붙여넣기 → Enter
    pyperclip.copy(file_path)
    send_keys("%n")                 # 파일 이름(N)
    time.sleep(0.05)
    send_keys("^a{BACKSPACE}")      # 지우기
    time.sleep(0.05)
    send_keys("^v")                 # 붙여넣기
    time.sleep(0.1)
    send_keys("{ENTER}")            # 열기

    # 창이 닫히면 성공
    try:
        dlg.wait_not("visible", timeout=6)
        return True
    except:
        raise RuntimeError(f"파일 선택 실패(열기창이 닫히지 않음). 경로/권한 확인: {file_path}")



def trigger_file_dialog(driver, timeout=6):
    """
    '파일추가' 버튼( onclick="control.openFileDialog();" )로 파일 선택창을 띄운다.
    1) 진짜 클릭(ActionChains)
    2) 안 되면 JS로 control.openFileDialog() 직접 호출
    """
    btn_sel = "#fileArea input.btn.btnView[value='파일추가']"

    # 1) 진짜 클릭
    try:
        btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, btn_sel))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.2)
        ActionChains(driver).move_to_element(btn).pause(0.05).click(btn).perform()
        time.sleep(0.4)
        return True
    except:
        pass

    # 2) JS 직접 호출(버튼 클릭이 막힐 때)
    try:
        driver.execute_script("""
            try {
                if (window.control && typeof window.control.openFileDialog === 'function') {
                    window.control.openFileDialog();
                    return true;
                }
                if (typeof control !== 'undefined' && control && typeof control.openFileDialog === 'function') {
                    control.openFileDialog();
                    return true;
                }
            } catch(e) {}
            return false;
        """)
        time.sleep(0.4)
        return True
    except:
        return False
        
        
def read_input_h7(excel_path: str) -> str:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["입력"] if "입력" in wb.sheetnames else wb[wb.sheetnames[0]]
    v = ws["H7"].value
    return "" if v is None else str(v).strip()

# ------------------------------------------------------------
# 🔽 PDF 최종본 저장
# ------------------------------------------------------------

def export_sheet_pdf(excel_path: str, sheet_name: str, out_pdf_path: str) -> bool:
    """엑셀 시트를 PDF로 내보내기 (pdf_utils.PDFExporter 사용)"""
    exporter = _PDFExporter(excel_path)
    try:
        return exporter.export_sheet(sheet_name, out_pdf_path)
    finally:
        exporter.close()

def merge_pdfs(pdf_list, out_pdf):
    """PDF 파일 병합 (pdf_utils.merge_pdfs 사용)"""
    _merge_pdfs(pdf_list, out_pdf)

def build_final_paths(excel_path: str, sample_no: str):
    yyyy = sample_no[1:3]  # 네 SN 규칙에 맞게 이미 함수가 있을 수도 있음
    yyyy_full = f"20{yyyy}"  # A26....이면 2026 같은

    h7 = read_input_h7(excel_path)
    p1_dir = FINAL_DONE_DIR
    p2_dir = os.path.join(PDF_BASE_DIR, yyyy_full, h7)

    os.makedirs(p1_dir, exist_ok=True)
    os.makedirs(p2_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    out_name = f"{base_name}.pdf"
    return os.path.join(p1_dir, out_name), os.path.join(p2_dir, out_name)


def build_final_path_water(excel_path: str, sample_no: str) -> str:
    """수질 PDF 최종 저장 경로.
    → PDF_BASE_DIR_WATER\YYYY년\업체명\파일명.pdf
    업체명은 엑셀 H7 셀 값(read_input_h7) 사용.
    """
    yyyy_full = f"20{sample_no[1:3]}"        # "26" → "2026"
    company   = read_input_h7(excel_path)    # 엑셀 H7 = 업체명
    out_dir   = os.path.join(PDF_BASE_DIR_WATER, f"{yyyy_full}년", company)
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    return os.path.join(out_dir, f"{base_name}.pdf")


def make_tab4_pdfs_water(excel_path: str, sample_no: str) -> dict:
    """수질 탭4 PDF 생성 + 최종본 저장.
    업1(anzeFile1): 분석일지 단독
    업2(anzeFile2): 대행기록부 단독
    최종본: 두 시트 병합 → build_final_path_water()
    """
    tmp_dir = PDF_TMP_DIR
    os.makedirs(tmp_dir, exist_ok=True)

    pdf_analy  = os.path.join(tmp_dir, f"{sample_no}__분석일지.pdf")
    pdf_record = os.path.join(tmp_dir, f"{sample_no}__대행기록부.pdf")

    export_sheet_pdf(excel_path, WATER_SHEET_ANALY,  pdf_analy)
    export_sheet_pdf(excel_path, WATER_SHEET_RECORD, pdf_record)

    # 최종본: 분석일지 + 대행기록부 병합
    merge_src = [p for p in (pdf_analy, pdf_record) if os.path.isfile(p)]
    final_tmp = os.path.join(tmp_dir, f"{sample_no}__FINAL.pdf")
    merge_pdfs(merge_src, final_tmp)

    import shutil
    p_out = build_final_path_water(excel_path, sample_no)
    shutil.copy2(final_tmp, p_out)

    return {
        "final_tmp" : final_tmp,
        "pdf_analy" : pdf_analy,    # #anzeFile1 업로드용
        "pdf_record": pdf_record,   # #anzeFile2 업로드용
        "p_out"     : p_out,
    }


def make_tab4_pdfs(excel_path: str, sample_no: str):
    tmp_dir = PDF_TMP_DIR
    os.makedirs(tmp_dir, exist_ok=True)

    # 업로드용 원본 시트 PDF
    pdf_analy = os.path.join(tmp_dir, f"{sample_no}__대기시료채취및분석일지.pdf")
    pdf_record = os.path.join(tmp_dir, f"{sample_no}__대기측정기록부.pdf")

    export_sheet_pdf(excel_path, "대기시료채취 및 분석일지", pdf_analy)
    export_sheet_pdf(excel_path, "대기측정기록부", pdf_record)

    # 커버(시험항목날짜서명) PDF
    pdf_cover = os.path.join(tmp_dir, f"{sample_no}__시험항목날짜서명.pdf")
    export_sheet_pdf(excel_path, "시험항목날짜서명", pdf_cover)

    # ✅ 업1만: 커버 + 분석일지 병합해서 새 파일 생성
    upload1 = os.path.join(tmp_dir, f"{sample_no}__분석일지.pdf")  # newFile1에 올릴 파일
    if os.path.isfile(pdf_cover) and os.path.isfile(pdf_analy):
        merge_pdfs([pdf_cover, pdf_analy], upload1)
    else:
        # 커버가 없거나 분석일지가 없으면 fallback (원래 파일이라도 올릴 수 있게)
        upload1 = pdf_analy

    # ✅ 업2는 "대기측정기록부 단독" 그대로
    upload2 = pdf_record

    # 최종 병합본(3시트) = 기존 그대로
    merge_src = [p for p in (pdf_cover, pdf_analy, pdf_record) if os.path.isfile(p)]
    final_tmp = os.path.join(tmp_dir, f"{sample_no}__FINAL.pdf")
    merge_pdfs(merge_src, final_tmp)

    # 두 경로에 복사(이 부분은 기존 유지)
    p1, p2 = build_final_paths(excel_path, sample_no)
    import shutil
    shutil.copy2(final_tmp, p1)
    shutil.copy2(final_tmp, p2)

    # ✅ 리턴에서 업로드용은 upload1/upload2로
    return {
        "final_tmp": final_tmp,
        "pdf_analy": upload1,     # 업1
        "pdf_record": upload2,    # 업2
        "p1": p1, "p2": p2
    }

def cleanup_tmp_pdfs(tmp_dir: str, sample_no: str, retries=6, delay=0.25):
    import glob
    pattern = os.path.join(tmp_dir, f"{sample_no}*.pdf")
    targets = glob.glob(pattern)

    for p in targets:
        ok = False
        for _ in range(retries):
            try:
                if os.path.isfile(p):
                    os.remove(p)
                ok = True
                break
            except PermissionError:
                time.sleep(delay)
            except Exception:
                time.sleep(delay)

        if not ok and os.path.isfile(p):
            print(f"⚠ tmp 삭제 실패(잠김): {p}")



def wait_file_selected(driver, css_sel, timeout=5.0):
    end = time.time() + timeout
    while time.time() < end:
        v = driver.execute_script("return arguments[0].value;", driver.find_element(By.CSS_SELECTOR, css_sel))
        if v and str(v).strip():
            return True
        time.sleep(0.1)
    return False


def upload_tab4_pdfs(driver, pdf_analy_path: str, pdf_record_path: str,
                     f1_sel: str = "#newFile1", f2_sel: str = "#newFile2"):
    """탭4 PDF 업로드.  대기: #newFile1/2,  수질: #anzeFile1/2"""

    f1 = driver.find_element(By.CSS_SELECTOR, f1_sel)
    f2 = driver.find_element(By.CSS_SELECTOR, f2_sel)

    driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", f1)
    driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", f2)

    # 1) 파일 선택(트리거: onchange)
    f1.send_keys(pdf_analy_path)
    if not wait_file_selected(driver, f1_sel, timeout=5.0):
        raise RuntimeError("⚠탭4 파일1 선택 확인 실패(newFile1)")

    f2.send_keys(pdf_record_path)
    if not wait_file_selected(driver, f2_sel, timeout=5.0):
        raise RuntimeError("⚠탭4 파일2 선택 확인 실패(newFile2)")

    # 2) setFile 처리 시간(서버 업로드/검증이 있을 수 있어 최소 대기)
    time.sleep(1.0)

# tab4_temp_save, tab4_comp_save → tab4_utils.py 로 이동


#============================================================
#                       탭2 리얼그리드 입력
#============================================================
def read_realgird_headers(driver, grid_root_css: str):
    """
    RealGrid 컨테이너(grid_root_css)에서 헤더 텍스트를 읽어
    [{'idx': 1, 'text': '대분류'}, ...] 형태로 반환
    idx는 tbody td:nth-child(idx) 와 맞추는 용도
    """
    js = r"""
    const root = document.querySelector(arguments[0]);
    if (!root) return null;

    const tries = [
      "div.rg-header table thead tr:last-child th",
      "div.rg-header table thead tr:last-child td",
      "table.rg-header-table thead tr:last-child th",
      "table.rg-header-table thead tr:last-child td",
      "thead tr:last-child th",
      "thead tr:last-child td"
    ];

    for (const sel of tries) {
      const els = root.querySelectorAll(sel);
      if (els && els.length) {
        return Array.from(els).map((el, i) => ({
          idx: i + 1,
          text: (el.innerText || "").trim()
        }));
      }
    }

    // fallback: 헤더를 못 찾으면 전체 thead 긁기
    const els = root.querySelectorAll("thead th, thead td");
    if (els && els.length) {
      return Array.from(els).map((el, i) => ({
        idx: i + 1,
        text: (el.innerText || "").trim()
      }));
    }
    return [];
    """
    headers = driver.execute_script(js, grid_root_css)
    if not headers:
        return []
    # 빈 텍스트 제거(가끔 공백 헤더가 섞임)
    return [h for h in headers if h.get("text")]


def build_header_map(headers):
    """
    headers(list) -> dict로 변환
    '단위'가 2번 나오므로
      - 시료채취량단위
      - 흡인속도단위
    로 분리해서 맵핑해줌
    """
    def find_idx(text):
        for h in headers:
            if h["text"] == text:
                return h["idx"]
        return None

    def find_next_idx_after(text, after_idx):
        for h in headers:
            if h["idx"] > after_idx and h["text"] == text:
                return h["idx"]
        return None

    m = {}
    # 기본 컬럼들
    for key in ["대분류","중분류","측정항목","시료채취량","흡인속도",
                "측정일(시작)","시작시간","측정일(종료)","종료시간","비고"]:
        m[key] = find_idx(key)

    # 단위 2개 분리
    vol_idx = m.get("시료채취량")
    spd_idx = m.get("흡인속도")

    if vol_idx:
        m["시료채취량단위"] = find_next_idx_after("단위", vol_idx)
    if spd_idx:
        m["흡인속도단위"] = find_next_idx_after("단위", spd_idx)

    return m


def rg_find_row_by_item(driver, grid_root_css, item_name):
    """
    tbody의 td div.rg-renderer 텍스트 중 item_name과 같은 행 찾기 (1-based row index)
    """
    js = r"""
    const root = document.querySelector(arguments[0]);
    const name = arguments[1];
    if (!root) return null;
    const body = root.querySelector('.rg-body table tbody');
    if (!body) return null;
    const rows = body.querySelectorAll('tr');
    for (let i=0;i<rows.length;i++){
      const tds = rows[i].querySelectorAll('td div.rg-renderer');
      for (const d of tds){
        if ((d.innerText||'').trim() === name) return i+1;
      }
    }
    return null;
    """
    return driver.execute_script(js, grid_root_css, item_name)


def rg_set_cell(driver, grid_root_css, row_idx, col_idx, value):
    """
    셀 클릭 → Ctrl+A → 값 → Enter
    """
    # td 선택 후 클릭 (RealGrid는 td 안에 div.rg-renderer가 보통 있음)
    cell_css = f"{grid_root_css} .rg-body table tbody tr:nth-child({row_idx}) td:nth-child({col_idx})"
    el = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, cell_css)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.1)
    ActionChains(driver).move_to_element(el).click(el).pause(0.05).perform()
    time.sleep(0.05)

    # 입력
    ActionChains(driver)\
        .key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL)\
        .send_keys(Keys.BACKSPACE)\
        .send_keys(str(value))\
        .send_keys(Keys.ENTER)\
        .perform()
    time.sleep(0.05)

def read_realgird_values(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    if "입력(분석값)" in wb.sheetnames:
        ws = wb["입력(분석값)"]
    elif "입력" in wb.sheetnames:
        ws = wb["입력"]
    else:
        ws = wb[wb.sheetnames[0]]

    # 1행에서 필요한 컬럼 찾기
    header_map = {}
    for col in range(1, 60):
        v = ws.cell(row=1, column=col).value
        if not v:
            continue
        t = str(v).strip()
        header_map[t] = col

    def col_of(name):
        return header_map.get(name)

    c_start = col_of("측정시작")
    c_end   = col_of("측정 종료") or col_of("측정종료")
    c_spd   = col_of("시료흡인속도")
    c_vol   = col_of("시료채취량")

    if not (c_start and c_end and c_spd and c_vol):
        raise RuntimeError(f"입력(분석값) 1행 헤더 매칭 실패: {list(header_map.keys())}")

    per_item = {}

    # B2~B64에 항목명
    for r in range(2, 65):
        
        # ✅ A열이 비어있으면(측정 안한 항목) 패스
        a_val = ws.cell(row=r, column=1).value
        if a_val is None or str(a_val).strip() == "":
            continue
        
        item = ws.cell(row=r, column=2).value
        if not item:
            continue
        item = str(item).strip()
        if not item:
            continue

        st = ws.cell(row=r, column=c_start).value
        et = ws.cell(row=r, column=c_end).value
        spd = ws.cell(row=r, column=c_spd).value
        vol = ws.cell(row=r, column=c_vol).value

        vol_unit = "Sm³" if item in sm3_items else "L"

        # 시료채취량: 사이트가 소수점 4자리까지만 지원 → 반올림
        vol_str = "" if vol is None else str(vol).strip()
        if vol_str:
            try:
                vol_str = f"{float(vol_str.replace(',', '')):.4f}"
            except ValueError:
                pass

        per_item[item] = {
            "시작시간": "" if st is None else str(st).strip(),
            "종료시간": "" if et is None else str(et).strip(),
            "흡인속도": "" if spd is None else str(spd).strip(),
            "시료채취량": vol_str,
            "채취량단위": vol_unit,
            "흡인속도단위": "L/min",
        }

    return per_item

def fill_tab2_realgird(driver, excel_path, sample_no, grid_root_css):
    date_str = sample_to_datestr(sample_no)
    if not date_str:
        raise RuntimeError("시료번호에서 날짜 파싱 실패")

    # 컬럼 찾기
    c_item = rg_find_col(driver, grid_root_css, ["측정항목"])
    c_vol  = rg_find_col(driver, grid_root_css, ["시료채취량"])
    c_spd  = rg_find_col(driver, grid_root_css, ["흡인속도"])
    c_sd   = rg_find_col(driver, grid_root_css, ["측정일(시작)"])
    c_st   = rg_find_col(driver, grid_root_css, ["시작시간"])
    c_ed   = rg_find_col(driver, grid_root_css, ["측정일(종료)"])
    c_et   = rg_find_col(driver, grid_root_css, ["종료시간"])

    unit_cols = rg_find_cols(driver, grid_root_css, "단위")
    if len(unit_cols) >= 2:
        c_vol_u = unit_cols[0]
        c_spd_u = unit_cols[1]
    else:
        c_vol_u = unit_cols[0] if unit_cols else (c_vol + 1 if c_vol is not None else None)
        c_spd_u = (c_spd + 1) if c_spd is not None else None

    need = [c_item, c_vol, c_vol_u, c_spd, c_spd_u, c_sd, c_st, c_ed, c_et]
    if any(x is None for x in need):
        raise RuntimeError(f"RealGrid 컬럼 탐색 실패: {rg_dump_headers(driver, grid_root_css)}")

    # 엑셀 -> 항목별 값
    per_item = read_realgird_values(excel_path)

    # 엑셀에 있는 항목만 처리(빠름)
    for item_name, v in per_item.items():
        tr = rg_find_tr_by_item(driver, grid_root_css, c_item, item_name)
        if not tr:
            continue

        st = v.get("시작시간", "")
        et = v.get("종료시간", "")
        vol = v.get("시료채취량", "")
        vol_u = v.get("채취량단위", "L")
        spd = v.get("흡인속도", "")
        spd_u = v.get("흡인속도단위", "L/min")

        # 예외별 붙여넣기 구성
        if item_name in SKIP_VOL_AND_SPEED:
            start_col = c_sd
            values = [date_str, st, date_str, et]   # 날짜/시간만
        elif item_name in SKIP_SPEED_ONLY:
            start_col = c_vol
            values = [vol, vol_u, "", "", date_str, st, date_str, et]  # 흡인 2칸 비움
        else:
            start_col = c_vol
            values = [vol, vol_u, spd, spd_u, date_str, st, date_str, et]

        rg_paste_to_tr(driver, tr, start_col, values)

    print("▶ RealGrid 붙여넣기 입력 완료")


# =====================================================================
# 메인
# =====================================================================

def main():
    """매체(대기/수질) 선택 후 해당 흐름으로 분기."""
    print("=== 자동 입력 시작 ===")
    media = input("매체 선택 (1=대기 / 2=수질) [기본:1]: ").strip() or "1"
    if media == "2":
        _main_water()
    else:
        _main_air()


# ══════════════════════════════════════════════════════════════
# 수질 흐름  ─  탭4(측정분석결과)만 입력
# ══════════════════════════════════════════════════════════════
def _main_water():
    """수질 흐름: 탭4 진입 → PDF 생성/업로드 → 저장만 수행 (그리드 입력 없음)"""
    login_id = input("측정인 아이디: ").strip()
    login_pw = input("측정인 비밀번호: ").strip()

    raw = input("처리할 수질 시료번호 (쉼표/공백/줄바꿈 구분):\n").strip()
    target_samples = parse_sample_input(raw)
    if not target_samples:
        print("⚠ 시료번호 없음 → 종료"); return

    # 수질 NAS에서 파일 찾기 (NAS_BASE_WATER\YYYY년\업체명\파일.xlsm)
    items = []
    for sn in target_samples:
        p = find_sample_file_in_water_nas(sn)
        if not p:
            print(f"⚠ 수질 NAS에서 파일 못 찾음: {sn} → 스킵")
            continue
        items.append({"sample": sn, "path": p})

    if not items:
        print("⚠ 처리할 시료 없음 → 종료"); return

    print(f"총 {len(items)}개 수질 시료 처리 예정")

    do_pdf_final = ask_yesno("PDF 생성/업로드까지 할래? (예/아니오) [기본: 예]: ", default_yes=True)

    # 드라이버 시작 + 수질 사이트 로그인
    driver = init_driver()
    login(driver, login_id, login_pw, field_url=FIELD_URL_WATER)

    for item in items:
        sample = item["sample"]
        path   = item["path"]

        print(f"\n{'='*51}")
        print(f"▶ 수질 시료: {sample}")
        print(f"{'='*51}")

        if not open_detail_by_sample(driver, sample):
            print("❌ 상세페이지 진입 실패 → 다음 시료로")
            continue

        # 탭4 클릭 + 로딩 대기
        safe_click(driver, TAB4_SELECTOR)
        wait_el(driver, "#smpl_rcpt_dt", timeout=10)

        pdfs = None
        if do_pdf_final:
            print("▶ PDF 생성/업로드중")
            pdfs = make_tab4_pdfs_water(path, sample)
            upload_tab4_pdfs(
                driver,
                pdfs["pdf_analy"],
                pdfs["pdf_record"],
                f1_sel=WATER_FILE_BTN1,
                f2_sel=WATER_FILE_BTN2,
            )
            tab4_comp_save(driver)
            print(f"✅ 완료  →  {pdfs['p_out']}")
        else:
            tab4_temp_save(driver)
            print("  임시저장")

        if pdfs:
            cleanup_tmp_pdfs(PDF_TMP_DIR, sample)

        back_to_list(driver)

    print("\n=== 수질 처리 완료 ===")
    try:
        input("엔터 누르면 종료...")
    except EOFError:
        pass


# ══════════════════════════════════════════════════════════════
# 대기 흐름  ─  기존 그대로
# ══════════════════════════════════════════════════════════════
def _main_air():
    job = input("작업 선택 (1=측정인 입력/저장/PDF/백데이터 / 2=백데이터만(로그인X)) [기본:1]: ").strip() or "1"

    login_id = ""
    login_pw = ""
    if job == "1":
        login_id = input("측정인 아이디: ").strip()
        login_pw = input("측정인 비밀번호: ").strip()

    mode = input("모드 선택 (1=시료번호 직접입력 / 2=팀+날짜 자동추출) [기본:2]: ").strip() or "2"

    do_tab1       = ask_yesno("현장측정정보(탭1) 입력할래? (예/아니오) [기본: 예]: ",          default_yes=True)
    do_tab2       = ask_yesno("시료채취/측정정보(탭2) 입력할래? (예/아니오) [기본: 예]: ",      default_yes=True)
    do_pdf_upload = ask_yesno("탭2 PDF 생성/업로드까지 할래? (예/아니오) [기본: 예]: ",        default_yes=True)
    do_backdata   = ask_yesno("백데이터(수분 CSV + THC CSV/FID)? (예/아니오) [기본: 예]: ",   default_yes=True)
    do_tab4       = ask_yesno("측정분석결과(탭4) 입력할래? (예/아니오) [기본: 아니오]: ",      default_yes=False)
    do_pdf_final  = ask_yesno("PDF최종본 생성(탭4용) 할래? (예/아니오) [기본: 아니오]: ",      default_yes=False)

    if job == "2":
        do_tab1 = do_tab2 = do_pdf_upload = do_tab4 = do_pdf_final = False
        do_backdata = True

    if not any([do_tab1, do_tab2, do_backdata, do_tab4]):
        print("⚠ 실행할 항목이 없어 종료"); return

    # 시료 목록 생성
    date_groups = {}

    if mode == "1":
        raw = input("처리할 시료번호 입력 (쉼표/공백/줄바꿈 구분):\n").strip()
        target_samples = parse_sample_input(raw)
        if not target_samples:
            print("⚠ 시료번호 없음 → 종료"); return

        for sn in target_samples:
            p = find_sample_file_in_nas(sn)
            if not p:
                print(f"⚠ NAS에서 파일 못 찾음: {sn} → 스킵"); continue
            dust = is_fugitive_dust_file(p)
            ds = sample_to_datestr(sn)
            if not ds:
                print(f"⚠ 날짜 파싱 실패: {sn} → 스킵"); continue
            date_groups.setdefault(ds, []).append({"sample": sn, "path": p, "dust": dust})

        if not date_groups:
            print("⚠ 처리 가능한 시료 없음 → 종료"); return
        total = sum(len(v) for v in date_groups.values())
        print(f"총 {total}개 시료(날짜 {len(date_groups)}개 그룹) 처리 예정")

    else:
        team_input = input("팀 번호(쉼표 구분, 예: 1,3 / 미입력 시 전체): ").strip()
        date_str = input("날짜(YYYY-MM-DD): ").strip()
        
        # 콤마로 구분된 팀 번호를 리스트로 파싱
        team_nos = [t.strip() for t in team_input.split(",")] if team_input else []
        
        # 숫자 검증 (선택사항, 빈 리스트면 전체 조회)
        for t in team_nos:
            if not t.isdigit():
                print(f"팀 번호 오류: '{t}'는 숫자가 아닙니다."); return

        samples = extract_samples_from_nas(team_nos, date_str)
        
        팀_표시 = ",".join(team_nos) if team_nos else "전체"
        print(f"▶ 팀({팀_표시}) 총 {len(samples)}개 시료 자동 입력 예정")
        date_groups = {date_str: samples}

    # 드라이버 / 로그인
    driver = None
    if job == "1":
        driver = init_driver()
        login(driver, login_id, login_pw)
    else:
        print("▶ 백데이터만 모드 → 브라우저 스킵")

    # 날짜 그룹별 처리
    for date_str, day_samples in date_groups.items():
        print(f"\n{'='*51}")
        print(f"▶ 날짜 검색: {date_str}  | 대상 시료: {len(day_samples)}개")
        print(f"{'='*51}")

        if job == "1":
            set_date_js(driver, "#search_meas_start_dt", date_str)
            set_date_js(driver, "#search_meas_end_dt",   date_str)
            safe_click(driver, "#btnSearch")
            wait(2)
        else:
            print("▶ 백데이터만 모드 → 날짜검색 스킵")

        for item in day_samples:
            sample  = item["sample"]
            path    = item["path"]
            is_dust = item["dust"]

            print(f"\n대기 시료: {sample}  | 비산먼지: {is_dust}")

            # 백데이터
            if do_backdata and not is_dust:
                try:
                    export_backdata_moist_thc(path, sample)
                except Exception as e:
                    print(f"⚠ 백데이터 실패: {e}")
            elif do_backdata and is_dust:
                print("▶ 비산먼지 → 백데이터 스킵")

            if job != "1":
                continue

            if not open_detail_by_sample(driver, sample):
                print("❌ 상세페이지 실패 → 다음 시료로"); continue

            excel = parse_measuring_record(path, sample)

            # 탭1
            if do_tab1:
                safe_click(driver, "#ui-id-1")
                fill_tab1(driver, excel, is_dust)
            else:
                print("▶ 탭1 스킵")

            # 탭2
            if do_tab2:
                safe_click(driver, "#ui-id-2")
                fill_tab2(driver, excel, is_dust)
                set_date_js(driver, SEL_DATE, date_str)

                if not is_dust:
                    fill_facility_rows(driver, parse_facility_from_excel(path))
                else:
                    print("▶ 비산먼지 → 배출시설 스킵")
                    ensure_gas_flow_checkbox_checked(driver)

                write_sampler_comment(driver)

                GRID_ROOT = "#measGridAnalySampAnzeDataAirItemList1"
                rg2_fill_measure_grid_api(driver, sample, read_realgird_values(path))

                # PDF 업로드
                did_pdf  = False
                pdf_path = None
                if do_pdf_upload:
                    try:
                        pdf_path = export_pdf_from_excel(path, sample, is_dust)
                        print(f"▶ PDF 생성 완료: {pdf_path}")
                        if not trigger_file_dialog(driver, timeout=8):
                            raise RuntimeError("파일추가 트리거 실패")
                        time.sleep(0.8)
                        pick_file_in_open_dialog(pdf_path, timeout=30)
                        print("✅ 파일 선택 완료")
                        did_pdf = True
                    except Exception as e:
                        print(f"❌ PDF 실패: {e}")
                else:
                    print("▶ PDF 업로드 스킵")
                time.sleep(1)

                draft_by_time  = _should_draft_by_sampling_end(date_str, excel.get("채취끝", ""))
                use_final_save = did_pdf and not draft_by_time
                print(f"▶ 탭2 {'✅입력완료' if use_final_save else '⚠임시저장'}")
                ok = save_tab2(driver, use_final_save=use_final_save)

                if ok and did_pdf and pdf_path and os.path.isfile(pdf_path):
                    try:
                        os.remove(pdf_path)
                        print(f"🗑 PDF 삭제: {pdf_path}")
                    except Exception as e:
                        print(f"⚠ PDF 삭제 실패: {e}")
            else:
                print("▶ 탭2 스킵")

            # 탭4
            if do_tab4:
                safe_click(driver, TAB4_SELECTOR)
                wait_el(driver, "#smpl_rcpt_dt", timeout=10)

                print("▶ 탭4 데이터 읽는중")
                tab4_meta = read_tab4_from_macro_xlsm(sample)
                print("✅ 탭4 데이터 읽기 완료")

                fill_tab4_dates(driver, sample, tab4_meta)
                fill_tab4_grid_only(driver, tab4_meta["rows"])

                pdfs = None
                if do_pdf_final:
                    print("▶ PDF 탭4 생성중")
                    pdfs = make_tab4_pdfs(path, sample)
                    upload_tab4_pdfs(driver, pdfs["pdf_analy"], pdfs["pdf_record"])
                    print("✅ PDF 탭4 완료")

                if do_pdf_final:
                    tab4_comp_save(driver)
                    print("✅ 탭4 입력완료")
                else:
                    tab4_temp_save(driver)
                    print("⚠ 탭4임시저장")

                if pdfs:
                    cleanup_tmp_pdfs(PDF_TMP_DIR, sample)

            back_to_list(driver)

    print("\n=== 대기 처리 완료 ===")
    try:
        input("엔터 누르면 종료...")
    except EOFError:
        pass


if __name__=="__main__":
    try:
        main()
    except Exception as e:
        # 공용 에러 로그에 남기고, 원래대로 예외는 다시 올려서 콘솔/GUI에서도 확인 가능하게 유지
        log_error("eco_input.main", e)
        raise