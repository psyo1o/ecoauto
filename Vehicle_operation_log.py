# -*- coding: utf-8 -*-
"""
차량 운행일지 검토 모듈
- 발송대장 / 주행기록부 / 환경기술인력 파일 3개를 비교
- 날짜별 출퇴근 시간, 주행거리, 근무시간 정합성 검사
- 주 52시간 초과 여부 보고서 생성
- GUI와 로직이 한 파일에 통합
"""
import re
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import warnings
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)

DEFAULT_DAEJANG_DIR = r"\\192.168.10.163\측정팀\0.시료접수발송대장"
DEFAULT_DRIVE_LOG_DIR = r"\\192.168.10.163\측정팀\10.검토\4.차랑운행일지 검토"
DEFAULT_ENGINEER_DIR = r"\\192.168.10.163\측정팀\2.성적서"

# ============================================================
# 공통 유틸 모듈 (모듈화)
# ============================================================
from data_utils import (
    normalize_company, normalize_plate, normalize_name, normalize_names_cell,
    parse_time, parse_time_range_text, parse_ymd_date
)

###############################################################
# (추가) 주 52시간 계산 유틸
###############################################################

def _time_to_minutes(t: dt.time):
    return t.hour * 60 + t.minute


def calc_work_hours_with_break(start_t: dt.time, end_t: dt.time, break_hours: float = 1.0):
    """
    출발~도착 -> 근무시간(시간, 소수) 계산
    - 자정 넘어가면 +24h 처리
    - 휴게시간 break_hours 차감
    - 0 미만이면 0
    """
    if not start_t or not end_t:
        return 0.0
    s = _time_to_minutes(start_t)
    e = _time_to_minutes(end_t)
    if e < s:
        e += 24 * 60
    mins = e - s
    hours = mins / 60.0
    hours -= break_hours
    return max(0.0, hours)


def week_start_monday(day: dt.date):
    """월요일 시작 주(월~일)"""
    return day - dt.timedelta(days=day.weekday())


def build_weekly_52_report(drive_list, break_hours=1.0):
    """
    drive_list(운행일지) 기반으로 인력별 주간합계 계산.
    L열 복수인력인 경우: 각 인력에게 동일한 (휴게차감된) 시간 부여
    """
    agg = {}  # (name, week_start) -> hours

    for d in drive_list:
        day = d.get("date")
        st = d.get("start")
        ed = d.get("end")
        names = d.get("engineers", [])  # 리스트
        if not day or not st or not ed or not names:
            continue

        wk = week_start_monday(day)
        h = calc_work_hours_with_break(st, ed, break_hours=break_hours)

        for nm in names:
            key = (nm, wk)
            agg[key] = agg.get(key, 0.0) + h

    rows = []
    for (nm, wk), h in agg.items():
        rows.append({
            "engineer": nm,
            "week_start": wk,
            "week_end": wk + dt.timedelta(days=6),
            "hours": h,
            "status": "❌ 초과" if h > 52 else "✅ 준수"
        })

    rows.sort(key=lambda x: (x["engineer"], x["week_start"]))
    return rows


###############################################################
# 1) 시료접수발송대장 파싱
###############################################################

def parse_daejang(path):
    """대장 파일: SN → 날짜/팀/시간을 추출 (3행~35행만 처리)"""
    xls = pd.ExcelFile(path)
    results = []

    for sheet in xls.sheet_names:
        # 시트명: 1일, 2일, … 31일 만 처리
        if not re.match(r"^\d{1,2}일$", sheet):
            continue

        df = pd.read_excel(path, sheet_name=sheet, header=None)

        # ----------- ★ 여기 핵심: 3~35행 범위 강제 -----------
        max_row = min(len(df), 35)   # 파일이 35행보다 작으면 그만큼만
        for r in range(2, max_row):  # r=2 → 엑셀 3행
        # ---------------------------------------------------

            sn = str(df.iloc[r, 0]).strip()      # A열 SN
            t_raw = str(df.iloc[r, 1]).strip()   # B열 시간
            comp_raw = str(df.iloc[r, 3]).strip() # D열 업소명

            if sn == "" or sn.lower() == "nan":
                continue

            start_t, end_t = parse_time_range_text(t_raw)
            company = normalize_company(comp_raw)

            results.append({
                "sn": sn,
                "start": start_t,
                "end": end_t,
                "company": company,
                "sheet": sheet,
                "row": r + 1
            })

    return results


###############################################################
# 2) SN → 날짜/팀/시료번호 파싱
###############################################################

def parse_sn(sn: str):
    """
    SN 예: A2512063-01
    → date=2025-12-06, team=3, idx=1
    """
    try:
        sn = sn.strip()

        yy = int(sn[1:3])
        mm = int(sn[3:5])
        dd = int(sn[5:7])
        team = int(sn[7])
        idx = int(sn.split("-")[1])

        date = dt.date(2000 + yy, mm, dd)
        return date, team, idx
    except:
        return None, None, None


###############################################################
# 3) 차량운행일지 파싱 (1개 시트 구조 확정)
###############################################################

def parse_drive_log(path):
    """
    차량운행일지 실제 구조 (확정):
    A 운행일
    B 차량번호
    D 경유지(업소명)
    F 출발시간
    G 도착시간
    L 사용인력  (예: 박수영,유경수)
    """

    df = pd.read_excel(path, header=None)
    results = []

    # row 0~1 은 헤더 → row=2부터 데이터
    for r in range(2, len(df)):
        raw_date      = df.iloc[r, 0]
        raw_plate     = df.iloc[r, 1]
        raw_company   = df.iloc[r, 3]
        raw_start     = df.iloc[r, 5]
        raw_end       = df.iloc[r, 6]
        raw_engineer  = df.iloc[r, 11]

        # 날짜 처리
        if isinstance(raw_date, dt.datetime):
            day = raw_date.date()
        elif isinstance(raw_date, dt.date):
            day = raw_date
        else:
            day = parse_ymd_date(raw_date)
            if day is None:
                continue

        engineers = normalize_names_cell(raw_engineer)

        results.append({
            "date": day,
            "plate": normalize_plate(raw_plate),
            "company": normalize_company(raw_company),
            "start": parse_time(raw_start),
            "end": parse_time(raw_end),
            "engineers": engineers,                 # ★ 리스트
            "engineer": ",".join(engineers),        # ★ 기존 compare_all 호환용 문자열
            "row": r + 1
        })

    return results


###############################################################
# 4) 기술인력파일 파싱 (가장 최근 날짜 적용 방식)
###############################################################

def parse_engineer_file(path):
    """
    기술인력파일 (1팀~5팀):
    A 날짜
    U 사수
    V 부사수
    W 차량번호

    날짜가 비어있지 않은 행 = “적용 시작점”
    날짜가 없는 행 = 바로 위의 날짜 적용
    """

    xls = pd.ExcelFile(path)
    results = []

    team_sheets = [s for s in xls.sheet_names if re.match(r"^\d팀$", s)]

    for sheet in team_sheets:
        df = pd.read_excel(path, sheet_name=sheet, header=None)

        last_date = None

        for r in range(len(df)):
            raw_date = df.iloc[r, 0]

            # 날짜 있으면 last_date 갱신
            if isinstance(raw_date, (dt.date, dt.datetime)):
                last_date = raw_date.date() if isinstance(raw_date, dt.datetime) else raw_date

            if last_date is None:
                continue   # 아직 시작점 못찾음

            # 인력 정보
            eng1 = normalize_name(df.iloc[r, 20])
            eng2 = normalize_name(df.iloc[r, 21])
            plate = normalize_plate(df.iloc[r, 22])

            engineers = []
            if eng1 != "":
                engineers.append(eng1)
            if eng2 != "":
                engineers.append(eng2)

            results.append({
                "date": last_date,
                "engineers": engineers,
                "plate": plate,
                "team": sheet,  # "3팀"
                "row": r + 1
            })

    return results


###############################################################
# 5) SN 날짜에 맞는 기술인력 데이터 찾기 (가장 최근 날짜)
###############################################################

def find_engineer_for_date(eng_list, sn_date, team):
    """
    기술인력 중 같은 팀이면서
    SN 날짜보다 이전 또는 같은 날짜 중 가장 최근 날짜 선택
    """
    candidates = [e for e in eng_list if e["team"] == f"{team}팀" and e["date"] <= sn_date]

    if not candidates:
        return None

    candidates.sort(key=lambda x: x["date"], reverse=True)
    return candidates[0]


###############################################################
# 6) 차량운행일지 중 차량번호 + 날짜로 필터링
###############################################################

def find_drive_for_date(drive_list, sn_date, plate):
    """같은 날짜 + 같은 차량번호"""
    return [d for d in drive_list if d["date"] == sn_date and d["plate"] == plate]

###############################################################
# 7) 대장 / 운행일지 / 기술인력 비교 통합 로직 (기술인력 중복 포함)
###############################################################

def is_engineer_duplicate_in_drive(drive_list, sn_date, matched_log):
    """
    쉬운버전 중복판정:
    - 같은 날짜(sn_date)의 운행일지 중
    - 이번에 매칭된 운행(matched_log) 말고 다른 운행(row가 다른 것)에도
    - 동일 인력이 한 명이라도 등장하면 → True(중복)
    """
    if not matched_log:
        return False

    base_names = set(matched_log.get("engineers", []))
    if not base_names:
        return False

    base_row = matched_log.get("row")

    for d in drive_list:
        if d.get("date") != sn_date:
            continue
        # 자기 자신(같은 행) 제외
        if base_row is not None and d.get("row") == base_row:
            continue

        other_names = set(d.get("engineers", []))
        if base_names & other_names:
            return True

    return False


def compare_all(dae, drive, eng):
    rows = []

    for d in dae:
        sn = d["sn"]
        comp = d["company"]
        st = d["start"]
        ed = d["end"]

        sn_date, sn_team, sn_idx = parse_sn(sn)

        row_status = {
            "SN": sn,
            "업소명": comp,
            "대장시간": f"{st}~{ed}",
            "운행일지_일치여부": "",
            "운행일지_출발시간": "",
            "운행일지_도착시간": "",
            "운행일지_차량": "",
            "운행일지_기술인력": "",
            "기술인력파일_기술인력": "",
            "기술인력파일_차량": "",
            "구분": ""
        }

        if sn_date is None:
            row_status["구분"] = "SN오류"
            rows.append(row_status)
            continue

        ########################################################
        # ① 기술인력파일 검색
        ########################################################

        eng_match = find_engineer_for_date(eng, sn_date, sn_team)

        if eng_match:
            row_status["기술인력파일_차량"] = eng_match["plate"]
            row_status["기술인력파일_기술인력"] = ",".join(eng_match["engineers"])
            target_plate = eng_match["plate"]
        else:
            target_plate = ""
            row_status["구분"] = "기술인력없음"



        ########################################################
        # ② 차량운행일지 검색
        ########################################################

        drv_candidates = find_drive_for_date(drive, sn_date, target_plate)

        if not drv_candidates:
            row_status["운행일지_일치여부"] = "없음"
            if row_status["구분"] == "":
                row_status["구분"] = "운행일지없음"
            rows.append(row_status)
            continue

        ########################################################
        # ③ 시간 포함 여부 검사
        ########################################################

        matched = False
        for log in drv_candidates:
            if st and ed and log["start"] and log["end"]:
                if log["start"] <= st and ed <= log["end"]:
                    matched = True
                    row_status["운행일지_일치여부"] = "OK"
                    row_status["운행일지_출발시간"] = log["start"]
                    row_status["운행일지_도착시간"] = log["end"]
                    row_status["운행일지_차량"] = log["plate"]
                    row_status["운행일지_기술인력"] = log.get("engineer", "")

                    ########################################################
                    # ★ 기술인력 중복 체크 (같은 날짜 기준으로 다른 팀과 중복)
                    ########################################################
                    # ✅ (쉬운버전) 운행일지 기준 기술인력 중복 체크
                    # - 같은 날짜 운행일지 전체에서
                    # - 이번 log 말고 다른 운행에도 같은 인력이 있으면 중복
                    if is_engineer_duplicate_in_drive(drive, sn_date, log):
                        row_status["구분"] = "기술인력중복"
                    else:
                        if row_status["구분"] == "":
                                row_status["구분"] = "OK"
                    break

        if not matched:
            row_status["운행일지_일치여부"] = "불일치"
            if row_status["구분"] == "":
                row_status["구분"] = "시간불일치"

        rows.append(row_status)

    return rows


###############################################################
# 8) 결과파일 생성 (+ 주52시간 시트 추가)
###############################################################

def write_report(rows, drive_list):
    save_dir = r"\\192.168.10.163\측정팀\10.검토\4.차랑운행일지 검토"

    # ✅ 운행일지 첫열(날짜) 기준으로 연월(YYYYMM) 만들기
    ym = dt.datetime.now().strftime("%Y%m")  # 기본값(운행일지 비어있을 때)
    if drive_list:
        v = drive_list[0].get("date")  # parse_drive_log에서 date 키로 넣는다고 가정
        try:
            if isinstance(v, (dt.date, dt.datetime)):
                ym = v.strftime("%Y%m")
            else:
                # 문자열이면 숫자만 뽑아서 YYYYMMDD / YYYY-MM-DD / YYYY.MM.DD 등 대응
                digits = re.sub(r"\D", "", str(v))
                if len(digits) >= 6:
                    ym = digits[:6]
        except:
            pass

    out_path = f"{save_dir}\\차량운행일지_검토결과_{ym}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "결과"
    # ✅ 첫 번째 시트(결과) 1행 1열 틀고정
    ws.freeze_panes = "B2"
    header = [
        "SN","업소명","대장시간",
        "운행일지_일치여부",
        "운행일지_출발시간","운행일지_도착시간",
        "운행일지_차량","운행일지_기술인력",
        "기술인력파일_기술인력","기술인력파일_차량",
        "구분"
    ]
    ws.append(header)

    red  = PatternFill(start_color="FF9999", fill_type="solid")
    blue = PatternFill(start_color="99CCFF", fill_type="solid")
    gray = PatternFill(start_color="DDDDDD", fill_type="solid")

    for row in rows:
        ws.append([
            row["SN"],
            row["업소명"],
            row["대장시간"],
            row["운행일지_일치여부"],
            row["운행일지_출발시간"],
            row["운행일지_도착시간"],
            row["운행일지_차량"],
            row["운행일지_기술인력"],
            row["기술인력파일_기술인력"],
            row["기술인력파일_차량"],
            row["구분"]
        ])

    # 색칠
    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, 11).value
        fill = gray
        if status == "OK":
            fill = blue
        elif status != "":
            fill = red
        for c in range(1, 12):
            ws.cell(r, c).fill = fill

    # ---------------------------------------------------------
    # ✅ 주52시간 시트 (운행일지 기반 / 월~일 / 휴게 1시간 차감)
    # ---------------------------------------------------------
    ws52 = wb.create_sheet("주52시간")
    
    # ✅ 두 번째 시트(주52시간)도 틀고정
    ws52.freeze_panes = "B2"
    
    ws52.append(["인력", "주시작(월)", "주종료(일)", "주간합계(시간)", "준수여부"])

    weekly = build_weekly_52_report(drive_list, break_hours=1.0)

    red_fill  = PatternFill(start_color="FF9999", fill_type="solid")
    blue_fill = PatternFill(start_color="99CCFF", fill_type="solid")

    for item in weekly:
        ws52.append([
            item["engineer"],
            item["week_start"].strftime("%Y-%m-%d"),
            item["week_end"].strftime("%Y-%m-%d"),
            round(item["hours"], 2),
            item["status"]
        ])

    for r in range(2, ws52.max_row + 1):
        status = ws52.cell(r, 5).value
        fill = red_fill if status == "❌ 초과" else blue_fill
        for c in range(1, 6):
            ws52.cell(r, c).fill = fill

    wb.save(out_path)
    return out_path


###############################################################
# 9) GUI
###############################################################
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import queue


def gui_start():
    root = tk.Tk()
    root.title("차량운행일지 검토")

    daejang_path = tk.StringVar()
    drive_path = tk.StringVar()
    eng_path = tk.StringVar()
    status = tk.StringVar(value="파일을 선택하세요.")

    def choose_daejang():
        p = filedialog.askopenfilename(
            title="대장파일 찾기",
            initialdir=DEFAULT_DAEJANG_DIR,
            filetypes=[("Excel Files","*.xlsx;*.xls;*.xlsm")]
        )
        if p: daejang_path.set(p)

    def choose_drive():
        p = filedialog.askopenfilename(
            title="운행일지 파일 찾기",
            initialdir=DEFAULT_DRIVE_LOG_DIR,
            filetypes=[("Excel Files","*.xlsx;*.xls;*.xlsm")]
        )
        if p: drive_path.set(p)

    def choose_eng():
        p = filedialog.askopenfilename(
            title="기술인력 파일 찾기",
            initialdir=DEFAULT_ENGINEER_DIR,
            filetypes=[("Excel Files","*.xlsx;*.xls;*.xlsm")]
        )
        if p: eng_path.set(p)


    # 작업 결과를 UI 스레드로 전달하기 위한 큐
    q = queue.Queue()

    def worker(dae_path, drv_path, eng_path):
        """무거운 작업은 여기서 실행 (백그라운드 스레드)."""
        try:
            q.put(("PROG", "1/5 대장 읽는 중…"))
            dae = parse_daejang(dae_path)

            q.put(("PROG", "2/5 운행일지 읽는 중…"))
            drv = parse_drive_log(drv_path)

            q.put(("PROG", "3/5 기술인력파일 읽는 중…"))
            eng = parse_engineer_file(eng_path)

            q.put(("PROG", "4/5 비교 로직 실행 중…"))
            rows = compare_all(dae, drv, eng)

            q.put(("PROG", "5/5 결과 엑셀 저장 중…"))
            out = write_report(rows, drv)

            q.put(("OK", out))
        except Exception as e:
            q.put(("ERR", str(e)))

    def poll_queue():
        """작업이 끝났는지 주기적으로 확인 (UI 스레드)."""
        try:
            kind, payload = q.get_nowait()
        except queue.Empty:
            root.after(150, poll_queue)
            return

        # 진행 메시지면 UI만 갱신하고 계속 폴링
        if kind == "PROG":
            status.set(payload)
            root.after(150, poll_queue)
            return

        # OK / ERR이면 작업 종료 처리
        root.config(cursor="")
        btn_run.config(state="normal")
        btn_d1.config(state="normal")
        btn_d2.config(state="normal")
        btn_d3.config(state="normal")

        if kind == "OK":
            status.set("완료!")
            messagebox.showinfo("완료", f"검사 완료!\n결과 파일:\n{payload}")
        else:
            status.set("오류")
            messagebox.showerror("오류", f"실행 중 오류:\n{payload}")


    def run():
        if not daejang_path.get() or not drive_path.get() or not eng_path.get():
            messagebox.showerror("오류", "세 파일을 모두 선택하세요.")
            return

        # UI: 실행 중 표시 + 버튼 비활성화
        status.set("검사 중…")
        root.config(cursor="watch")
        btn_run.config(state="disabled")
        btn_d1.config(state="disabled")
        btn_d2.config(state="disabled")
        btn_d3.config(state="disabled")
        root.update_idletasks()

        # 백그라운드 작업 시작
        t = threading.Thread(
            target=worker,
            args=(daejang_path.get(), drive_path.get(), eng_path.get()),
            daemon=True
        )
        t.start()

        # 결과 폴링 시작
        root.after(150, poll_queue)


    frm = tk.Frame(root)
    frm.pack(pady=10)

    tk.Label(frm,text="대장 파일").grid(row=0,column=0)
    tk.Entry(frm,textvariable=daejang_path,width=45).grid(row=0,column=1)
    btn_d1 = tk.Button(frm, text="찾기", command=choose_daejang)
    btn_d1.grid(row=0, column=2)

    tk.Label(frm,text="운행일지").grid(row=1,column=0)
    tk.Entry(frm,textvariable=drive_path,width=45).grid(row=1,column=1)
    btn_d2 = tk.Button(frm, text="찾기", command=choose_drive)
    btn_d2.grid(row=1, column=2)

    tk.Label(frm,text="기술인력 파일").grid(row=2,column=0)
    tk.Entry(frm,textvariable=eng_path,width=45).grid(row=2,column=1)
    btn_d3 = tk.Button(frm, text="찾기", command=choose_eng)
    btn_d3.grid(row=2, column=2)

    btn_run = tk.Button(root, text="검사 시작", width=20, command=run)
    btn_run.pack(pady=15)
    tk.Label(root,textvariable=status,fg="blue").pack()

    root.mainloop()


###############################################################
# 실행부
###############################################################
if __name__ == "__main__":
    gui_start()