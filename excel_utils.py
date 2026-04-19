# -*- coding: utf-8 -*-
"""
엑셀 읽기 통합 (eco_input, eco_check, report_check 공통 파서)
"""
import os
import warnings
warnings.filterwarnings("ignore")  # 무조건 모든 경고 차단 (조건 없음)
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from format_utils import format_time, to_float1, to_float2

def find_sheet_by_candidates(wb, candidates: list):
    """후보 시트명 리스트에서 워크북에 존재하는 첫 번째 시트 반환."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        norm = name.replace(" ", "")
        for cand in candidates:
            if norm == cand.replace(" ", ""):
                return wb[name]
    return None

def _clean_text(s):
    """자잘한 특수문자 제거 (인력/장비용)"""
    if not s: return ""
    s = str(s).strip()
    if s.startswith(("×", "✕", ",")): return s[1:].strip()
    return s

def parse_measuring_record(excel_path: str, sample_no: str) -> dict:
    """
    대기측정기록부 및 입력 시트에서 모든 데이터를 긁어와 딕셔너리로 반환합니다.
    (eco_input, eco_check, report_check 모두 호환)
    """
    wb = load_workbook(excel_path, data_only=True)
    is_dust = "비산먼지" in os.path.basename(excel_path) or "비산" in os.path.basename(excel_path)

    # 1. 대기측정기록부 시트 파싱
    ws_name = "대기측정기록부" if "대기측정기록부" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    def cv(addr):
        v = ws[addr].value
        if isinstance(v, datetime): return v.strftime("%Y-%m-%d %H:%M:%S")
        return "" if v is None else str(v).strip()

    data = {
        "is_dust": is_dust,
        "기상": cv("D12"),
        "기온": cv("F12"),
        "습도": cv("H12"),
        "기압": cv("J12"),
        "풍향": cv("M12"),
        "풍속": to_float1(cv("O12")),
        "표준산소농도": to_float1(cv("D14")),
        "실측산소농도": to_float1(cv("G14")),
        "배출가스유량전": to_float1(cv("J14")),
        "배출가스유량후": to_float1(cv("N14")),
        "수분량": cv("D16"),
        "배출가스온도": cv("G16"),
        "배출가스유속": to_float2(cv("J16")),
        "엑셀시료번호": cv("P1")
    }

    # 날짜
    date_cell = ws["D37"].value
    if isinstance(date_cell, datetime):
        data["날짜"] = date_cell.strftime("%Y-%m-%d")
    else:
        data["날짜"] = str(date_cell).split()[0] if date_cell else ""

    # 채취시간 (비산먼지 vs 일반 행 위치 다름)
    if is_dust:
        data["채취시작"] = format_time(ws["F17"].value)
        data["채취끝"] = format_time(ws["I17"].value)
    else:
        data["채취시작"] = format_time(ws["F23"].value)
        data["채취끝"] = format_time(ws["I23"].value)

    # DT 생성 (eco_check 비교용)
    data["측정시작DT"] = f"{data['날짜']} {data['채취시작']}" if data["날짜"] and data["채취시작"] else ""
    data["측정종료DT"] = f"{data['날짜']} {data['채취끝']}" if data["날짜"] and data["채취끝"] else ""

    # 측정항목 파싱
    items = []
    for rng in ["B31:B36", "B49:B77", "B90:B118"]:
        min_c, min_r, max_c, max_r = range_boundaries(rng)
        for r in range(min_r, max_r + 1):
            v = ws.cell(row=r, column=min_c).value
            if v and str(v).strip():
                items.append(str(v).strip())
    data["측정항목"] = items

    # 2. '입력' 시트 파싱 (인력, 차량, 장비, 업소명)
    data["업소명"] = ""
    data["인력"] = []
    data["차량"] = []
    data["장비"] = []

    if "입력" in wb.sheetnames:
        ws_input = wb["입력"]
        # 업소명 (eco_check용)
        v_comp = ws_input["H7"].value
        data["업소명"] = "" if v_comp is None else str(v_comp).strip()

        # 시료번호에서 팀 번호 추출하여 해당 행(row) 파싱
        try:
            team_no = int(sample_no[7]) if len(sample_no) > 7 and sample_no[7].isdigit() else None
        except:
            team_no = None

        if team_no and 1 <= team_no <= 5:
            row = 8 + team_no
            # 인력 (AC, AD)
            staff = []
            for c in (29, 30):
                v = ws_input.cell(row=row, column=c).value
                if v and _clean_text(v): staff.append(_clean_text(v))
            data["인력"] = list(dict.fromkeys(staff))

            # 차량 (AE)
            car = ws_input.cell(row=row, column=31).value
            data["차량"] = [_clean_text(car)] if car and _clean_text(car) else []

            # 장비 (AH~AQ)
            eq = []
            for c in range(34, 44):
                v = ws_input.cell(row=row, column=c).value
                if v and _clean_text(v): eq.append(_clean_text(v))
            data["장비"] = list(dict.fromkeys(eq))

    wb.close()
    return data